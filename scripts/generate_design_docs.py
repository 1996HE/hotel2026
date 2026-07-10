from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/heniantong/Documents/codex/hotel-management")
OUT = ROOT / "docs/仕様書/民宿管理システム.xlsx"

CREATED_AT = datetime(2026, 7, 1, 0, 0)
UPDATED_AT = datetime(2026, 7, 8, 0, 0)
CREATED_TEXT = "2026-07-01"
UPDATED_TEXT = "2026-07-08"
AUTHOR = "何念童"
UPDATED_BY = "何念童"

thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

fill_title = PatternFill("solid", fgColor="D9EAF7")
fill_header = PatternFill("solid", fgColor="EAF3F8")
fill_body = PatternFill("solid", fgColor="FFFFFF")
fill_note = PatternFill("solid", fgColor="F7F9FC")

font_title = Font(name="Meiryo", size=12, bold=True)
font_header = Font(name="Meiryo", size=10, bold=True)
font_body = Font(name="Meiryo", size=10)
font_large = Font(name="Meiryo", size=16, bold=True)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)


def set_cell(cell, value, *, font=font_body, fill=fill_body, alignment=align_left):
    cell.value = value
    cell.font = font
    cell.fill = fill
    cell.border = border
    cell.alignment = alignment


def setup_sheet(ws, title: str, max_col: int):
    ws.sheet_view.zoomScale = 90
    for col_idx in range(1, max(max_col, 5) + 1):
        letter = ws.cell(1, col_idx).column_letter
        ws.column_dimensions[letter].width = 28 if col_idx == 2 else 18

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    set_cell(ws["A1"], f"{title}　作成・更新情報", font=font_title, fill=fill_title, alignment=align_center)

    headers = ["作成日時", "作成者", "更新日時", "更新者"]
    values = [CREATED_TEXT, AUTHOR, UPDATED_TEXT, UPDATED_BY]
    for idx, value in enumerate(headers, start=1):
        set_cell(ws.cell(2, idx), value, font=font_header, fill=fill_header, alignment=align_center)
    for idx, value in enumerate(values, start=1):
        set_cell(ws.cell(3, idx), value, fill=fill_note)

    set_cell(ws["A4"], "備考", font=font_header, fill=fill_header, alignment=align_center)
    set_cell(ws["B4"], "現行 Spring Boot + React + MyBatis 実装を基準に更新", fill=fill_note)
    set_cell(ws["C4"], "対象シート", font=font_header, fill=fill_header, alignment=align_center)
    set_cell(ws["D4"], title, fill=fill_note)

    for row in range(1, 5):
        ws.row_dimensions[row].height = 22
    ws.row_dimensions[5].height = 10


def add_table(ws, start_row: int, headers: list[str], rows: list[list[object]]):
    for col, header in enumerate(headers, start=1):
        set_cell(ws.cell(start_row, col), header, font=font_header, fill=fill_header, alignment=align_center)

    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c_idx, value in enumerate(row, start=1):
            set_cell(ws.cell(r_idx, c_idx), value)
        ws.row_dimensions[r_idx].height = 36


def add_table_after(ws, headers: list[str], rows: list[list[object]], *, gap: int = 2):
    add_table(ws, ws.max_row + gap, headers, rows)


def text_len(value: object) -> int:
    if value is None:
        return 0
    return len(str(value))


def write_cover(ws):
    setup_sheet(ws, "表紙", 5)
    ws.merge_cells("A6:E6")
    set_cell(ws["A6"], "民宿管理システム 仕様書", font=font_large, fill=fill_body, alignment=align_center)

    rows = [
        ["文書名", "民宿管理システム 仕様書", "", "仕様書構成", ""],
        ["プロジェクト名", "民宿管理システム開発", "", "シート", "内容"],
        ["対象業務", "民宿予約・客室・料金・清掃状態の社内管理", "", "概要", "システムの目的・範囲"],
        ["作成日", CREATED_TEXT, "", "業務要件", "利用者・管理者・共通要件"],
        ["版数", "1.0", "", "機能一覧", "開発対象機能と Java 処理一覧"],
        ["更新日", UPDATED_TEXT, "", "画面一覧", "React 画面一覧"],
        ["作成者", AUTHOR, "", "DB設計", "テーブル・Domain・DAO 対応"],
        ["更新者", UPDATED_BY, "", "API一覧", "React API と DTO"],
        ["対象URL", "http://localhost:8001/jukai-internal", "", "テスト仕様", "単体テスト仕様のみ"],
        ["詳細設計", "コード行別処理仕様", "", "コード行別処理仕様", "Java ファイル・行番号別の処理追跡表"],
    ]
    for r_idx, row in enumerate(rows, start=8):
        for c_idx, value in enumerate(row, start=1):
            set_cell(ws.cell(r_idx, c_idx), value, font=font_header if c_idx in (1, 4) else font_body,
                     fill=fill_header if c_idx in (1, 4) else fill_body)
    add_table_after(ws, ["優先順", "工程区分", "対象シート", "記載目的", "確認観点"], [
        [1, "文書管理", "表紙 / 変更履歴", "文書の版、作成者、更新者、変更内容を確認する。", "提出物としての識別性があること。"],
        [2, "全体方針", "概要", "システム目的、対象範囲、基本仕様、アーキテクチャを確認する。", "システム全体像が先に把握できること。"],
        [3, "要件定義", "業務要件", "利用者別要件、受入条件、業務規則を確認する。", "要件から実装・テストへ追跡できること。"],
        [4, "基本設計", "機能一覧 / 画面一覧", "機能、画面、Java 責務、画面遷移を確認する。", "業務機能と画面/APIの対応が分かること。"],
        [5, "外部・データ設計", "API一覧 / DB設計", "API、DTO、状態値、DB、Mapper、Domain 対応を確認する。", "外部入出力と永続化仕様が分かること。"],
        [6, "詳細設計", "コード行別処理仕様", "実際の Java 行番号に沿って処理内容を確認する。", "コード行と仕様行が対応していること。"],
        [7, "品質設計", "非機能要件 / テスト仕様", "セキュリティ、性能、保守性、テスト要望、期待結果を確認する。", "品質要件と検証観点が揃っていること。"],
        [8, "管理情報", "スケジュール / 参考資料", "工程、成果物、参照コード、関連資料を確認する。", "提出・保守時の参照先が明確であること。"],
    ])


def write_history(ws):
    setup_sheet(ws, "変更履歴", 5)
    add_table(ws, 6, ["版数", "日付", "変更内容", "作成者", "承認者"], [
        ["1.0", CREATED_TEXT, "初版作成", AUTHOR, ""],
        ["1.1", "2026-07-06", "React 化、Java 分層仕様、DTO/API/DAO/Service 仕様を追加", UPDATED_BY, ""],
        ["1.2", "2026-07-06", "参考サンプル形式に合わせ、テスト仕様は単体テストまでに整理", UPDATED_BY, ""],
        ["1.3", UPDATED_TEXT, "Java 処理ロジックを Controller / Service / Mapper / SQL 単位に分解し、詳細仕様を補強", UPDATED_BY, ""],
        ["1.4", UPDATED_TEXT, "各シートの優先順をプロジェクト標準工程に合わせて整理し、コード行別処理仕様を追加", UPDATED_BY, ""],
        ["1.5", UPDATED_TEXT, "テストケース番号、テスト注解、System.out.print 文字列結果、システム日付基準の不正利用防止テストを追加", UPDATED_BY, ""],
        ["1.6", UPDATED_TEXT, "予約番号 R000001 形式、取消済み予約削除、客室復元・完全削除、清掃更新制限、前端ページング表示を反映", UPDATED_BY, ""],
        ["1.7", UPDATED_TEXT, "チェックアウト済み予約削除、予約番号 DB 制約・索引、予約サンプルデータ追加を反映", UPDATED_BY, ""],
    ])


def write_overview(ws):
    setup_sheet(ws, "概要", 4)
    add_table(ws, 6, ["優先順", "項目", "内容", "確認観点"], [
        [1, "目的", "社内担当者が民宿の予約登録、同行者管理、客室管理、料金ルール管理を Web 画面で効率的に行えるようにする。", "システムが解決する業務目的を最初に確認する。"],
        [2, "対象利用者", "予約管理担当者、施設管理担当者、社内運用担当者", "誰が利用する業務システムか確認する。"],
        [3, "対象範囲", "ダッシュボード、予約管理、客室管理、料金管理、React API、単体テスト", "開発・提出対象の範囲を確認する。"],
        [4, "対象外", "外部決済連携、会員向け公開予約サイト、スマートロック連携、清掃業者自動手配", "今回含めない範囲を確認する。"],
        [5, "背景", "民宿運用では、予約、客室状態、清掃状態、期間別料金を一元管理する必要がある。", "業務上の必要性を確認する。"],
        [6, "構成", "Spring Boot 3.3.5 / React / MyBatis / PostgreSQL。Controller -> Service -> Mapper -> DB の層構成。", "技術構成と分層を確認する。"],
        [7, "文書方針", "参考サンプルに合わせ、各シート先頭に作成・更新情報を付与し、本文は第6行から表形式で記載する。", "提出形式を確認する。"],
    ])
    add_table_after(ws, ["基本仕様ID", "分類", "仕様項目", "仕様内容"], [
        ["BASE-001", "アーキテクチャ", "画面方式", "Spring MVC Controller は /dashboard、/reservations、/rooms、/prices で React shell の app.html を返し、実データは /api 配下の JSON API から取得する。"],
        ["BASE-002", "アーキテクチャ", "業務処理方式", "Controller は入力受領とレスポンス整形を担当し、予約・客室・料金の業務判断は Service に集約する。"],
        ["BASE-003", "アーキテクチャ", "永続化方式", "Service は MyBatis Mapper interface を呼び出し、SQL 詳細は src/main/resources/mapper の XML に定義する。"],
        ["BASE-004", "トランザクション", "更新系処理", "予約登録、予約状態更新、客室登録、料金登録・削除は @Transactional で一貫性を確保する。"],
        ["BASE-005", "トランザクション", "参照系処理", "一覧取得、件数取得、単票取得は @Transactional(readOnly=true) として定義する。"],
        ["BASE-006", "例外処理", "JSON API", "Service の IllegalArgumentException は ApiController の @ExceptionHandler で HTTP 400 の ErrorResponse に変換する。"],
        ["BASE-007", "例外処理", "MVC POST", "HTML Controller の POST は IllegalArgumentException を flash attribute の error に変換し、元画面へ redirect する。"],
        ["BASE-008", "状態同期", "予約と客室", "予約成立・取消・チェックアウト・予約状態変更時は、予約状態と客室の宿泊状態・清掃状態を同じ業務処理内で同期する。"],
        ["BASE-009", "料金計算", "日別積算", "宿泊日ごとに料金ルールを検索し、該当ルールがない日は客室基本単価を利用して人数分を合算する。"],
        ["BASE-010", "セキュリティ", "公開範囲", "実画面、API、静的 JS/CSS のみ permitAll とし、ソース・設定・SQL・.git 等は denyAll とする。"],
    ])


def write_business_requirements(ws):
    setup_sheet(ws, "業務要件", 6)
    add_table(ws, 6, ["要件ID", "分類", "要件名", "要件内容", "優先度", "備考"], [
        ["BR-001", "予約", "予約登録", "空室かつ清掃済みの客室に対して、宿泊者・日程・人数・同行者を登録できる。", "高", "ReservationService.create"],
        ["BR-002", "予約", "重複予約防止", "同一客室で宿泊期間が重なる予約を登録できない。", "高", "countOverlapping"],
        ["BR-003", "予約", "予約番号発番", "予約登録時に R + 6桁以上の連番予約番号を発番する。", "高", "reservation_no_seq"],
        ["BR-004", "予約", "同行者管理", "宿泊人数が2名以上の場合、代表者以外の同行者情報を保存する。", "高", "reservation_guests"],
        ["BR-005", "予約", "支払更新", "予約一覧から未払い・支払済を更新できる。", "中", "paymentStatus"],
        ["BR-006", "予約", "取消", "予約を取消済みにし、客室を空室・清掃済みに戻せる。", "高", "cancel"],
        ["BR-006A", "予約", "チェックアウト済み予約削除", "チェックアウト済み予約を一覧から完全削除できる。", "中", "deleteCheckedOut"],
        ["BR-007", "客室", "客室登録", "客室番号、名称、種別、定員、基本料金を登録できる。", "高", "RoomService.create"],
        ["BR-008", "客室", "状態更新", "宿泊状態と清掃状態を同時に更新できる。", "高", "RoomService.updateStatuses"],
        ["BR-009", "客室", "論理削除", "客室は物理削除せず無効化し、過去予約との整合性を保つ。", "中", "active=false"],
        ["BR-010", "料金", "期間別料金", "客室ごとに期間別の1人料金を登録できる。", "高", "RoomPriceRuleService.create"],
        ["BR-011", "料金", "料金重複防止", "同一客室で期間が重なる料金ルールを拒否する。", "高", "countOverlapping"],
        ["BR-012", "自動同期", "チェックアウト同期", "宿泊終了日を過ぎた予約をチェックアウト済みにし、客室を清掃待ちにする。", "中", "syncDueCheckouts"],
    ])
    add_table_after(ws, ["要件定義ID", "利用者", "業務", "要求", "受入条件", "関連機能"], [
        ["REQ-001", "予約管理担当者", "予約登録", "電話・メールを保持しない予約も登録できること。", "電話非保持とメール非保持を同時選択した場合、DB には NULL が保存される。", "F-003"],
        ["REQ-002", "予約管理担当者", "予約登録", "複数名宿泊では代表者以外の同行者を登録できること。", "宿泊人数が2名以上の場合、人数-1件分の同行者名が必須になる。", "F-003"],
        ["REQ-003", "予約管理担当者", "予約登録", "過去日や宿泊期間が逆転した予約を防止できること。", "DB CURRENT_DATE より前のチェックイン日、または checkIn >= checkOut は登録できない。", "F-003"],
        ["REQ-004", "予約管理担当者", "予約登録", "同一客室の重複宿泊予約を防止できること。", "booked 状態の既存予約と期間交差する場合、登録できない。", "F-003"],
        ["REQ-005", "予約管理担当者", "支払管理", "予約一覧から未払い・支払済みを更新できること。", "payment_status は unpaid / paid のみ受け付ける。", "F-005"],
        ["REQ-006", "予約管理担当者", "取消管理", "予約取消時に客室を再販売可能状態へ戻せること。", "取消後、予約は cancelled、客室は vacant / cleaned になる。", "F-004"],
        ["REQ-006A", "予約管理担当者", "取消管理", "取消済み予約を一覧から完全削除できること。", "reservation_status=cancelled の予約だけ deleteCancelled で削除できる。", "F-004A"],
        ["REQ-006B", "予約管理担当者", "チェックアウト管理", "チェックアウト済み予約を一覧から完全削除できること。", "reservation_status=checked_out の予約だけ deleteCheckedOut で削除できる。", "F-004B"],
        ["REQ-007", "施設管理担当者", "客室管理", "客室を論理削除でき、過去予約参照を維持できること。", "rooms.active=false になり、reservations.room_id は維持される。", "F-008"],
        ["REQ-008", "施設管理担当者", "客室管理", "削除済みと同じ部屋番号を再登録する場合は再有効化できること。", "同一 room_number の inactive 客室は reactivate で更新される。", "F-008"],
        ["REQ-008A", "施設管理担当者", "客室管理", "削除済み客室を復元でき、予約履歴がない削除済み客室だけ完全削除できること。", "restore は inactive のみ、deletePermanently は inactive かつ予約履歴0件のみ実行できる。", "F-008A"],
        ["REQ-009", "施設管理担当者", "清掃管理", "チェックアウト後の清掃状態を管理できること。", "checked_out 予約の客室を needs_cleaning / cleaned に更新できる。", "F-006"],
        ["REQ-010", "施設管理担当者", "料金管理", "客室ごとに期間別料金を登録できること。", "同一客室で期間重複する active 料金ルールは登録できない。", "F-011"],
        ["REQ-011", "社内運用担当者", "ダッシュボード", "現在の客室・予約状況を確認できること。", "画面表示前に期限到来チェックアウトを同期した集計値を表示する。", "F-001"],
        ["REQ-012", "社内運用担当者", "セキュリティ", "社内画面以外のソース・設定ファイルを公開しないこと。", "SOURCE_LIKE_PATHS に該当する URL は denyAll になる。", "F-013"],
    ])
    add_table_after(ws, ["規則ID", "対象", "処理タイミング", "判定条件", "正常時", "異常時"], [
        ["RULE-001", "予約登録", "ReservationService.create 開始時", "roomId、宿泊日、宿泊者名、宿泊人数が入力済み", "次の検証へ進む", "業務エラーメッセージを返す"],
        ["RULE-002", "予約登録", "日付検証", "checkInDate >= DB CURRENT_DATE かつ checkInDate < checkOutDate", "宿泊期間として採用", "過去日または逆転日付を拒否"],
        ["RULE-003", "予約登録", "代表者連絡先検証", "フリガナは全角カタカナ、電話は000-0000-0000、メールは一般的なメール形式", "予約情報に保持", "形式不正を拒否"],
        ["RULE-004", "予約登録", "連絡先非保持", "noPhoneInfo と noEmailInfo がともに true", "電話番号・メールを NULL 保存", "該当なし"],
        ["RULE-005", "予約登録", "客室検証", "active=true、occupancy_status=vacant、cleaning_status=cleaned、人数 <= 定員", "予約対象として確定", "利用不可客室として拒否"],
        ["RULE-006", "予約登録", "重複予約検証", "同一 room_id で booked 予約の期間が交差しない", "予約番号発番へ進む", "指定期間予約済みとして拒否"],
        ["RULE-007", "予約登録", "同行者検証", "guestCount - 1 件分の同行者名が入力済み", "reservation_guests へ分割保存", "同行者情報不足を拒否"],
        ["RULE-008", "予約登録", "料金計算", "宿泊日ごとに最優先料金ルールを検索。未設定日は客室基本単価", "日別単価 * 人数を total_amount に合算", "該当なし"],
        ["RULE-009", "予約状態更新", "booked へ変更", "reservationStatus=booked", "客室を reserved / cleaned に同期", "許可値以外は拒否"],
        ["RULE-010", "予約状態更新", "checked_out へ変更", "reservationStatus=checked_out", "客室を vacant / needs_cleaning に同期", "許可値以外は拒否"],
        ["RULE-011", "予約取消", "cancel 実行時", "対象予約が存在する", "予約を cancelled、客室を vacant / cleaned に同期", "予約なしを拒否"],
        ["RULE-012", "客室登録", "RoomService.create", "部屋番号・部屋名・定員・部屋タイプ・状態が妥当", "新規登録または削除済み客室を再有効化", "重複中の有効客室を拒否"],
        ["RULE-013", "客室状態更新", "RoomService.updateStatuses", "occupancyStatus と cleaningStatus が許可値", "rooms を同時更新", "不正値または存在なしを拒否"],
        ["RULE-014", "料金登録", "RoomPriceRuleService.create", "客室、名称、期間、金額、優先度、有効状態が妥当", "料金ルールを登録", "期間不正・負数・重複期間を拒否"],
        ["RULE-015", "料金削除", "deleteByIds", "ID 配列から NULL と重複を除外後、1件以上残る", "対象 ID を一括削除", "未選択または対象なしを拒否"],
    ])


def write_functions(ws):
    setup_sheet(ws, "機能一覧", 9)
    add_table(ws, 6, ["機能ID", "大分類", "機能名", "概要", "入力", "出力", "権限", "優先度", "関連Java"], [
        ["F-001", "画面", "ダッシュボード表示", "客室数、空室数、予約中件数、直近予約を表示する。", "なし", "集計・予約一覧", "社内", "高", "DashboardController / ApiController"],
        ["F-002", "予約", "予約一覧取得", "予約、取消、チェックアウト一覧と予約可能客室を取得する。", "なし", "ReservationsResponse", "社内", "高", "ApiController.reservations"],
        ["F-003", "予約", "予約登録", "入力検証、客室ロック、重複確認、料金計算、同行者保存を行う。", "ReservationCreateRequest", "登録完了", "社内", "高", "ReservationService.create"],
        ["F-004", "予約", "予約取消", "予約状態を取消済みにし、他予約がなければ客室状態を空室・清掃済みに戻す。", "予約ID", "取消完了", "社内", "高", "ReservationService.cancel"],
        ["F-004A", "予約", "取消済み予約削除", "取消済み予約のみを予約一覧から完全削除する。", "予約ID", "削除完了", "社内", "中", "ReservationService.deleteCancelled"],
        ["F-004B", "予約", "チェックアウト済み予約削除", "チェックアウト済み予約のみをチェックアウト一覧から完全削除する。", "予約ID", "削除完了", "社内", "中", "ReservationService.deleteCheckedOut"],
        ["F-005", "予約", "支払状態更新", "支払状態を unpaid / paid の範囲で更新する。", "PaymentRequest", "更新完了", "社内", "中", "updatePaymentStatus"],
        ["F-006", "予約", "清掃状態更新", "チェックアウト後の客室清掃状態を更新する。", "CleaningRequest", "更新完了", "社内", "中", "updateCheckoutCleaningStatus"],
        ["F-007", "客室", "客室一覧取得", "有効、削除済み、予約可能客室を取得する。", "なし", "RoomsResponse", "社内", "高", "RoomService"],
        ["F-008", "客室", "客室登録", "入力検証後、新規登録または論理削除済み客室を再有効化する。", "Room", "登録完了", "社内", "高", "RoomService.create"],
        ["F-008A", "客室", "削除済み客室復元・完全削除", "論理削除済み客室を有効化し、予約履歴がない削除済み客室は物理削除できる。", "客室ID", "復元/完全削除完了", "社内", "中", "RoomService.restore / deletePermanently"],
        ["F-009", "客室", "客室状態更新", "宿泊状態と清掃状態を許可値で検証し更新する。", "RoomStatusRequest", "更新完了", "社内", "高", "RoomService.updateStatuses"],
        ["F-010", "料金", "料金一覧取得", "料金ルールと有効客室を取得する。", "なし", "PricesResponse", "社内", "中", "RoomPriceRuleService"],
        ["F-011", "料金", "料金ルール登録", "期間、価格、重複を検証して登録する。", "RoomPriceRule", "登録完了", "社内", "高", "RoomPriceRuleService.create"],
        ["F-012", "料金", "料金削除", "単一または選択 ID 一覧で料金ルールを削除する。", "id / ids", "削除完了", "社内", "中", "delete / deleteByIds"],
        ["F-013", "設定", "セキュリティ制御", "公開 URL、拒否 URL、CSRF、CSP、セキュリティヘッダーを設定する。", "HTTP request", "HTTP response", "社内", "高", "SecurityConfig"],
    ])
    add_table_after(ws, ["処理ID", "層", "Java / XML", "メソッド / SQL", "処理分解", "主な入力", "主な出力", "トランザクション", "例外・備考"], [
        ["LOGIC-000", "Application", "MinshukuManagementApplication", "main", "Spring Boot アプリケーションを起動する。業務ロジックは持たず、コンポーネントスキャンの起点になる。", "起動引数", "ApplicationContext", "-", "context-path は application.yml で /jukai-internal"],
        ["LOGIC-000A", "Config", "SecurityConfig", "securityFilterChain", "公開パス、拒否パス、CSRF、CSP、各種セキュリティヘッダーを HttpSecurity に設定する。", "HTTP request", "SecurityFilterChain", "-", "認証画面は無効。社内画面前提で permitAll と denyAll を明示"],
        ["LOGIC-000B", "Controller", "DashboardController", "dashboard", "/ と /dashboard への GET で React shell の app を返す。", "GET request", "templates/app.html", "-", "集計データは /api/dashboard が担当"],
        ["LOGIC-000C", "Controller", "ReservationController", "reservations / create / update", "MVC 互換の画面表示・POST 入口。GET は app を返し、POST は Service 実行後に /reservations へ redirect する。", "ModelAttribute / RequestParam", "redirect / flash", "Service 側", "React 化後も旧フォーム送信に近い入口を保持"],
        ["LOGIC-000D", "Controller", "RoomController", "rooms / create / updateStatuses / delete", "客室画面の app 返却、客室登録、状態更新、論理削除を Service に委譲する。", "Room / statuses / id", "redirect / flash", "Service 側", "RuntimeException は登録失敗メッセージへ変換"],
        ["LOGIC-000E", "Controller", "PriceRuleController", "prices / create / delete / deleteSelected", "料金画面の app 返却、単一削除、旧URL互換削除、一括削除を Service に委譲する。", "RoomPriceRule / id / ids", "redirect / flash", "Service 側", "入力エラーは flash error"],
        ["LOGIC-001", "Controller", "ApiController", "dashboard", "期限到来チェックアウトを同期し、客室総数・空室数・予約中件数・直近予約1ページ目を組み立てる。", "GET /api/dashboard", "DashboardResponse", "Service 側", "表示前に syncDueCheckouts を必ず実行"],
        ["LOGIC-002", "Controller", "ApiController", "reservations", "予約中・取消済み・チェックアウト済みの各一覧、予約可能客室、業務日付を返す。", "GET /api/reservations", "ReservationsResponse", "Service 側", "各一覧は PAGE_SIZE=5 の初期ページ"],
        ["LOGIC-003", "Controller", "ApiController", "createReservation", "JSON DTO から予約本体と同行者配列を取り出し、Service へ委譲する。", "ReservationCreateRequest", "MessageResponse", "Service 側", "noPhoneInfo と noEmailInfo が両方 true の場合のみ連絡先非保持"],
        ["LOGIC-003A", "Controller", "ApiController", "deleteCancelledReservation", "取消済み予約削除 API を受け取り、Service の deleteCancelled へ委譲する。", "POST /api/reservations/{id}/delete", "MessageResponse", "Service 側", "取消済み予約のみ削除"],
        ["LOGIC-004", "Service", "ReservationService", "create", "入力検証、同行者検証、客室 FOR UPDATE、重複予約確認、予約番号発番、初期値補完、料金計算、本体保存、同行者保存、客室 reserved 化を1トランザクションで行う。", "Reservation / companion arrays", "DB更新", "REQUIRED", "IllegalArgumentException を API で 400 JSON 化"],
        ["LOGIC-005", "Service", "ReservationService", "validateReservation", "必須、過去日、宿泊期間、代表者名、フリガナ、連絡先形式、人数1-10名を検証する。", "Reservation / noContactInfo", "検証通過", "create 内", "連絡先非保持時は電話・メールを NULL にする"],
        ["LOGIC-006", "Service", "ReservationService", "validateRoomForReservation", "営業中、空室、清掃済み、定員内を検証する。", "Reservation / Room", "予約可能客室", "create 内", "Room は findByIdForUpdate で取得"],
        ["LOGIC-007", "Service", "ReservationService", "validateNoOverlappingReservation", "booked 状態の同一客室予約について check_in < 新check_out かつ check_out > 新check_in を重複とする。", "roomId / dates", "重複なし", "create 内", "重複時は登録拒否"],
        ["LOGIC-008", "Service", "ReservationService", "calculateTotalAmount", "チェックイン日からチェックアウト前日まで日別に単価を取得し、人数分を合算する。", "Reservation / Room", "totalAmount", "create 内", "findBestRule がない日は basePricePerPerson"],
        ["LOGIC-009", "Service", "ReservationService", "syncDueCheckouts", "booked かつ check_out_date <= CURRENT_DATE の予約を checked_out にし、客室を vacant / needs_cleaning にする。", "DB current date", "予約・客室更新", "REQUIRED", "dashboard / reservations 表示前に呼び出し"],
        ["LOGIC-010", "Service", "ReservationService", "updateReservationStatus", "予約状態許可値を検証し、予約本体更新後に booked / checked_out の客室状態を同期する。他の booked 予約が同室に残る場合は客室を reserved / cleaned に維持する。", "id / reservationStatus", "DB更新", "REQUIRED", "cancelled は cancel 専用 API でも扱う"],
        ["LOGIC-011", "Service", "ReservationService", "updateCheckoutCleaningStatus", "予約を取得し、checked_out 予約だけ清掃状態許可値を検証して客室を vacant のまま更新する。", "id / cleaningStatus", "DB更新", "REQUIRED", "booked / cancelled は清掃更新不可"],
        ["LOGIC-012", "Service", "RoomService", "create", "部屋番号、名称、定員、基本料金、種別、風呂有無、宿泊状態、清掃状態、有効フラグを補完・検証し、重複判定後に insert または reactivate する。", "Room", "DB更新", "REQUIRED", "有効客室の同番号は拒否"],
        ["LOGIC-012A", "Service", "RoomService", "restore / deletePermanently", "削除済み客室の復元と、予約履歴がない削除済み客室の完全削除を行う。", "roomId", "DB更新", "REQUIRED", "active 客室や予約履歴あり客室は完全削除不可"],
        ["LOGIC-013", "Service", "RoomService", "updateStatuses", "宿泊状態 vacant/reserved/occupied、清掃状態 cleaned/needs_cleaning を検証し、客室状態を同時更新する。", "id / statuses", "DB更新", "REQUIRED", "更新0件は部屋なし"],
        ["LOGIC-014", "Service", "RoomPriceRuleService", "create", "対象客室、名称、期間、金額、優先度、有効フラグを検証し、同一客室の期間重複がない場合だけ登録する。", "RoomPriceRule", "DB更新", "REQUIRED", "優先度未指定は10、有効未指定はtrue"],
        ["LOGIC-015", "Service", "RoomPriceRuleService", "deleteByIds", "ID 一覧から null と重複を除去し、残った ID を一括削除する。", "List<Integer>", "DB更新", "REQUIRED", "未選択、削除0件は業務エラー"],
        ["LOGIC-016", "Mapper", "RoomMapper.xml", "findBookable", "active=true、vacant、cleaned の客室を予約フォーム候補として取得する。", "なし", "List<Room>", "readOnly", "部屋番号順"],
        ["LOGIC-017", "Mapper", "ReservationMapper.xml", "CompanionSummary", "reservation_guests を予約単位で string_agg し、同行者名・性別・年齢・電話を複数行文字列にする。", "reservation_id", "companionSummary", "readOnly", "未入力は既定文言で補完"],
        ["LOGIC-018", "Mapper", "RoomPriceRuleMapper.xml", "findBestRule", "対象客室・対象日・active=true に一致する料金ルールを priority 昇順、price 降順で1件取得する。", "roomId / stayDate", "RoomPriceRule", "readOnly", "料金計算で日別に呼び出す"],
    ])
    add_table_after(ws, ["Java分類ID", "パッケージ", "クラス", "責務", "主要メンバー / メソッド", "関連DB/API", "基本仕様上の位置付け", "備考", ""], [
        ["JAVA-001", "root", "MinshukuManagementApplication", "Spring Boot 起動クラス。業務処理は保持しない。", "main", "-", "アプリケーション起点", ""],
        ["JAVA-002", "config", "SecurityConfig", "社内画面の HTTP セキュリティ設定を集約する。", "SOURCE_LIKE_PATHS / PUBLIC_PAGE_PATHS / CONTENT_SECURITY_POLICY / securityFilterChain", "全URL", "非機能・セキュリティ基本仕様", "CSRF token は CookieCsrfTokenRepository"],
        ["JAVA-003", "controller", "DashboardController", "ダッシュボード画面の React shell を返す。", "dashboard", "/, /dashboard", "画面ルーティング", ""],
        ["JAVA-004", "controller", "ReservationController", "予約管理の MVC 互換入口を提供する。", "create / updatePayment / cancel / updateStatus / updateCleaning", "/reservations", "画面ルーティング・旧POST互換", ""],
        ["JAVA-005", "controller", "RoomController", "客室管理の MVC 互換入口を提供する。", "create / updateStatuses / delete", "/rooms", "画面ルーティング・旧POST互換", ""],
        ["JAVA-006", "controller", "PriceRuleController", "料金管理の MVC 互換入口を提供する。", "create / delete / deleteSelected", "/prices", "画面ルーティング・旧POST互換", "旧 /prices/{id} 削除URLも受け付ける"],
        ["JAVA-007", "controller", "ApiController", "React フロントエンド用 JSON API と DTO record を提供する。", "dashboard / rooms / prices / reservations / ExceptionHandler", "/api/**", "API 基本仕様", "PAGE_SIZE=5"],
        ["JAVA-008", "service", "ReservationService", "予約登録、状態更新、料金計算、同行者保存、チェックアウト同期を扱う。", "create / syncDueCheckouts / calculateTotalAmount", "reservations / reservation_guests / rooms / room_price_rules", "予約業務詳細仕様", "予約業務の中核"],
        ["JAVA-009", "service", "RoomService", "客室登録、再有効化、状態更新、論理削除、件数集計を扱う。", "create / updateStatuses / delete / countAll / countVacant", "rooms", "客室業務詳細仕様", ""],
        ["JAVA-010", "service", "RoomPriceRuleService", "料金ルール登録、期間重複チェック、削除を扱う。", "create / delete / deleteByIds", "room_price_rules", "料金業務詳細仕様", ""],
        ["JAVA-011", "mapper", "ReservationMapper", "予約検索・登録・状態更新・件数取得の Mapper interface。", "findRecentPage / insert / countOverlapping / markCheckedOut / cancel", "ReservationMapper.xml", "DAO 基本仕様", ""],
        ["JAVA-012", "mapper", "ReservationGuestMapper", "同行者明細の保存 Mapper interface。", "insert", "ReservationGuestMapper.xml", "DAO 基本仕様", ""],
        ["JAVA-013", "mapper", "RoomMapper", "客室検索・登録・状態更新の Mapper interface。", "findBookable / findByIdForUpdate / reactivate / updateStatuses", "RoomMapper.xml", "DAO 基本仕様", ""],
        ["JAVA-014", "mapper", "RoomPriceRuleMapper", "料金ルール検索・登録・削除の Mapper interface。", "findBestRule / countOverlapping / deleteByIds", "RoomPriceRuleMapper.xml", "DAO 基本仕様", ""],
        ["JAVA-015", "domain", "Reservation", "予約本体、表示用客室情報、代表者、状態、料金、同行者要約を保持する。", "getReservationStatusLabel / getPaymentStatusLabel", "reservations", "ドメイン基本仕様", "ラベル変換を保持"],
        ["JAVA-016", "domain", "ReservationGuest", "予約に紐づく同行者明細を保持する。", "reservationId / guestName / guestKana / guestGender / guestAge / guestPhone", "reservation_guests", "ドメイン基本仕様", ""],
        ["JAVA-017", "domain", "Room", "客室マスタ、販売条件、宿泊状態、清掃状態、論理削除情報を保持する。", "roomNumber / capacity / basePricePerPerson / occupancyStatus / cleaningStatus / active", "rooms", "ドメイン基本仕様", ""],
        ["JAVA-018", "domain", "RoomPriceRule", "客室別・期間別料金ルール、優先度、有効状態を保持する。", "roomId / startDate / endDate / pricePerPerson / priority / active", "room_price_rules", "ドメイン基本仕様", ""],
    ])


def write_code_trace(ws):
    setup_sheet(ws, "コード行別処理仕様", 9)
    add_table(ws, 6, ["Trace ID", "実行順", "ファイル", "行番号", "メソッド", "コード行 / 条件", "処理内容", "入出力・状態変化", "関連仕様"], [
        ["TRACE-001", 1, "MinshukuManagementApplication.java", "13-14", "main", "SpringApplication.run(...)", "Spring Boot アプリケーションを起動し、Controller / Service / Mapper / Config を DI 対象として読み込む。", "ApplicationContext 起動", "BASE-001"],
        ["TRACE-002", 2, "SecurityConfig.java", "22-39", "class init", "SOURCE_LIKE_PATHS", "ソース・設定・SQL・ビルドファイル系 URL の拒否対象を定義する。", "denyAll 候補パス", "BNF-001"],
        ["TRACE-003", 3, "SecurityConfig.java", "41-55", "class init", "PUBLIC_PAGE_PATHS", "業務画面、API、静的 JS/CSS の公開対象 URL を定義する。", "permitAll 候補パス", "BNF-002"],
        ["TRACE-004", 4, "SecurityConfig.java", "57-67", "class init", "CONTENT_SECURITY_POLICY", "外部送信・外部埋め込みを抑止する CSP 文字列を組み立てる。", "CSP header 値", "BNF-004"],
        ["TRACE-005", 5, "SecurityConfig.java", "71", "sourceLikeMatchers", "new RequestMatcher[SOURCE_LIKE_PATHS.length]", "拒否パス配列と同じ長さの matcher 配列を作成する。", "RequestMatcher[]", "BNF-001"],
        ["TRACE-006", 6, "SecurityConfig.java", "72-74", "sourceLikeMatchers", "for / AntPathRequestMatcher.antMatcher", "SOURCE_LIKE_PATHS を AntPathRequestMatcher に変換する。", "denyAll 用 matcher", "BNF-001"],
        ["TRACE-007", 7, "SecurityConfig.java", "85", "securityFilterChain", "requestMatchers(sourceLikeMatchers()).denyAll()", "ソース類似パスへのアクセスを拒否する。", "HTTP 403 対象", "BNF-001"],
        ["TRACE-008", 8, "SecurityConfig.java", "86", "securityFilterChain", "requestMatchers(PUBLIC_PAGE_PATHS).permitAll()", "業務画面・API・静的リソースを許可する。", "公開 URL", "BNF-002"],
        ["TRACE-009", 9, "SecurityConfig.java", "87", "securityFilterChain", "anyRequest().denyAll()", "定義外 URL をすべて拒否する。", "未定義 URL 拒否", "BNF-001"],
        ["TRACE-010", 10, "SecurityConfig.java", "88-90", "securityFilterChain", "httpBasic/formLogin/logout disable", "標準認証画面、Basic 認証、logout を無効化する。", "社内画面前提の認証なし構成", "BASE-010"],
        ["TRACE-011", 11, "SecurityConfig.java", "91", "securityFilterChain", "CookieCsrfTokenRepository.withHttpOnlyFalse()", "React が CSRF token を読める Cookie CSRF 方式を設定する。", "CSRF token cookie", "BNF-003"],
        ["TRACE-012", 12, "SecurityConfig.java", "93-102", "securityFilterChain", "headers(...)", "CSP、Referrer、Robots、Permissions、FrameOptions、ContentTypeOptions を付与する。", "Security headers", "BNF-004 / BNF-005"],
        ["TRACE-013", 13, "SecurityConfig.java", "103", "securityFilterChain", "return http.build()", "SecurityFilterChain を Spring Security へ返す。", "SecurityFilterChain bean", "BASE-010"],
        ["TRACE-014", 14, "DashboardController.java", "14-16", "dashboard", "return \"app\"", "/ または /dashboard への GET で React shell を返す。", "viewName=app", "SCR-001"],
        ["TRACE-015", 15, "ApiController.java", "47", "dashboard", "reservationService.syncDueCheckouts()", "ダッシュボード集計前に期限到来チェックアウトを同期する。", "予約・客室状態更新", "LOGIC-001"],
        ["TRACE-016", 16, "ApiController.java", "48-52", "dashboard", "new DashboardResponse(...)", "客室数、空室数、予約中件数、直近予約ページをレスポンスへ詰める。", "DashboardResponse", "DTO-001"],
        ["TRACE-017", 17, "ApiController.java", "57", "rooms", "new RoomsResponse(...)", "有効客室、削除済み客室、予約可能客室を返す。", "RoomsResponse", "DTO-004"],
        ["TRACE-018", 18, "ApiController.java", "62", "createRoom", "roomService.create(room)", "JSON の客室情報を客室登録 Service へ渡す。", "rooms insert/reactivate", "API-003"],
        ["TRACE-019", 19, "ApiController.java", "63", "createRoom", "new MessageResponse(...)", "客室登録成功メッセージを返す。", "message JSON", "DTO-015"],
        ["TRACE-020", 20, "ApiController.java", "68", "updateRoomStatuses", "roomService.updateStatuses(...)", "客室の宿泊状態と清掃状態を同時更新する。", "rooms status update", "API-004"],
        ["TRACE-021", 21, "ApiController.java", "74", "deleteRoom", "roomService.delete(id)", "客室を論理削除する。", "rooms.active=false", "API-005"],
        ["TRACE-022", 22, "ApiController.java", "80", "prices", "new PricesResponse(...)", "料金ルール一覧と登録対象客室を返す。", "PricesResponse", "DTO-005"],
        ["TRACE-023", 23, "ApiController.java", "85", "createPriceRule", "priceRuleService.create(rule)", "料金ルール登録 Service を実行する。", "room_price_rules insert", "API-007"],
        ["TRACE-024", 24, "ApiController.java", "91", "deletePriceRule", "priceRuleService.delete(id)", "料金ルールを1件削除する。", "room_price_rules delete", "API-008"],
        ["TRACE-025", 25, "ApiController.java", "97", "deleteSelectedPriceRules", "priceRuleService.deleteByIds(request.ids())", "選択された料金ルールを一括削除する。", "room_price_rules bulk delete", "API-009"],
        ["TRACE-026", 26, "ApiController.java", "104", "reservations", "reservationService.syncDueCheckouts()", "予約一覧取得前にチェックアウト対象を同期する。", "予約・客室状態更新", "LOGIC-002"],
        ["TRACE-027", 27, "ApiController.java", "105-110", "reservations", "new ReservationsResponse(...)", "予約中、取消、チェックアウト、予約可能客室、業務日付を返す。", "ReservationsResponse", "DTO-006"],
        ["TRACE-028", 28, "ApiController.java", "115-122", "createReservation", "reservationService.create(...)", "予約本体と同行者配列を予約登録 Service へ渡す。", "reservations / reservation_guests / rooms 更新", "API-011"],
        ["TRACE-029", 29, "ApiController.java", "117", "createReservation", "request.noPhoneInfo() && request.noEmailInfo()", "電話とメールの両方が非保持指定の場合だけ noContactInfo=true にする。", "連絡先 NULL 保存条件", "RULE-004"],
        ["TRACE-030", 30, "ApiController.java", "128", "updatePayment", "reservationService.updatePaymentStatus(...)", "支払状態を更新する。", "payment_status update", "API-012"],
        ["TRACE-031", 31, "ApiController.java", "134", "cancelReservation", "reservationService.cancel(id)", "予約を取消し、客室状態を戻す。", "cancelled / vacant / cleaned", "API-013"],
        ["TRACE-032", 32, "ApiController.java", "142", "updateReservationStatus", "reservationService.updateReservationStatus(...)", "予約状態を更新し、客室状態も同期する。", "reservation_status / room status", "API-014"],
        ["TRACE-033", 33, "ApiController.java", "148", "updateCleaning", "reservationService.updateCheckoutCleaningStatus(...)", "チェックアウト後の清掃状態を更新する。", "room cleaning_status", "API-015"],
        ["TRACE-033A", 33, "ApiController.java", "156", "deleteCheckedOutReservation", "reservationService.deleteCheckedOut(...)", "チェックアウト済み予約を一覧から完全削除する。", "reservation deleted", "API-014A"],
        ["TRACE-034", 34, "ApiController.java", "156", "handleIllegalArgument", "new ErrorResponse(ex.getMessage())", "Service の業務例外を React 表示用 JSON に変換する。", "HTTP 400 / error", "API-016"],
        ["TRACE-035", 35, "ApiController.java", "161", "firstPage", "new PageResponse<>(items, 1, totalPages(...))", "初期表示用のページ情報を page=1 で統一する。", "PageResponse", "DTO-003"],
        ["TRACE-036", 36, "ApiController.java", "166", "totalPages", "Math.max(1, ...)", "0件でも totalPages=1 として画面表示を安定させる。", "totalPages", "DTO-003"],
        ["TRACE-037", 37, "ReservationService.java", "54-57", "constructor", "this.*Mapper = ...", "予約 Service に必要な Mapper を保持する。", "DI 済み Mapper", "JAVA-008"],
        ["TRACE-038", 38, "ReservationService.java", "62", "currentDate", "reservationMapper.currentDate()", "DB 基準の業務日付を取得する。", "LocalDate", "SQL-001"],
        ["TRACE-039", 39, "ReservationService.java", "82-83", "findRecentPage", "safePageSize / pageOffset", "ページサイズを補正し、予約中一覧を取得する。", "List<Reservation>", "LOGIC-002"],
        ["TRACE-040", 40, "ReservationService.java", "88-89", "findCancelledPage", "safePageSize / pageOffset", "取消済み一覧をページング取得する。", "List<Reservation>", "LOGIC-002"],
        ["TRACE-041", 41, "ReservationService.java", "94-95", "findCheckedOutPage", "safePageSize / pageOffset", "チェックアウト済み一覧をページング取得する。", "List<Reservation>", "LOGIC-002"],
        ["TRACE-042", 42, "ReservationService.java", "116", "syncDueCheckouts", "reservationMapper.findDueCheckouts()", "期限到来した予約を抽出する。", "List<Reservation>", "SQL-006"],
        ["TRACE-043", 43, "ReservationService.java", "117", "syncDueCheckouts", "for (Reservation dueReservation ...)", "期限到来予約を1件ずつ処理する。", "loop", "LOGIC-009"],
        ["TRACE-044", 44, "ReservationService.java", "118", "syncDueCheckouts", "reservationMapper.markCheckedOut(...)", "予約状態を checked_out に更新する。", "reservation_status=checked_out", "STAT-004"],
        ["TRACE-045", 45, "ReservationService.java", "120", "syncDueCheckouts", "roomMapper.updateStatuses(..., vacant, needs_cleaning)", "対象客室を空室・清掃待ちへ戻す。", "rooms status update", "STAT-010"],
        ["TRACE-046", 46, "ReservationService.java", "137", "create", "validateReservation(...)", "予約本体の必須、日付、代表者、連絡先、人数を検証する。", "検証通過または例外", "RULE-001〜004"],
        ["TRACE-047", 47, "ReservationService.java", "138", "create", "validateCompanions(...)", "宿泊人数に応じた同行者名の必須入力を検証する。", "同行者検証", "RULE-007"],
        ["TRACE-048", 48, "ReservationService.java", "139", "create", "validateCompanionContacts(...)", "同行者フリガナ・電話番号の形式を検証する。", "同行者連絡先検証", "RULE-007"],
        ["TRACE-049", 49, "ReservationService.java", "142", "create", "roomMapper.findByIdForUpdate(...)", "予約対象客室を行ロック付きで取得する。", "Room / FOR UPDATE", "BNF-006"],
        ["TRACE-050", 50, "ReservationService.java", "143", "create", "validateRoomForReservation(...)", "客室が営業中・空室・清掃済み・定員内か確認する。", "予約可能客室", "RULE-005"],
        ["TRACE-051", 51, "ReservationService.java", "144", "create", "validateNoOverlappingReservation(...)", "同一客室の重複予約を検証する。", "重複なし", "RULE-006"],
        ["TRACE-052", 52, "ReservationService.java", "147", "create", "setReservationNo(nextReservationNo())", "DB シーケンスから R + 6桁予約番号を採番して設定する。", "reservation_no", "BR-003"],
        ["TRACE-053", 53, "ReservationService.java", "148", "create", "setReservationStatus(\"booked\")", "新規予約の予約状態を予約中に設定する。", "reservation_status=booked", "STAT-003"],
        ["TRACE-054", 54, "ReservationService.java", "149-152", "create", "if paymentStatus empty -> unpaid", "支払状態未入力時、未払いを初期値にする。", "payment_status=unpaid", "STAT-001"],
        ["TRACE-055", 55, "ReservationService.java", "153-156", "create", "if reservationForm empty -> 公式", "予約経路未入力時、公式予約として扱う。", "reservation_form=公式", "RULE-001"],
        ["TRACE-056", 56, "ReservationService.java", "159", "create", "setTotalAmount(calculateTotalAmount(...))", "宿泊日ごとの単価から合計金額を算出して設定する。", "total_amount", "RULE-008"],
        ["TRACE-057", 57, "ReservationService.java", "160", "create", "reservationMapper.insert(reservation)", "予約本体を DB に登録する。", "reservations insert / id 採番", "SQL-012"],
        ["TRACE-058", 58, "ReservationService.java", "162-169", "create", "saveCompanions(...)", "予約 ID 確定後、同行者明細を保存する。", "reservation_guests insert", "BR-004"],
        ["TRACE-059", 59, "ReservationService.java", "171", "create", "roomMapper.updateStatuses(... reserved ...)", "予約成立後、客室を予約済みに切り替える。", "occupancy_status=reserved", "STAT-007"],
        ["TRACE-060", 60, "ReservationService.java", "182", "updatePaymentStatus", "requireAllowed(paymentStatus,...)", "支払状態が unpaid / paid のどちらかを検証する。", "許可値検証", "STAT-001 / STAT-002"],
        ["TRACE-061", 61, "ReservationService.java", "183-185", "updatePaymentStatus", "updatePaymentStatus == 0 -> throw", "支払更新対象が存在しない場合は業務エラーにする。", "予約なしエラー", "RULE-001"],
        ["TRACE-062", 62, "ReservationService.java", "194", "updateReservationStatus", "requireAllowed(reservationStatus,...)", "予約状態が booked / checked_out / cancelled のいずれかを検証する。", "許可値検証", "STAT-003〜005"],
        ["TRACE-063", 63, "ReservationService.java", "196-199", "updateReservationStatus", "findById / null check", "状態更新対象の予約を取得し、存在しない場合はエラーにする。", "Reservation または例外", "LOGIC-010"],
        ["TRACE-064", 64, "ReservationService.java", "200-202", "updateReservationStatus", "updateReservationStatus == 0 -> throw", "予約状態更新結果が0件なら対象なしエラーにする。", "予約なしエラー", "LOGIC-010"],
        ["TRACE-065", 65, "ReservationService.java", "203-206", "updateReservationStatus", "if booked -> reserved / cleaned", "予約中へ戻した場合、客室を予約済み・清掃済みに同期する。", "rooms status update", "STAT-003"],
        ["TRACE-066", 66, "ReservationService.java", "207-210", "updateReservationStatus", "if checked_out -> vacant / needs_cleaning", "チェックアウト済みにした場合、客室を空室・清掃待ちに同期する。", "rooms status update", "STAT-004"],
        ["TRACE-067", 67, "ReservationService.java", "215-218", "updateCheckoutCleaningStatus", "findById / null check", "清掃状態更新対象の予約を取得し、存在しない場合はエラーにする。", "Reservation または例外", "LOGIC-011"],
        ["TRACE-067A", 67, "ReservationService.java", "220-222", "updateCheckoutCleaningStatus", "reservationStatus != checked_out -> error", "チェックアウト済み予約以外の清掃状態更新を拒否する。", "清掃更新拒否", "LOGIC-011"],
        ["TRACE-068", 68, "ReservationService.java", "221", "updateCheckoutCleaningStatus", "requireAllowed(cleaningStatus,...)", "清掃状態が needs_cleaning / cleaned のどちらかを検証する。", "許可値検証", "STAT-009 / STAT-010"],
        ["TRACE-069", 69, "ReservationService.java", "222", "updateCheckoutCleaningStatus", "roomMapper.updateStatuses(... vacant, cleaningStatus)", "客室を空室のまま、指定清掃状態へ更新する。", "rooms status update", "REQ-009"],
        ["TRACE-070", 70, "ReservationService.java", "227", "cancel", "reservationMapper.findById(id)", "取消対象予約を取得し、客室 ID を後続同期に使う。", "Reservation", "LOGIC-011"],
        ["TRACE-070A", 70, "ReservationService.java", "243-253", "deleteCancelled", "findById -> status check -> deleteCancelled", "取消済み予約だけを一覧から完全削除する。", "予約削除または業務エラー", "F-004A"],
        ["TRACE-070AA", 70, "ReservationService.java", "256-265", "deleteCheckedOut", "findById -> status check -> deleteCheckedOut", "チェックアウト済み予約だけを一覧から完全削除する。", "予約削除または業務エラー", "F-004B"],
        ["TRACE-070B", 70, "ReservationService.java", "256-265", "updateRoomAfterReservationRelease", "countOtherBookedByRoomId -> room status", "同室に他の予約中データが残る場合は客室を reserved / cleaned に維持し、ない場合だけ vacant に戻す。", "客室状態更新", "LOGIC-010"],
        ["TRACE-070C", 70, "ReservationService.java", "267-269", "nextReservationNo", "'R' + format('%06d', nextReservationSequence)", "DB 連番を R000001 形式へ整形する。", "予約番号", "SQL-002"],
        ["TRACE-071", 71, "ReservationService.java", "228-230", "cancel", "reservationMapper.cancel(id) == 0", "取消更新が0件なら予約なしエラーにする。", "reservation_status=cancelled または例外", "STAT-005"],
        ["TRACE-072", 72, "ReservationService.java", "231-234", "cancel", "if reservation != null -> vacant / cleaned", "取消後、客室を空室・清掃済みに戻す。", "rooms status update", "REQ-006"],
        ["TRACE-073", 73, "ReservationService.java", "238-240", "requireAllowed", "blank or not contains -> throw", "状態値が空または許可値外なら業務エラーにする。", "IllegalArgumentException", "STAT-001〜010"],
        ["TRACE-074", 74, "ReservationService.java", "245", "safePageSize", "Math.min(MAX_PAGE_SIZE, Math.max(1, pageSize))", "ページサイズを 1〜100 に補正する。", "safePageSize", "BNF-007"],
        ["TRACE-075", 75, "ReservationService.java", "250-251", "pageOffset", "safePage / offset calculation", "ページ番号を1以上に補正し、SQL OFFSET を算出する。", "offset", "LOGIC-002"],
        ["TRACE-076", 76, "ReservationService.java", "256-257", "validateReservation", "roomId == null", "部屋未選択を拒否する。", "部屋選択エラー", "RULE-001"],
        ["TRACE-077", 77, "ReservationService.java", "259-260", "validateReservation", "checkInDate/checkOutDate == null", "宿泊日未入力を拒否する。", "宿泊日エラー", "RULE-002"],
        ["TRACE-078", 78, "ReservationService.java", "262-264", "validateReservation", "checkInDate.isBefore(currentDate())", "過去チェックイン日を拒否する。", "過去日エラー", "REQ-003"],
        ["TRACE-079", 79, "ReservationService.java", "266-267", "validateReservation", "!checkInDate.isBefore(checkOutDate)", "チェックアウト日がチェックイン日以前の予約を拒否する。", "宿泊期間エラー", "REQ-003"],
        ["TRACE-080", 80, "ReservationService.java", "269-270", "validateReservation", "!hasText(guestName)", "宿泊者名未入力を拒否する。", "宿泊者名エラー", "RULE-001"],
        ["TRACE-081", 81, "ReservationService.java", "273", "validateReservation", "validateOptionalContact(guestKana,...)", "代表者フリガナ形式を検証する。", "カナ形式検証", "RULE-003"],
        ["TRACE-082", 82, "ReservationService.java", "275-278", "validateReservation", "if !noContactInfo", "連絡先保持時は電話番号とメール形式を検証する。", "電話・メール形式検証", "RULE-003"],
        ["TRACE-083", 83, "ReservationService.java", "279-283", "validateReservation", "else setGuestPhone/Email null", "連絡先非保持時は電話・メールを NULL にする。", "guest_phone/email=NULL", "RULE-004"],
        ["TRACE-084", 84, "ReservationService.java", "284-285", "validateReservation", "guestCount null or < 1", "宿泊人数1名未満を拒否する。", "人数エラー", "RULE-001"],
        ["TRACE-085", 85, "ReservationService.java", "287-288", "validateReservation", "guestCount > MAX_GUEST_COUNT", "宿泊人数10名超を拒否する。", "人数上限エラー", "UT-003"],
        ["TRACE-086", 86, "ReservationService.java", "294-295", "validateRoomForReservation", "room null or !active", "存在しない、または無効客室を拒否する。", "利用可能部屋エラー", "RULE-005"],
        ["TRACE-087", 87, "ReservationService.java", "297-298", "validateRoomForReservation", "occupancyStatus != vacant", "空室以外の客室を拒否する。", "空室エラー", "RULE-005"],
        ["TRACE-088", 88, "ReservationService.java", "300-301", "validateRoomForReservation", "cleaningStatus != cleaned", "未清掃客室を拒否する。", "清掃済みエラー", "RULE-005"],
        ["TRACE-089", 89, "ReservationService.java", "303-304", "validateRoomForReservation", "guestCount > capacity", "定員超過の予約を拒否する。", "定員超過エラー", "RULE-005"],
        ["TRACE-090", 90, "ReservationService.java", "310-313", "validateNoOverlappingReservation", "countOverlapping(...)", "同一客室・期間交差する booked 予約数を取得する。", "overlaps", "SQL-007"],
        ["TRACE-091", 91, "ReservationService.java", "315-316", "validateNoOverlappingReservation", "overlaps > 0", "重複予約がある場合は登録を拒否する。", "重複予約エラー", "REQ-004"],
        ["TRACE-092", 92, "ReservationService.java", "322", "validateCompanions", "requiredCount = guestCount - 1", "代表者を除いた同行者必須件数を算出する。", "requiredCount", "REQ-002"],
        ["TRACE-093", 93, "ReservationService.java", "323-324", "validateCompanions", "requiredCount == 0 -> return", "1名予約では同行者検証を終了する。", "検証終了", "REQ-002"],
        ["TRACE-094", 94, "ReservationService.java", "326-327", "validateCompanions", "names null or size < requiredCount", "同行者配列が不足している場合は拒否する。", "同行者情報エラー", "REQ-002"],
        ["TRACE-095", 95, "ReservationService.java", "330-333", "validateCompanions", "for / !hasText(companionNames[i])", "同行者名の空欄を拒否する。", "同行者名エラー", "REQ-002"],
        ["TRACE-096", 96, "ReservationService.java", "342-345", "validateCompanionContacts", "for validateOptionalContact(...)", "同行者フリガナと電話番号形式を人数分検証する。", "同行者形式検証", "UT-011"],
        ["TRACE-097", 97, "ReservationService.java", "350-351", "validateOptionalContact", "!hasText(value) -> return", "任意項目が空の場合は形式検証を省略する。", "検証通過", "RULE-003"],
        ["TRACE-098", 98, "ReservationService.java", "353-354", "validateOptionalContact", "!pattern.matcher(value).matches()", "入力が指定パターンに合わない場合は業務エラーにする。", "形式エラー", "RULE-003"],
        ["TRACE-099", 99, "ReservationService.java", "367", "saveCompanions", "companionCount = guestCount - 1", "保存対象の同行者数を算出する。", "companionCount", "BR-004"],
        ["TRACE-100", 100, "ReservationService.java", "368-376", "saveCompanions", "for / new ReservationGuest / setters / insert", "同行者ごとに明細オブジェクトを作成し DB に登録する。", "reservation_guests insert", "BR-004"],
        ["TRACE-101", 101, "ReservationService.java", "381", "valueAt", "values null or size <= index ? null : values.get(index)", "配列不足時は null として扱い、同行者任意項目を安全に取得する。", "String or null", "LOGIC-004"],
        ["TRACE-102", 102, "ReservationService.java", "385", "integerAt", "values null or size <= index ? null : values.get(index)", "配列不足時は null として扱い、同行者年齢を安全に取得する。", "Integer or null", "LOGIC-004"],
        ["TRACE-103", 103, "ReservationService.java", "390", "calculateTotalAmount", "BigDecimal total = BigDecimal.ZERO", "料金合計を0円で初期化する。", "total=0", "RULE-008"],
        ["TRACE-104", 104, "ReservationService.java", "391", "calculateTotalAmount", "stayDate = checkInDate", "料金計算の開始日をチェックイン日に設定する。", "stayDate", "RULE-008"],
        ["TRACE-105", 105, "ReservationService.java", "392", "calculateTotalAmount", "while stayDate before checkOutDate", "チェックアウト日前日まで日別に繰り返す。", "日別 loop", "RULE-008"],
        ["TRACE-106", 106, "ReservationService.java", "393", "calculateTotalAmount", "priceRuleMapper.findBestRule(...)", "対象日の最優先料金ルールを取得する。", "RoomPriceRule or null", "SQL-011"],
        ["TRACE-107", 107, "ReservationService.java", "394", "calculateTotalAmount", "rule == null ? basePrice : rulePrice", "料金ルール未設定日は客室基本単価を採用する。", "price", "RULE-008"],
        ["TRACE-108", 108, "ReservationService.java", "395", "calculateTotalAmount", "total.add(price * guestCount)", "日別単価に人数を掛けて合計へ加算する。", "total updated", "RULE-008"],
        ["TRACE-109", 109, "ReservationService.java", "396", "calculateTotalAmount", "stayDate.plus(1, DAYS)", "次の宿泊日に進める。", "stayDate + 1", "RULE-008"],
        ["TRACE-110", 110, "ReservationService.java", "398", "calculateTotalAmount", "return total", "算出した宿泊合計金額を返す。", "totalAmount", "RULE-008"],
        ["TRACE-111", 111, "RoomService.java", "61-62", "create", "!hasText(roomNumber)", "部屋番号未入力を拒否する。", "部屋番号エラー", "RULE-012"],
        ["TRACE-112", 112, "RoomService.java", "64-65", "create", "!hasText(roomName)", "部屋名未入力を拒否する。", "部屋名エラー", "RULE-012"],
        ["TRACE-113", 113, "RoomService.java", "67-68", "create", "capacity null or < 1", "定員1名未満を拒否する。", "定員エラー", "RULE-012"],
        ["TRACE-114", 114, "RoomService.java", "70-72", "create", "basePricePerPerson null -> ZERO", "基本料金未入力時は0円に補完する。", "basePricePerPerson=0", "RULE-012"],
        ["TRACE-115", 115, "RoomService.java", "74-76", "create", "roomType empty -> washitsu", "部屋タイプ未入力時は和室を初期値にする。", "roomType=washitsu", "RULE-012"],
        ["TRACE-116", 116, "RoomService.java", "79", "create", "requireAllowed(roomType,...)", "部屋タイプが許可値か検証する。", "washitsu/yoshitsu/suite/family", "RULE-012"],
        ["TRACE-117", 117, "RoomService.java", "81-82", "create", "privateBath null -> false", "専用風呂未入力時は false に補完する。", "privateBath=false", "RULE-012"],
        ["TRACE-118", 118, "RoomService.java", "84-86", "create", "occupancyStatus empty -> vacant", "宿泊状態未入力時は空室に補完する。", "occupancyStatus=vacant", "STAT-006"],
        ["TRACE-119", 119, "RoomService.java", "88-90", "create", "cleaningStatus empty -> cleaned", "清掃状態未入力時は清掃済みに補完する。", "cleaningStatus=cleaned", "STAT-009"],
        ["TRACE-120", 120, "RoomService.java", "93-94", "create", "requireAllowed(statuses,...)", "宿泊状態と清掃状態が許可値か検証する。", "状態値検証", "RULE-013"],
        ["TRACE-121", 121, "RoomService.java", "96-97", "create", "active null -> true", "有効フラグ未入力時は有効に補完する。", "active=true", "REQ-007"],
        ["TRACE-122", 122, "RoomService.java", "100", "create", "findByRoomNumberIncludingInactive", "同一部屋番号の既存客室を論理削除済み含めて検索する。", "Room or null", "REQ-008"],
        ["TRACE-123", 123, "RoomService.java", "102-103", "create", "existing active -> throw", "同一番号の有効客室がある場合は重複エラーにする。", "重複エラー", "RULE-012"],
        ["TRACE-124", 124, "RoomService.java", "105-108", "create", "existing inactive -> reactivate / return", "同一番号の削除済み客室は再有効化して処理終了する。", "rooms active=true", "REQ-008"],
        ["TRACE-125", 125, "RoomService.java", "111", "create", "roomMapper.insert(room)", "既存客室がない場合は新規客室を登録する。", "rooms insert", "REQ-007"],
        ["TRACE-126", 126, "RoomService.java", "120-121", "updateStatuses", "requireAllowed(...)", "宿泊状態と清掃状態の許可値を検証する。", "状態値検証", "RULE-013"],
        ["TRACE-127", 127, "RoomService.java", "123-124", "updateStatuses", "updateStatuses == 0 -> throw", "状態更新対象が存在しない場合はエラーにする。", "部屋なしエラー", "RULE-013"],
        ["TRACE-128", 128, "RoomService.java", "131-132", "delete", "deactivate == 0 -> throw", "客室を論理削除し、対象なしならエラーにする。", "active=false or error", "REQ-007"],
        ["TRACE-129", 129, "RoomPriceRuleService.java", "36", "create", "validateRule(rule)", "料金ルールの必須、期間、金額、初期値を検証する。", "検証通過", "RULE-014"],
        ["TRACE-130", 130, "RoomPriceRuleService.java", "37-40", "create", "countOverlapping > 0 -> throw", "同一客室・重複期間の有効料金ルールを拒否する。", "期間重複エラー", "REQ-010"],
        ["TRACE-131", 131, "RoomPriceRuleService.java", "42", "create", "priceRuleMapper.insert(rule)", "検証済み料金ルールを登録する。", "room_price_rules insert", "REQ-010"],
        ["TRACE-132", 132, "RoomPriceRuleService.java", "47-48", "delete", "id == null -> throw", "削除対象未選択を拒否する。", "料金ルール選択エラー", "RULE-015"],
        ["TRACE-133", 133, "RoomPriceRuleService.java", "50-51", "delete", "delete(id) == 0 -> throw", "単一削除対象が存在しない場合はエラーにする。", "削除 or 対象なしエラー", "RULE-015"],
        ["TRACE-134", 134, "RoomPriceRuleService.java", "57-58", "deleteByIds", "ids null or empty -> throw", "一括削除の未選択を拒否する。", "選択エラー", "RULE-015"],
        ["TRACE-135", 135, "RoomPriceRuleService.java", "62", "deleteByIds", "filter nonNull / distinct / toList", "NULL と重複を除外して削除対象 ID を正規化する。", "targetIds", "RULE-015"],
        ["TRACE-136", 136, "RoomPriceRuleService.java", "64-65", "deleteByIds", "targetIds empty -> throw", "正規化後に対象がない場合は未選択エラーにする。", "選択エラー", "RULE-015"],
        ["TRACE-137", 137, "RoomPriceRuleService.java", "67-68", "deleteByIds", "deleteByIds == 0 -> throw", "一括削除結果が0件なら対象なしエラーにする。", "削除 or 対象なしエラー", "RULE-015"],
        ["TRACE-138", 138, "RoomPriceRuleService.java", "74-75", "validateRule", "roomId == null", "料金対象客室未選択を拒否する。", "部屋選択エラー", "RULE-014"],
        ["TRACE-139", 139, "RoomPriceRuleService.java", "77-78", "validateRule", "!hasText(ruleName)", "料金ルール名未入力を拒否する。", "名称エラー", "RULE-014"],
        ["TRACE-140", 140, "RoomPriceRuleService.java", "80-81", "validateRule", "startDate/endDate == null", "開始日・終了日未入力を拒否する。", "日付エラー", "RULE-014"],
        ["TRACE-141", 141, "RoomPriceRuleService.java", "83-84", "validateRule", "startDate.isAfter(endDate)", "開始日が終了日より後の料金ルールを拒否する。", "期間エラー", "RULE-014"],
        ["TRACE-142", 142, "RoomPriceRuleService.java", "86-87", "validateRule", "price null or < 0", "料金未入力または負数を拒否する。", "料金エラー", "RULE-014"],
        ["TRACE-143", 143, "RoomPriceRuleService.java", "89-91", "validateRule", "priority null -> 10", "優先度未入力時は10に補完する。", "priority=10", "RULE-014"],
        ["TRACE-144", 144, "RoomPriceRuleService.java", "93-95", "validateRule", "active null -> true", "有効フラグ未入力時は true に補完する。", "active=true", "RULE-014"],
        ["TRACE-145", 145, "ReservationController.java", "29-30", "reservations", "return \"app\"", "予約管理画面 URL で React shell を返す。", "viewName=app", "SCR-002"],
        ["TRACE-146", 146, "ReservationController.java", "49-56", "create", "reservationService.create(...)", "フォーム投稿の予約情報と同行者配列を Service へ渡す。", "予約登録", "SCR-002"],
        ["TRACE-147", 147, "ReservationController.java", "57", "create", "addFlashAttribute(\"message\", ...)", "予約登録成功メッセージを flash に設定する。", "flash message", "SCR-002"],
        ["TRACE-148", 148, "ReservationController.java", "58-59", "create", "catch IllegalArgumentException", "予約登録の業務エラーを flash error に設定する。", "flash error", "SCR-002"],
        ["TRACE-149", 149, "ReservationController.java", "61", "create", "return redirect:/reservations", "処理後に予約画面へリダイレクトする。", "redirect", "SCR-002"],
        ["TRACE-150", 150, "ReservationController.java", "70-72", "updatePayment", "updatePaymentStatus / message / redirect", "支払状態更新後、成功メッセージ付きで予約画面へ戻る。", "payment update / redirect", "F-005"],
        ["TRACE-151", 151, "ReservationController.java", "78-80", "cancel", "cancel / message / redirect", "予約取消後、成功メッセージ付きで予約画面へ戻る。", "cancel / redirect", "F-004"],
        ["TRACE-151A", 151, "ReservationController.java", "80-84", "deleteCancelled", "deleteCancelled / message / redirect", "取消済み予約削除後、成功メッセージ付きで予約画面へ戻る。", "delete / redirect", "F-004A"],
        ["TRACE-152", 152, "ReservationController.java", "89-91", "updateStatus", "updateReservationStatus / message / redirect", "予約状態更新後、成功メッセージ付きで予約画面へ戻る。", "status update / redirect", "F-014"],
        ["TRACE-153", 153, "ReservationController.java", "100-102", "updateCleaning", "updateCheckoutCleaningStatus / message / redirect", "清掃状態更新後、成功メッセージ付きで予約画面へ戻る。", "cleaning update / redirect", "F-006"],
        ["TRACE-154", 154, "RoomController.java", "28-29", "rooms", "return \"app\"", "客室管理画面 URL で React shell を返す。", "viewName=app", "SCR-003"],
        ["TRACE-155", 155, "RoomController.java", "39-40", "create", "roomService.create / message", "客室登録成功時、成功メッセージを flash に設定する。", "rooms insert/reactivate", "F-008"],
        ["TRACE-155A", 155, "RoomService.java", "143-153", "restore", "findById -> active check -> restore", "削除済み客室を有効状態へ戻す。", "rooms active=true", "F-008A"],
        ["TRACE-155B", 155, "RoomService.java", "159-174", "deletePermanently", "inactive check + countReservationsByRoomId == 0 -> delete", "予約履歴のない削除済み客室だけ完全削除する。", "rooms delete", "F-008A"],
        ["TRACE-156", 156, "RoomController.java", "41-42", "create", "catch IllegalArgumentException", "客室登録の業務エラーを flash error に設定する。", "flash error", "SCR-003"],
        ["TRACE-157", 157, "RoomController.java", "43-46", "create", "catch RuntimeException", "想定外登録失敗を汎用エラーメッセージに変換する。", "flash error", "SCR-003"],
        ["TRACE-158", 158, "RoomController.java", "58-60", "updateStatuses", "updateStatuses / message / redirect", "客室状態更新後、成功メッセージ付きで客室画面へ戻る。", "room status update", "F-009"],
        ["TRACE-159", 159, "RoomController.java", "66-68", "delete", "delete / message / redirect", "客室論理削除後、成功メッセージ付きで客室画面へ戻る。", "active=false", "REQ-007"],
        ["TRACE-160", 160, "PriceRuleController.java", "29-30", "prices", "return \"app\"", "料金管理画面 URL で React shell を返す。", "viewName=app", "SCR-004"],
        ["TRACE-161", 161, "PriceRuleController.java", "40-41", "create", "priceRuleService.create / message", "料金ルール登録成功時、成功メッセージを flash に設定する。", "room_price_rules insert", "F-011"],
        ["TRACE-162", 162, "PriceRuleController.java", "42-43", "create", "catch IllegalArgumentException", "料金登録の業務エラーを flash error に設定する。", "flash error", "SCR-004"],
        ["TRACE-163", 163, "PriceRuleController.java", "52-57", "delete", "priceRuleService.delete / catch / redirect", "単一削除または旧URL互換削除を実行し、結果を flash に設定する。", "delete / redirect", "F-012"],
        ["TRACE-164", 164, "PriceRuleController.java", "66-71", "deleteSelected", "deleteByIds / catch / redirect", "複数選択削除を実行し、結果を flash に設定する。", "bulk delete / redirect", "F-012"],
    ])


def write_screens(ws):
    setup_sheet(ws, "画面一覧", 7)
    add_table(ws, 6, ["画面ID", "画面名", "利用者", "概要", "主な項目", "遷移元", "遷移先"], [
        ["S-001", "ダッシュボード", "社内担当者", "予約・客室の集計と直近予約を表示する。", "客室数、空室数、予約中件数、予約一覧", "-", "予約一覧、客室一覧、料金設定"],
        ["S-002", "予約一覧", "予約管理担当者", "予約登録、予約一覧、取消一覧、チェックアウト一覧を管理する。", "宿泊者、同行者、日程、支払、清掃", "ダッシュボード", "同画面更新"],
        ["S-003", "客室一覧", "施設管理担当者", "客室登録、状態更新、論理削除を行う。", "部屋番号、定員、料金、宿泊状態、清掃状態", "ダッシュボード", "同画面更新"],
        ["S-004", "料金設定", "施設管理担当者", "客室別・期間別の料金ルールを管理する。", "客室、期間、料金、優先度、状態", "ダッシュボード", "同画面更新"],
    ])
    add_table_after(ws, ["画面仕様ID", "URL", "MVC Controller", "React API", "初期表示データ", "主操作", "基本仕様"], [
        ["SCR-001", "/jukai-internal/ または /dashboard", "DashboardController.dashboard", "GET /api/dashboard", "客室総数、空室数、予約中件数、直近予約", "各管理画面へ遷移", "画面 Controller は app を返し、API が集計値を返す。"],
        ["SCR-002", "/jukai-internal/reservations", "ReservationController.reservations", "GET /api/reservations", "予約中、取消済み、チェックアウト済み、予約可能客室、業務日付", "予約登録、支払更新、取消、状態更新、清掃更新", "表示前に期限到来チェックアウト同期を実行する。"],
        ["SCR-003", "/jukai-internal/rooms", "RoomController.rooms", "GET /api/rooms", "有効客室、削除済み客室、予約可能客室", "客室登録、状態更新、論理削除", "同一部屋番号の削除済み客室は再有効化対象。"],
        ["SCR-004", "/jukai-internal/prices", "PriceRuleController.prices", "GET /api/prices", "料金ルール一覧、有効客室", "料金ルール登録、単一削除、一括削除", "料金ルールは予約金額計算の根拠として使用する。"],
    ])


def write_db(ws):
    setup_sheet(ws, "DB設計", 7)
    add_table(ws, 6, ["テーブル名", "カラム名", "型", "PK", "FK", "NULL", "説明"], [
        ["rooms", "id", "SERIAL", "YES", "", "NO", "客室ID。Room.id に対応。"],
        ["rooms", "room_number", "VARCHAR", "", "", "NO", "客室番号。重複登録判定に使用。"],
        ["rooms", "room_name", "VARCHAR", "", "", "NO", "客室名。"],
        ["rooms", "room_type", "VARCHAR", "", "", "NO", "washitsu / yoshitsu / suite / family。"],
        ["rooms", "capacity", "INTEGER", "", "", "NO", "定員。予約人数上限チェックに使用。"],
        ["rooms", "base_price_per_person", "NUMERIC", "", "", "NO", "料金ルール未設定日の基本単価。"],
        ["rooms", "occupancy_status", "VARCHAR", "", "", "NO", "vacant / reserved / occupied。"],
        ["rooms", "cleaning_status", "VARCHAR", "", "", "NO", "cleaned / needs_cleaning。"],
        ["rooms", "active", "BOOLEAN", "", "", "NO", "論理削除フラグ。"],
        ["reservations", "id", "SERIAL", "YES", "", "NO", "予約ID。"],
        ["reservations", "reservation_no", "VARCHAR(16)", "", "", "NO", "R + 6桁以上の予約番号。"],
        ["reservations", "room_id", "INTEGER", "", "rooms.id", "NO", "予約対象客室。"],
        ["reservations", "check_in_date / check_out_date", "DATE", "", "", "NO", "宿泊期間。重複予約・料金計算に使用。"],
        ["reservations", "guest_*", "VARCHAR/INTEGER", "", "", "一部YES", "代表宿泊者情報。"],
        ["reservations", "payment_status", "VARCHAR", "", "", "NO", "unpaid / paid。"],
        ["reservations", "reservation_status", "VARCHAR", "", "", "NO", "booked / checked_out / cancelled。"],
        ["reservations", "total_amount", "NUMERIC", "", "", "NO", "登録時点の確定金額。"],
        ["reservation_guests", "id", "SERIAL", "YES", "", "NO", "同行者ID。"],
        ["reservation_guests", "reservation_id", "INTEGER", "", "reservations.id", "NO", "予約本体への紐付け。"],
        ["reservation_guests", "guest_*", "VARCHAR/INTEGER", "", "", "一部YES", "同行者情報。"],
        ["room_price_rules", "id", "SERIAL", "YES", "", "NO", "料金ルールID。"],
        ["room_price_rules", "room_id", "INTEGER", "", "rooms.id", "NO", "対象客室。"],
        ["room_price_rules", "start_date / end_date", "DATE", "", "", "NO", "適用期間。"],
        ["room_price_rules", "price_per_person", "NUMERIC", "", "", "NO", "1人料金。"],
        ["room_price_rules", "priority", "INTEGER", "", "", "NO", "最適料金選択の優先度。"],
        ["reservation_no_seq", "value", "SEQUENCE", "", "", "NO", "予約番号連番発番。"],
    ])
    add_table_after(ws, ["DB仕様ID", "区分", "対象", "定義内容", "利用箇所", "仕様上の意味", "備考"], [
        ["DBR-001", "制約", "rooms.capacity", "CHECK capacity > 0", "RoomService.create", "定員0名以下の客室をDBでも拒否する。", "Service と DB の二重防御"],
        ["DBR-002", "制約", "rooms.base_price_per_person", "CHECK base_price_per_person >= 0", "料金計算", "基本料金が負数になることを防止する。", ""],
        ["DBR-003", "制約", "room_price_rules", "CHECK start_date <= end_date / price_per_person >= 0", "RoomPriceRuleService.create", "期間逆転・負数料金をDBでも拒否する。", ""],
        ["DBR-004", "制約", "reservations", "CHECK check_in_date < check_out_date / guest_count > 0 / total_amount >= 0", "ReservationService.create", "宿泊期間・人数・金額の最低整合性を保証する。", ""],
        ["DBR-005", "制約", "reservations.reservation_no", "CHECK reservation_no ~ '^R[0-9]{6,}$'", "nextReservationNo", "画面表示用予約番号を R + 6桁以上の数字に固定する。", "例 R000001"],
        ["DBR-006", "外部キー", "reservations.room_id", "REFERENCES rooms(id) ON DELETE RESTRICT", "予約登録", "予約がある客室の物理削除を防ぐ。", "画面削除は active=false"],
        ["DBR-007", "外部キー", "reservation_guests.reservation_id", "REFERENCES reservations(id) ON DELETE CASCADE", "同行者保存", "予約本体削除時に同行者を連動削除できる。", "現行画面は予約物理削除なし"],
        ["DBR-008", "View", "checkout_reservations", "reservation_status='checked_out' の予約を抽出", "findCheckedOutPage / countCheckedOut", "チェックアウト一覧の参照元を単純化する。", ""],
        ["DBR-009", "Index", "ix_reservations_room_status_dates", "room_id, reservation_status, check_in_date, check_out_date", "countOverlapping", "重複予約検出の検索効率を確保する。", ""],
        ["DBR-010", "Index", "ix_room_price_rules_lookup", "room_id, active, start_date, end_date, priority", "findBestRule", "日別料金ルール検索を支援する。", ""],
    ])
    add_table_after(ws, ["Mapper ID", "Mapper", "SQL ID", "対象テーブル", "検索・更新条件", "返却・更新内容", "呼出元"], [
        ["SQL-001", "ReservationMapper", "currentDate", "-", "SELECT CURRENT_DATE", "DB基準業務日付", "ReservationService.currentDate / validateReservation"],
        ["SQL-002", "ReservationMapper", "nextReservationSequence", "reservation_no_seq", "nextval を long で返し Service が R + 6桁へ整形", "予約番号", "ReservationService.create"],
        ["SQL-002A", "ReservationMapper", "deleteCancelled", "reservations", "DELETE WHERE id = ? AND reservation_status = 'cancelled'", "取消済み予約削除", "ReservationService.deleteCancelled"],
        ["SQL-002B", "ReservationMapper", "countOtherBookedByRoomId", "reservations", "同一 room_id の他 booked 予約件数を取得", "客室状態維持判定", "updateRoomAfterReservationRelease"],
        ["SQL-002C", "ReservationMapper", "deleteCheckedOut", "reservations", "DELETE WHERE id = ? AND reservation_status = 'checked_out'", "チェックアウト済み予約削除", "ReservationService.deleteCheckedOut"],
        ["SQL-003", "ReservationMapper", "findRecentPage", "reservations / rooms / reservation_guests", "reservation_status='booked'、limit/offset", "予約中一覧と同行者要約", "ApiController.reservations"],
        ["SQL-004", "ReservationMapper", "findCancelledPage", "reservations / rooms / reservation_guests", "reservation_status='cancelled'、updated_at desc", "取消済み一覧", "ApiController.reservations"],
        ["SQL-005", "ReservationMapper", "findCheckedOutPage", "checkout_reservations / rooms / reservation_guests", "view 参照、updated_at desc", "チェックアウト済み一覧", "ApiController.reservations"],
        ["SQL-006", "ReservationMapper", "findDueCheckouts", "reservations / rooms", "booked かつ check_out_date <= CURRENT_DATE", "期限到来予約", "syncDueCheckouts"],
        ["SQL-007", "ReservationMapper", "countOverlapping", "reservations", "booked かつ期間交差", "重複件数", "validateNoOverlappingReservation"],
        ["SQL-008", "RoomMapper", "findByIdForUpdate", "rooms", "id一致 FOR UPDATE", "ロック済み客室", "ReservationService.create"],
        ["SQL-009", "RoomMapper", "reactivate", "rooms", "room_number一致", "削除済み客室を active=true で更新", "RoomService.create"],
        ["SQL-010", "RoomPriceRuleMapper", "countOverlapping", "room_price_rules", "同一客室、active=true、期間交差", "重複件数", "RoomPriceRuleService.create"],
        ["SQL-011", "RoomPriceRuleMapper", "findBestRule", "room_price_rules", "roomId、active、stayDate BETWEEN start/end", "最優先料金ルール1件", "calculateTotalAmount"],
        ["SQL-012", "ReservationGuestMapper", "insert", "reservation_guests", "予約IDと同行者項目", "同行者レコード", "saveCompanions"],
    ])
    add_table_after(ws, ["Domain ID", "Domain", "フィールド群", "DB対応", "業務意味", "表示・処理用途", "備考"], [
        ["DOM-001", "Reservation", "id / reservationNo / roomId", "reservations.id / reservation_no / room_id", "予約識別と客室紐付け", "状態更新、取消、予約番号表示", ""],
        ["DOM-002", "Reservation", "roomNumber / roomName / roomCleaningStatus", "rooms 結合列", "予約一覧に客室情報を表示するための派生項目", "一覧表示、清掃状態更新", "予約テーブルには保持しない"],
        ["DOM-003", "Reservation", "checkInDate / checkOutDate", "reservations.check_in_date / check_out_date", "宿泊期間", "過去日検証、期間重複、料金計算、チェックアウト同期", "checkOutDate は宿泊日数計算の終端"],
        ["DOM-004", "Reservation", "guestName / guestKana / guestGender / guestAge / guestPhone / guestEmail / guestCount", "reservations.guest_*", "代表宿泊者情報", "予約登録、一覧表示、連絡先非保持", ""],
        ["DOM-005", "Reservation", "reservationForm / paymentStatus / reservationStatus", "reservations.reservation_form / payment_status / reservation_status", "予約経路、支払、予約状態", "状態更新、一覧分類、ラベル表示", "ラベル変換メソッドあり"],
        ["DOM-006", "Reservation", "totalAmount / note / companionSummary", "reservations.total_amount / note、reservation_guests 集約", "確定金額、業務メモ、同行者要約", "予約一覧表示", "companionSummary は SQL string_agg"],
        ["DOM-007", "ReservationGuest", "reservationId / guestName / guestKana / guestGender / guestAge / guestPhone", "reservation_guests", "同行者明細", "予約登録後の同行者保存、一覧要約生成", ""],
        ["DOM-008", "Room", "roomNumber / roomName / roomType / capacity / basePricePerPerson / privateBath", "rooms", "客室販売条件", "客室登録、定員検証、基本料金計算", ""],
        ["DOM-009", "Room", "occupancyStatus / cleaningStatus / active", "rooms", "予約可否・清掃状態・論理削除", "予約可能客室抽出、状態同期、削除済み一覧", "bookable は active=true/vacant/cleaned"],
        ["DOM-010", "RoomPriceRule", "roomId / ruleName / startDate / endDate / pricePerPerson / priority / active", "room_price_rules", "期間別料金条件", "料金登録、重複判定、日別料金検索", "priority 昇順、同順位は料金高い順"],
    ])


def write_api(ws):
    setup_sheet(ws, "API一覧", 8)
    add_table(ws, 6, ["API ID", "分類", "メソッド", "URL", "概要", "リクエスト", "レスポンス", "認証"], [
        ["API-001", "Dashboard", "GET", "/api/dashboard", "集計と直近予約を取得", "-", "DashboardResponse", "社内"],
        ["API-002", "Room", "GET", "/api/rooms", "客室一覧を取得", "-", "RoomsResponse", "社内"],
        ["API-003", "Room", "POST", "/api/rooms", "客室登録", "Room", "MessageResponse", "社内"],
        ["API-004", "Room", "POST", "/api/rooms/{id}/statuses", "客室状態更新", "RoomStatusRequest", "MessageResponse", "社内"],
        ["API-005", "Room", "POST", "/api/rooms/{id}/delete", "客室論理削除", "path id", "MessageResponse", "社内"],
        ["API-005A", "Room", "POST", "/api/rooms/{id}/restore", "削除済み客室復元", "path id", "MessageResponse", "社内"],
        ["API-005B", "Room", "POST", "/api/rooms/{id}/delete-permanently", "削除済み客室完全削除", "path id", "MessageResponse", "社内"],
        ["API-006", "Price", "GET", "/api/prices", "料金一覧取得", "-", "PricesResponse", "社内"],
        ["API-007", "Price", "POST", "/api/prices", "料金ルール登録", "RoomPriceRule", "MessageResponse", "社内"],
        ["API-008", "Price", "POST", "/api/prices/{id}/delete", "料金ルール削除", "path id", "MessageResponse", "社内"],
        ["API-009", "Price", "POST", "/api/prices/delete-selected", "料金ルール一括削除", "IdsRequest", "MessageResponse", "社内"],
        ["API-010", "Reservation", "GET", "/api/reservations", "予約一覧取得", "-", "ReservationsResponse", "社内"],
        ["API-011", "Reservation", "POST", "/api/reservations", "予約登録", "ReservationCreateRequest", "MessageResponse", "社内"],
        ["API-012", "Reservation", "POST", "/api/reservations/{id}/payment", "支払更新", "PaymentRequest", "MessageResponse", "社内"],
        ["API-013", "Reservation", "POST", "/api/reservations/{id}/cancel", "予約取消", "path id", "MessageResponse", "社内"],
        ["API-013A", "Reservation", "POST", "/api/reservations/{id}/delete", "取消済み予約削除", "path id", "MessageResponse", "社内"],
        ["API-014", "Reservation", "POST", "/api/reservations/{id}/status", "予約状態更新", "ReservationStatusRequest", "MessageResponse", "社内"],
        ["API-014A", "Reservation", "POST", "/api/reservations/{id}/delete-checked-out", "チェックアウト済み予約削除", "path id", "MessageResponse", "社内"],
        ["API-015", "Reservation", "POST", "/api/reservations/{id}/cleaning", "清掃状態更新", "CleaningRequest", "MessageResponse", "社内"],
        ["API-016", "Error", "-", "ExceptionHandler", "業務エラー JSON 変換", "IllegalArgumentException", "ErrorResponse", "社内"],
    ])
    add_table_after(ws, ["DTO ID", "DTO / record", "方向", "フィールド", "型", "必須", "説明", "利用API"], [
        ["DTO-001", "DashboardResponse", "Response", "roomCount / vacantCount / bookedCount", "int", "YES", "客室総数、空室数、予約中件数", "API-001"],
        ["DTO-002", "DashboardResponse", "Response", "recentReservations", "PageResponse<Reservation>", "YES", "直近予約1ページ目", "API-001"],
        ["DTO-003", "PageResponse<T>", "Response", "items / page / totalPages", "List<T> / int / int", "YES", "一覧データとページ情報。0件でも totalPages=1", "API-001, API-010"],
        ["DTO-004", "RoomsResponse", "Response", "rooms / deletedRooms / bookableRooms", "List<Room>", "YES", "有効客室、論理削除済み客室、予約可能客室", "API-002"],
        ["DTO-005", "PricesResponse", "Response", "rules / rooms", "List<RoomPriceRule> / List<Room>", "YES", "料金ルール一覧と登録対象客室", "API-006"],
        ["DTO-006", "ReservationsResponse", "Response", "reservations / cancelledReservations / checkedOutReservations / rooms / today", "PageResponse / List<Room> / LocalDate", "YES", "予約管理画面の初期表示データ", "API-010"],
        ["DTO-007", "RoomStatusRequest", "Request", "occupancyStatus / cleaningStatus", "String", "YES", "宿泊状態と清掃状態", "API-004"],
        ["DTO-008", "IdsRequest", "Request", "ids", "List<Integer>", "YES", "一括削除対象料金ルールID", "API-009"],
        ["DTO-009", "ReservationCreateRequest", "Request", "reservation", "Reservation", "YES", "予約本体。roomId、宿泊日、代表者、人数等を含む", "API-011"],
        ["DTO-010", "ReservationCreateRequest", "Request", "noPhoneInfo / noEmailInfo", "boolean", "YES", "電話・メール非保持チェック。両方 true の場合だけ連絡先非保持", "API-011"],
        ["DTO-011", "ReservationCreateRequest", "Request", "companionNames / companionKanas / companionGenders / companionAges / companionPhones", "List", "NO", "同行者配列。宿泊人数-1件分を保存対象にする", "API-011"],
        ["DTO-012", "PaymentRequest", "Request", "paymentStatus", "String", "YES", "unpaid / paid", "API-012"],
        ["DTO-013", "ReservationStatusRequest", "Request", "reservationStatus", "String", "YES", "booked / checked_out / cancelled", "API-014"],
        ["DTO-014", "CleaningRequest", "Request", "cleaningStatus", "String", "YES", "needs_cleaning / cleaned", "API-015"],
        ["DTO-015", "MessageResponse", "Response", "message", "String", "YES", "更新系APIの正常終了メッセージ", "POST系"],
        ["DTO-016", "ErrorResponse", "Response", "error", "String", "YES", "業務エラーメッセージ", "ExceptionHandler"],
    ])
    add_table_after(ws, ["状態ID", "分類", "値", "表示意味", "設定箇所", "遷移・制約", "備考", ""], [
        ["STAT-001", "payment_status", "unpaid", "未払い", "予約登録初期値 / 支払更新", "paid へ変更可", "未指定時の初期値", ""],
        ["STAT-002", "payment_status", "paid", "支払済み", "支払更新", "unpaid へ戻し可", "許可値以外は拒否", ""],
        ["STAT-003", "reservation_status", "booked", "予約中", "予約登録 / 状態更新", "checked_out / cancelled へ変更可", "客室は reserved / cleaned", ""],
        ["STAT-004", "reservation_status", "checked_out", "チェックアウト済み", "自動同期 / 状態更新", "客室は vacant / needs_cleaning", "checkout_reservations view 対象", ""],
        ["STAT-005", "reservation_status", "cancelled", "取消済み", "予約取消", "客室は vacant / cleaned", "取消一覧対象", ""],
        ["STAT-006", "occupancy_status", "vacant", "空室", "客室登録 / 取消 / チェックアウト", "reserved / occupied へ変更可", "予約候補は vacant かつ cleaned", ""],
        ["STAT-007", "occupancy_status", "reserved", "予約済み", "予約登録 / 状態更新", "vacant へ戻し可", "予約成立後に設定", ""],
        ["STAT-008", "occupancy_status", "occupied", "宿泊中", "客室状態更新", "vacant へ変更可", "手動運用値", ""],
        ["STAT-009", "cleaning_status", "cleaned", "清掃済み", "客室登録 / 清掃更新 / 取消", "needs_cleaning へ変更可", "予約可能条件", ""],
        ["STAT-010", "cleaning_status", "needs_cleaning", "清掃待ち", "チェックアウト同期 / 清掃更新", "cleaned へ変更可", "予約不可条件", ""],
    ])


def write_non_functional(ws):
    setup_sheet(ws, "非機能要件", 5)
    add_table(ws, 6, ["要件ID", "分類", "要件", "基準", "優先度"], [
        ["NF-001", "性能", "一覧表示", "通常データ量で3秒以内に初期表示する。", "高"],
        ["NF-002", "操作性", "業務画面密度", "予約、客室、料金一覧は横スクロール前提で列崩れを抑止する。", "高"],
        ["NF-003", "セキュリティ", "ソース公開抑止", "Java、SQL、設定ファイル、.git 等のパスを denyAll にする。", "高"],
        ["NF-004", "セキュリティ", "CSRF", "CookieCsrfTokenRepository を使用し、React POST でも CSRF を送信する。", "高"],
        ["NF-005", "セキュリティ", "ブラウザ保護", "CSP、Referrer-Policy、X-Robots-Tag、Permissions-Policy を付与する。", "中"],
        ["NF-006", "保守性", "分層", "Controller/API、Service、DAO/Mapper、Domain、Config を分離する。", "高"],
        ["NF-007", "テスト", "JUnit5", "単体テストは JUnit Platform で実行し、業務制約を検証する。", "高"],
    ])
    add_table_after(ws, ["基本非機能ID", "分類", "実装箇所", "設定・実装内容", "確認観点"], [
        ["BNF-001", "URL公開制御", "SecurityConfig.SOURCE_LIKE_PATHS", "*.java、*.class、*.xml、*.yml、*.sql、*.map、/.git/**、/.env、pom.xml 等を denyAll", "ソースや設定ファイルの直接取得が拒否されること。"],
        ["BNF-002", "URL公開制御", "SecurityConfig.PUBLIC_PAGE_PATHS", "/, /dashboard, /rooms/**, /reservations/**, /prices/**, /api/**, /js/**, /styles/** を permitAll", "業務画面と必要静的リソースのみアクセスできること。"],
        ["BNF-003", "CSRF", "SecurityConfig.securityFilterChain", "CookieCsrfTokenRepository.withHttpOnlyFalse()", "React POST が CSRF token を送信できること。"],
        ["BNF-004", "CSP", "SecurityConfig.CONTENT_SECURITY_POLICY", "default/script/style/img/font/connect/form/frame/base の各 directive を self 中心に制限", "外部送信・外部埋め込みが抑止されること。"],
        ["BNF-005", "ブラウザ保護", "SecurityConfig.headers", "No Referrer、X-Robots-Tag、X-Permitted-Cross-Domain-Policies、Permissions-Policy、FrameOptions、ContentTypeOptions", "検索エンジン・外部機能・クリックジャッキング等を抑止すること。"],
        ["BNF-006", "並行制御", "RoomMapper.findByIdForUpdate", "予約登録時に対象客室を FOR UPDATE で取得", "同一客室の同時予約競合を抑止できること。"],
        ["BNF-007", "性能", "ReservationService.safePageSize", "ページサイズを1-100に制限", "過大 limit によるDB負荷を抑止すること。"],
        ["BNF-008", "監査性", "DB updated_at / created_at", "登録・更新日時を各主要テーブルに保持", "一覧並び順と運用確認に利用できること。"],
    ])


def write_tests(ws):
    setup_sheet(ws, "テスト仕様", 11)
    add_table(ws, 6, ["要望ID", "分類", "テスト要望", "テスト目標", "対象範囲", "完了条件", "優先度", "対応要件", "備考", "", ""], [
        ["TREQ-001", "予約", "予約登録の正常系と入力拒否条件を確認する。", "過去日、期間逆転、人数上限、代表者形式、同行者、重複予約、客室状態の各制約が仕様どおり動作すること。", "ReservationServiceTest / ReservationServiceLocalDbTest", "正常登録と主要エラーが自動テストで確認済み。", "高", "REQ-001〜REQ-006", "単体テスト対象", "", ""],
        ["TREQ-002", "客室", "客室登録、再有効化、状態更新、論理削除を確認する。", "客室マスタの必須入力、許可値、同一番号、削除済み再利用、状態更新が仕様どおり動作すること。", "RoomServiceLocalDbTest / RoomControllerTest", "Service と Controller の主要操作が確認済み。", "高", "REQ-007〜REQ-009", "単体テスト対象", "", ""],
        ["TREQ-003", "料金", "料金ルール登録と削除を確認する。", "期間重複、負数料金、初期値、一括削除の入力正規化が仕様どおり動作すること。", "RoomPriceRuleServiceTest / RoomPriceRuleServiceLocalDbTest / PriceRuleControllerTest", "登録・削除・エラーが確認済み。", "高", "REQ-010", "単体テスト対象", "", ""],
        ["TREQ-004", "API", "React 用 JSON API のレスポンスとエラー変換を確認する。", "Dashboard、Rooms、Prices、Reservations、登録 API、業務エラー JSON、CSRF 拒否が仕様どおり動作すること。", "ApiControllerTest", "HTTP status、JSON key、message/error が確認済み。", "高", "BASE-001 / API-001〜API-016", "MockMvc", "", ""],
        ["TREQ-005", "画面Controller", "MVC Controller の画面返却、redirect、flash message を確認する。", "React shell 返却、POST 成功時 message、失敗時 error、CSRF 拒否が仕様どおり動作すること。", "DashboardControllerTest / ReservationControllerTest / RoomControllerTest / PriceRuleControllerTest", "主要 URL と POST が確認済み。", "中", "SCR-001〜SCR-004", "MockMvc", "", ""],
        ["TREQ-006", "セキュリティ", "ソース・設定ファイル公開抑止と CSRF を確認する。", "SOURCE_LIKE_PATHS が拒否され、POST は CSRF token なしで拒否されること。", "DashboardControllerTest / Controller tests", "拒否ステータスが確認済み。", "高", "BNF-001 / BNF-003", "Spring Security Test", "", ""],
        ["TREQ-007", "予約・監査", "端末のローカル日時変更による不正な過去日予約を防止する。", "予約可否判定に端末ローカル日付ではなくシステム日付を使用し、システム日付より前のチェックイン日を拒否すること。", "ReservationService.currentDate / ReservationServiceTest", "currentDate が Mapper 由来のシステム日付を返し、ローカル日付が早くても過去日予約が拒否される。", "高", "REQ-003 / RULE-002 / SQL-001", "不正利用防止", "", ""],
        ["TREQ-008", "テスト運用", "コンソールで各テストケースを識別しやすくする。", "全 @Test に test_XX 番号、テストケース注解、DisplayName、System.out.print 文字列結果を付与すること。", "全 JUnit5 テスト / TestResultLogger", "@Test 85件、DisplayName test_XX 85件、テストケース名 test_XX 85件が一致する。", "中", "テスト証跡", "VS Code / Maven console", "", ""],
    ])
    add_table_after(ws, ["テストID", "対象", "テスト要望", "テスト目標", "前提条件", "入力データ / 操作", "期待結果", "判定基準", "対応仕様", "結果", "備考"], [
        ["UT-001", "ReservationService.create", "正常な予約を登録したい。", "予約本体、同行者、予約番号、金額、客室状態が一括で確定すること。", "空室・清掃済み・有効客室、重複予約なし、料金ルールあり", "代表者、宿泊日、2名、同行者1名で create を実行", "予約が insert され、同行者が insert され、客室が reserved になる。", "Mapper insert と updateStatuses が期待引数で呼ばれる。", "REQ-001 / REQ-002 / RULE-008", "実施済", "ReservationServiceTest"],
        ["UT-002", "ReservationService.create", "過去日の予約を拒否したい。", "業務日付より前のチェックイン日を登録できないこと。", "DB currentDate が取得できる", "過去 checkInDate で create を実行", "「チェックイン日は本日以降」エラーになる。", "IllegalArgumentException の message が一致する。", "REQ-003 / RULE-002", "実施済", "ReservationServiceTest"],
        ["UT-003", "ReservationService.create", "宿泊人数の上限を守りたい。", "11名以上の予約を登録できないこと。", "予約入力あり", "guestCount=11 で create を実行", "「宿泊人数は10名以下」エラーになる。", "予約 insert が呼ばれない。", "REQ-002 / RULE-001", "実施済", "ReservationServiceTest"],
        ["UT-004", "ReservationService.create", "代表者フリガナ形式を確認したい。", "全角カタカナ以外を拒否すること。", "予約入力あり", "guestKana に英数字を指定して create", "フリガナ形式エラーになる。", "IllegalArgumentException の message が一致する。", "RULE-003", "実施済", "ReservationServiceTest"],
        ["UT-005", "ReservationService.create", "電話番号形式を確認したい。", "000-0000-0000 以外の電話番号を拒否すること。", "連絡先保持あり", "不正 guestPhone で create", "電話番号形式エラーになる。", "予約 insert が呼ばれない。", "RULE-003", "実施済", "ReservationServiceTest"],
        ["UT-006", "ReservationService.create", "メール形式を確認したい。", "メール形式でない文字列を拒否すること。", "連絡先保持あり", "不正 guestEmail で create", "メール形式エラーになる。", "予約 insert が呼ばれない。", "RULE-003", "実施済", "ReservationServiceTest"],
        ["UT-007", "ReservationService.create", "連絡先非保持予約を登録したい。", "電話・メール非保持選択時、電話とメールを NULL 保存すること。", "noPhoneInfo=true かつ noEmailInfo=true", "電話・メール入力なしで create", "guestPhone と guestEmail が NULL になる。", "insert 対象 Reservation の該当値が null。", "REQ-001 / RULE-004", "実施済", "ReservationServiceTest"],
        ["UT-008", "ReservationService.create", "重複予約を防ぎたい。", "同一客室で宿泊期間が重なる booked 予約を拒否すること。", "countOverlapping が1以上", "同一 roomId、交差期間で create", "指定期間予約済みエラーになる。", "予約 insert と同行者 insert が呼ばれない。", "REQ-004 / RULE-006 / SQL-007", "実施済", "ReservationServiceTest"],
        ["UT-009", "ReservationService.create", "予約不可客室を選ばせない。", "空室以外の客室を拒否すること。", "対象 Room の occupancyStatus が reserved", "create を実行", "空室の部屋のみ予約可能エラーになる。", "予約 insert が呼ばれない。", "RULE-005", "実施済", "ReservationServiceTest"],
        ["UT-010", "ReservationService.create", "未清掃客室を選ばせない。", "cleaned 以外の客室を拒否すること。", "対象 Room の cleaningStatus が needs_cleaning", "create を実行", "清掃済みの部屋のみ予約可能エラーになる。", "予約 insert が呼ばれない。", "REQ-009 / RULE-005", "実施済", "ReservationServiceTest"],
        ["UT-011", "ReservationService.create", "同行者連絡先も検証したい。", "同行者電話番号の形式不正を拒否すること。", "2名以上の予約", "companionPhones に不正値を指定", "電話番号形式エラーになる。", "同行者 insert が呼ばれない。", "REQ-002 / RULE-007", "実施済", "ReservationServiceTest"],
        ["UT-012", "ReservationService.syncDueCheckouts", "期限到来予約を自動同期したい。", "宿泊終了日に達した booked 予約が checked_out になり、客室が清掃待ちになること。", "findDueCheckouts が予約を返す", "syncDueCheckouts を実行", "予約 markCheckedOut、客室 vacant / needs_cleaning 更新。", "対象件数分 Mapper が呼ばれる。", "REQ-011 / LOGIC-009", "仕様化済", "追加観点"],
        ["UT-013", "ReservationService.updateReservationStatus", "予約状態変更時に客室も同期したい。", "booked 復帰、checked_out 変更で客室状態が仕様どおりになること。", "対象予約あり", "reservationStatus を booked または checked_out に更新", "booked は reserved/cleaned、checked_out は他予約なしなら vacant/needs_cleaning、他予約ありなら reserved/cleaned。", "ReservationMapper と RoomMapper の更新内容が一致する。", "LOGIC-010 / STAT-003", "仕様化済", "追加観点"],
        ["UT-014", "ReservationService.updateCheckoutCleaningStatus", "清掃状態だけを更新したい。", "checked_out 予約だけ清掃状態更新を受け付け、客室は vacant のまま更新されること。", "対象予約あり", "cleaned または needs_cleaning で更新", "checked_out なら RoomMapper.updateStatuses(roomId, vacant, cleaningStatus) が実行される。booked ならエラー。", "不正値はエラー、正値は更新。", "REQ-009 / LOGIC-011", "仕様化済", "追加観点"],
        ["UT-015", "RoomService.create", "客室を正常登録したい。", "必須項目と初期値が整った客室が登録されること。", "同一 roomNumber なし", "roomNumber、roomName、capacity を指定して create", "roomType、料金、状態、active が補完され insert される。", "RoomMapper.insert が1回呼ばれる。", "REQ-007 / RULE-012", "仕様化済", "追加観点"],
        ["UT-016", "RoomService.create", "削除済み客室を再有効化したい。", "同一番号の inactive 客室は新規登録ではなく reactivate されること。", "findByRoomNumberIncludingInactive が active=false を返す", "同じ roomNumber で create", "RoomMapper.reactivate が実行される。", "insert が呼ばれない。", "REQ-008 / SQL-009", "仕様化済", "追加観点"],
        ["UT-017", "RoomService.create", "有効客室の重複番号を拒否したい。", "active=true の同一 roomNumber を登録できないこと。", "同一 roomNumber の有効客室あり", "create を実行", "部屋番号重複エラーになる。", "insert / reactivate が呼ばれない。", "RULE-012", "仕様化済", "追加観点"],
        ["UT-018", "RoomService.updateStatuses", "客室状態の許可値を守りたい。", "不正 occupancyStatus または cleaningStatus を拒否すること。", "客室あり", "不正状態値で updateStatuses", "状態エラーになる。", "RoomMapper.updateStatuses が呼ばれない。", "RULE-013 / STAT-006", "実施済", "RoomService 系テスト"],
        ["UT-019", "RoomPriceRuleService.create", "料金ルールを正常登録したい。", "対象客室、名称、期間、料金が妥当なら登録できること。", "重複期間なし", "RoomPriceRule で create", "priority 未指定は10、active 未指定は true で insert される。", "RoomPriceRuleMapper.insert が呼ばれる。", "REQ-010 / RULE-014", "実施済", "RoomPriceRuleServiceTest"],
        ["UT-020", "RoomPriceRuleService.create", "期間重複を拒否したい。", "同一客室で期間が重なる active ルールを登録できないこと。", "countOverlapping が1以上", "重複期間で create", "指定期間設定済みエラーになる。", "insert が呼ばれない。", "REQ-010 / SQL-010", "実施済", "RoomPriceRuleServiceTest"],
        ["UT-021", "RoomPriceRuleService.create", "負数料金を拒否したい。", "pricePerPerson が0円未満の場合は登録できないこと。", "料金ルール入力あり", "price=-1 で create", "料金は0円以上エラーになる。", "insert が呼ばれない。", "RULE-014", "実施済", "RoomPriceRuleServiceTest"],
        ["UT-022", "RoomPriceRuleService.deleteByIds", "一括削除入力を正規化したい。", "NULL と重複を除外し、残った ID だけ削除すること。", "ids に null と重複あり", "deleteByIds を実行", "distinct 後の ID で Mapper.deleteByIds が実行される。", "空配列はエラー、削除0件もエラー。", "RULE-015", "実施済", "RoomPriceRuleServiceTest"],
        ["UT-023", "ApiController.dashboard", "ダッシュボード API を確認したい。", "集計値と直近予約ページが JSON で返ること。", "Service mock 設定済み", "GET /api/dashboard", "roomCount、vacantCount、bookedCount、recentReservations を返す。", "HTTP 200 と JSON path が一致する。", "API-001 / DTO-001", "実施済", "ApiControllerTest"],
        ["UT-024", "ApiController.reservations", "予約 API を確認したい。", "予約中・取消・チェックアウト・客室・業務日付が JSON で返ること。", "Service mock 設定済み", "GET /api/reservations", "ReservationsResponse の各 key を返す。", "HTTP 200 と JSON path が一致する。", "API-010 / DTO-006", "実施済", "ApiControllerTest"],
        ["UT-025", "ApiController", "業務エラーを JSON 化したい。", "Service の IllegalArgumentException が HTTP 400 と error JSON になること。", "Service が例外を throw", "POST /api/rooms", "{error: メッセージ} を返す。", "HTTP 400 と error key が一致する。", "API-016 / DTO-016", "実施済", "ApiControllerTest"],
        ["UT-026", "Controller", "CSRF 防御を確認したい。", "CSRF token なしの POST が拒否されること。", "Spring Security 有効", "POST API または画面 POST を token なしで実行", "403 Forbidden になる。", "更新処理が実行されない。", "BNF-003 / TREQ-006", "実施済", "Controller tests"],
        ["UT-027", "DashboardController", "React shell を返したい。", "/ と /dashboard が app view を返すこと。", "アプリ起動", "GET / または /dashboard", "viewName=app になる。", "HTTP 200 と viewName が一致する。", "SCR-001", "実施済", "DashboardControllerTest"],
        ["UT-028", "SecurityConfig", "ソース公開を拒否したい。", "SOURCE_LIKE_PATHS に該当する URL が拒否されること。", "Spring Security 有効", "GET /pom.xml 等", "アクセス拒否になる。", "HTTP 403 または拒否ステータス。", "BNF-001", "実施済", "DashboardControllerTest"],
        ["UT-029", "ReservationController", "予約画面 POST 成功を確認したい。", "予約登録成功時に message を flash へ入れて redirect すること。", "Service mock 正常", "POST /reservations", "/reservations へ redirect し成功 message を持つ。", "3xx redirect と flash message が一致する。", "SCR-002", "実施済", "ReservationControllerTest"],
        ["UT-030", "RoomController", "客室登録エラー表示を確認したい。", "Service の業務エラーを flash error に変換すること。", "RoomService.create が IllegalArgumentException", "POST /rooms", "/rooms へ redirect し error を持つ。", "3xx redirect と flash error が一致する。", "SCR-003", "実施済", "RoomControllerTest"],
        ["UT-031", "PriceRuleController", "旧削除 URL 互換を確認したい。", "/prices/{id} と /prices/{id}/delete の両方で削除できること。", "Service mock 正常", "POST /prices/{id}", "料金ルール削除 message 付きで redirect する。", "priceRuleService.delete(id) が呼ばれる。", "SCR-004", "実施済", "PriceRuleControllerTest"],
        ["UT-032", "ReservationService.currentDate", "システム日付を使用したい。", "業務日付取得が端末ローカル環境日付ではなく Mapper / DB 基準日付に委譲されること。", "reservationMapper.currentDate() が 2099-01-01 を返す。JVM default timezone を Pacific/Pago_Pago に変更する。", "reservationService.currentDate() を実行", "戻り値が 2099-01-01 になり、LocalDate.now() と一致しない。", "Mapper currentDate が呼ばれ、ローカル環境日付に依存しない。", "REQ-003 / SQL-001 / TREQ-007", "実施済", "ReservationServiceTest test_01"],
        ["UT-033", "ReservationService.create", "ローカル日時改ざんによる過去日予約を拒否したい。", "端末ローカル日付上は未来でも、システム日付より前のチェックイン日は登録できないこと。", "localEnvironmentDate=現在日、systemDate=現在日+30日、checkInDate=現在日+1日", "create を実行し、System.out.print で systemInput.currentDate / localEnvironmentDate / input.checkInDate / input.checkOutDate を出力", "「チェックイン日は本日以降を選択してください。」エラーになり、客室検索へ進まない。", "IllegalArgumentException message と roomMapper.findByIdForUpdate 未呼び出しを確認する。", "REQ-003 / RULE-002 / TREQ-007", "実施済", "ReservationServiceTest test_02"],
        ["UT-034", "全 JUnit5 テスト", "テスト証跡を文字列で確認したい。", "全テストケースに test_XX の識別子、Javadoc 形式のテスト注解、DisplayName、文字列結果が出力されること。", "@LoggedTest と TestResultLogger が全テストクラスに適用されている。", "mvn test または VS Code task mvn test を実行", "テスト開始、テスト内容、テストコード注解、テスト文字列結果、テストクラス文字列結果が System.out.print で出力される。", "@Test 85件、DisplayName test_XX 85件、テストケース名 test_XX 85件、mvn test 成功。", "TREQ-008", "実施済", "TestResultLogger / 全 Test"],
        ["UT-035", "ReservationService.deleteCancelled", "取消済み予約だけを完全削除したい。", "cancelled 以外の予約を削除できず、cancelled 予約だけ deleteCancelled SQL を実行すること。", "対象予約あり", "deleteCancelled(id) を実行", "cancelled は削除され、booked は「取消済み予約のみ削除できます。」エラーになる。", "deleteCancelled の呼び出し有無と例外 message が一致する。", "REQ-006A / API-013A / SQL-002A", "実施済", "ReservationServiceTest / ReservationServiceLocalDbTest"],
        ["UT-036", "RoomService.restore / deletePermanently", "削除済み客室を復元・完全削除したい。", "inactive 客室だけ復元でき、予約履歴のない inactive 客室だけ完全削除できること。", "削除済み客室、予約履歴件数", "restore または deletePermanently を実行", "復元は active に戻り、予約履歴ありの完全削除は拒否される。", "RoomMapper.restore / deletePermanently の呼び出しと例外 message を確認する。", "REQ-008A / API-005A / API-005B", "実施済", "RoomServiceTest / RoomServiceLocalDbTest"],
        ["UT-037", "ReservationService.deleteCheckedOut", "チェックアウト済み予約だけを完全削除したい。", "checked_out 以外の予約を削除できず、checked_out 予約だけ deleteCheckedOut SQL を実行すること。", "対象予約あり", "deleteCheckedOut(id) を実行", "checked_out は削除され、booked は「チェックアウト済み予約のみ削除できます。」エラーになる。", "deleteCheckedOut の呼び出し有無と例外 message が一致する。", "REQ-006B / API-014A / SQL-002C", "実施済", "ReservationServiceTest / ReservationServiceLocalDbTest"],
    ])


def write_schedule(ws):
    setup_sheet(ws, "スケジュール", 7)
    add_table(ws, 6, ["工程", "開始日", "終了日", "期間(日)", "担当", "成果物", "状態"], [
        ["要件定義", "2026-07-01", "2026-07-01", 1, AUTHOR, "要件定義", "完了"],
        ["基本設計", "2026-07-01", "2026-07-02", 2, AUTHOR, "基本設計", "完了"],
        ["詳細設計", "2026-07-02", "2026-07-06", 5, AUTHOR, "Java 分層仕様", "完了"],
        ["開発", "2026-07-02", "2026-07-06", 5, AUTHOR, "Spring Boot + React 実装", "完了"],
        ["単体テスト", "2026-07-06", "2026-07-06", 1, AUTHOR, "JUnit5 テスト", "完了"],
    ])


def write_references(ws):
    setup_sheet(ws, "参考資料", 4)
    add_table(ws, 6, ["優先順", "資料名", "内容", "URL"], [
        [1, "民宿管理システム.xlsx", "本プロジェクト向けに現行コードへ合わせて作成した仕様書。", "docs/仕様書/民宿管理システム.xlsx"],
        [2, "src/main/java/com/example/minshuku", "Application、Config、Controller、Service、Mapper、Domain の Java 実装。", "src/main/java/com/example/minshuku"],
        [3, "src/main/resources/mapper", "MyBatis XML。Reservation、Room、RoomPriceRule、ReservationGuest のSQL仕様。", "src/main/resources/mapper"],
        [4, "src/main/resources/db/schema.sql", "PostgreSQL テーブル、制約、索引、予約番号シーケンス、チェックアウト view。", "src/main/resources/db/schema.sql"],
        [5, "src/main/frontend/App.jsx", "React 画面と API 呼び出し。", "src/main/frontend/App.jsx"],
        [6, "src/test/java", "JUnit5 単体テストおよび controller/service テスト。", "src/test/java/com/example/minshuku"],
        [7, "src/main/resources/application.yml", "context-path、DB接続、MyBatis、静的リソース、セキュリティ除外設定。", "src/main/resources/application.yml"],
        [8, "src/main/java/com/example/minshuku/config/SecurityConfig.java", "公開URL、拒否URL、CSRF、CSP、セキュリティヘッダーの基本仕様。", "src/main/java/com/example/minshuku/config/SecurityConfig.java"],
        [9, "src/main/java/com/example/minshuku/controller", "MVC Controller と React JSON API の入口仕様。", "src/main/java/com/example/minshuku/controller"],
        [10, "src/main/java/com/example/minshuku/domain", "予約、同行者、客室、料金ルールの Domain 項目仕様。", "src/main/java/com/example/minshuku/domain"],
        [11, "src/test/java/com/example/minshuku/support/TestResultLogger.java", "全テストケースの System.out.print 文字列結果、テスト内容、テストコード注解、クラス集計結果の出力仕様。", "src/test/java/com/example/minshuku/support/TestResultLogger.java"],
        [12, "src/test/java/com/example/minshuku/service/ReservationServiceTest.java", "システム日付基準の予約可否判定、不正利用防止テスト、test_XX DisplayName とテストケース注解。", "src/test/java/com/example/minshuku/service/ReservationServiceTest.java"],
        [13, "日本民宿予約管理システム_仕様書_ヘッダー追加.xlsx", "仕様書フォーマット参考。作成・更新情報ヘッダー、シート構成、表形式を参照。", "docs/日本民宿予約管理システム_仕様書_ヘッダー追加.xlsx"],
    ])


def add_code_section(ws, file_name: str, summary: str, rows: list[list[object]]):
    start_row = ws.max_row + 2
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
    set_cell(ws.cell(start_row, 1), file_name, font=font_title, fill=fill_title, alignment=align_left)

    set_cell(ws.cell(start_row + 1, 1), "処理内容", font=font_header, fill=fill_header, alignment=align_center)
    ws.merge_cells(start_row=start_row + 1, start_column=2, end_row=start_row + 1, end_column=4)
    set_cell(ws.cell(start_row + 1, 2), summary, fill=fill_note)

    add_table(ws, start_row + 3, ["No", "コード / 要素", "入出力・状態変化"], rows)
    ws.row_dimensions[start_row].height = 26
    ws.row_dimensions[start_row + 1].height = 48


def write_code_trace(ws):
    setup_sheet(ws, "コード行別処理仕様", 4)
    add_table(ws, 6, ["記法", "意味", "例", ""], [
        ["name = value", "定数・変数・初期値を示す。", "paymentStatus = unpaid", ""],
        ["method(args)", "メソッド呼び出しと引数を示す。", "reservationService.create(request)", ""],
        ["if (condition) -> result", "条件分岐と結果を示す。", "if (room == null) -> error", ""],
        ["A -> B -> C", "処理順序または状態遷移を示す。", "validate -> insert -> update room", ""],
        ["target = permitAll / denyAll", "URL や権限の許可・拒否を示す。", "/api/** = permitAll", ""],
    ])

    sections = [
        (
            "MinshukuManagementApplication.java",
            "`MinshukuManagementApplication` は Spring Boot の起動クラス。業務処理は持たず、`main(args)` から `SpringApplication.run(...)` を実行してアプリケーション全体を起動する。",
            [
                [1, "package = com.example.minshuku", "Java パッケージを `com.example.minshuku` に設定する。"],
                [2, "import = SpringApplication", "Spring Boot 起動 API を使用可能にする。"],
                [3, "import = SpringBootApplication", "Spring Boot アプリケーション宣言用 annotation を使用可能にする。"],
                [4, "@SpringBootApplication", "自動設定 + コンポーネントスキャン + 設定読み込みの起点になる。"],
                [5, "class = MinshukuManagementApplication", "アプリケーション起動専用クラスとして定義する。"],
                [6, "main(args)", "Java アプリケーションの起動入口。起動引数 `args` を受け取る。"],
                [7, "SpringApplication.run(MinshukuManagementApplication.class, args)", "Spring Boot を起動し、Controller / Service / Mapper / Config を Spring 管理対象にする。"],
            ],
        ),
        (
            "SecurityConfig.java",
            "`SecurityConfig` は社内向け管理画面の HTTP セキュリティ設定クラス。`SOURCE_LIKE_PATHS = denyAll`、`PUBLIC_PAGE_PATHS = permitAll` として URL 公開範囲を分離し、CSRF / CSP / security header を一元設定する。",
            [
                [1, "@Configuration", "Spring 設定クラスとして登録される。"],
                [2, "SOURCE_LIKE_PATHS = { *.java, *.xml, *.yml, *.sql, .git, pom.xml, ... }", "ソース・設定・SQL・ビルドファイル系 URL を拒否対象にする。"],
                [3, "PUBLIC_PAGE_PATHS = { /, /dashboard, /rooms/**, /reservations/**, /prices/**, /api/**, /js/**, /styles/** }", "業務画面、API、静的リソースを公開対象にする。"],
                [4, "CONTENT_SECURITY_POLICY = default-src('self') + connect-src('self') + frame-ancestors('none')", "外部 script / 外部通信 / iframe 埋め込みを制限する CSP 文字列になる。"],
                [5, "sourceLikeMatchers()", "`SOURCE_LIKE_PATHS` を `RequestMatcher[]` に変換する。"],
                [6, "for (i = 0; i < SOURCE_LIKE_PATHS.length; i++)", "拒否対象パスを1件ずつ `AntPathRequestMatcher` に変換する。"],
                [7, "requestMatchers(sourceLikeMatchers()).denyAll()", "ソース類似パスへのアクセスを拒否する。"],
                [8, "requestMatchers(PUBLIC_PAGE_PATHS).permitAll()", "業務画面と API へのアクセスを許可する。"],
                [9, "anyRequest().denyAll()", "定義外 URL はすべて拒否する。"],
                [10, "httpBasic() / formLogin() / logout() = disable", "標準認証画面、Basic 認証、logout を無効化する。"],
                [11, "csrf = CookieCsrfTokenRepository.withHttpOnlyFalse()", "React から CSRF token を扱えるようにする。"],
                [12, "headers = CSP + ReferrerPolicy + X-Robots-Tag + PermissionsPolicy", "ブラウザ側の保護ヘッダーを付与する。"],
                [13, "return http.build()", "完成した `SecurityFilterChain` を Spring Security に登録する。"],
            ],
        ),
        (
            "DashboardController.java",
            "`DashboardController` はダッシュボード URL の画面入口。画面データは API 側で取得するため、この Controller は React shell の `app` を返すだけにする。",
            [
                [1, "@Controller", "Spring MVC Controller として登録される。"],
                [2, "@GetMapping({ '/', '/dashboard' })", "ルート URL とダッシュボード URL を同じ画面入口にする。"],
                [3, "dashboard() -> 'app'", "React shell テンプレート `app.html` を返す。"],
            ],
        ),
        (
            "ApiController.java",
            "`ApiController` は React フロントエンド用 JSON API。画面初期表示データ、登録・更新処理、業務エラー JSON 変換、DTO record をまとめて提供する。",
            [
                [1, "@RestController + @RequestMapping('/api')", "`/api/**` 配下の JSON API Controller として登録される。"],
                [2, "PAGE_SIZE = 5", "一覧 API の初期表示件数を5件に固定する。"],
                [3, "constructor(roomService, reservationService, priceRuleService)", "客室・予約・料金 Service を DI で受け取る。"],
                [4, "dashboard() -> syncDueCheckouts()", "集計前に期限到来チェックアウトを同期する。"],
                [5, "dashboard() -> DashboardResponse(roomCount, vacantCount, bookedCount, recentReservations)", "ダッシュボード用集計値と直近予約一覧を返す。"],
                [6, "rooms() -> RoomsResponse(findAll(), findInactive(), findBookable())", "有効客室、削除済み客室、予約可能客室を返す。"],
                [7, "createRoom(room) -> roomService.create(room)", "JSON の客室情報を登録 Service へ渡す。"],
                [8, "updateRoomStatuses(id, request) -> roomService.updateStatuses(id, occupancyStatus, cleaningStatus)", "宿泊状態と清掃状態を同時更新する。"],
                [9, "deleteRoom(id) -> roomService.delete(id)", "客室を論理削除する。"],
                [10, "prices() -> PricesResponse(findAllWithRoom(), findActive())", "料金ルール一覧と登録対象客室を返す。"],
                [11, "createPriceRule(rule) -> priceRuleService.create(rule)", "期間別料金ルールを登録する。"],
                [12, "deletePriceRule(id) -> priceRuleService.delete(id)", "料金ルールを1件削除する。"],
                [13, "deleteSelectedPriceRules(ids) -> priceRuleService.deleteByIds(ids)", "選択された料金ルールを一括削除する。"],
                [14, "reservations() -> syncDueCheckouts()", "予約一覧取得前にチェックアウト対象を同期する。"],
                [15, "reservations() -> ReservationsResponse(reservations, cancelled, checkedOut, rooms, today)", "予約管理画面の初期表示データをまとめて返す。"],
                [16, "createReservation(request) -> reservationService.create(...)", "予約本体と同行者配列を Service へ渡す。"],
                [17, "noContactInfo = noPhoneInfo && noEmailInfo", "電話・メールの両方が非保持指定された場合だけ連絡先非保持として扱う。"],
                [18, "updatePayment(id, paymentStatus)", "支払状態を `unpaid / paid` の範囲で更新する。"],
                [19, "cancelReservation(id)", "予約を取消し、客室状態を空室・清掃済みに戻す。"],
                [20, "updateReservationStatus(id, reservationStatus)", "予約状態を更新し、必要に応じて客室状態も同期する。"],
                [21, "updateCleaning(id, cleaningStatus)", "チェックアウト後の客室清掃状態を更新する。"],
                [22, "@ExceptionHandler(IllegalArgumentException) -> ErrorResponse(error)", "Service 層の業務エラーを HTTP 400 の JSON に変換する。"],
                [23, "firstPage(items, totalCount) -> PageResponse(items, 1, totalPages)", "初期表示用ページ情報を `page = 1` で統一する。"],
                [24, "totalPages = max(1, ceil(totalCount / PAGE_SIZE))", "0件でも画面側のページ数を1として扱う。"],
            ],
        ),
        (
            "ReservationService.java",
            "`ReservationService` は予約業務の中核。予約登録、入力検証、客室ロック、重複予約防止、予約番号発番、料金計算、同行者保存、予約状態と客室状態の同期を1つの業務単位として扱う。",
            [
                [1, "MAX_PAGE_SIZE = 100", "一覧 API の過大ページサイズを抑止する上限値。"],
                [2, "KANA_PATTERN = ^[ァ-ヶー\\s]+$ / PHONE_PATTERN = ^\\d{3}-\\d{4}-\\d{4}$ / EMAIL_PATTERN = ^[^\\s@]+@[^\\s@]+\\.[^\\s@]+$", "代表者・同行者の入力形式チェックに使用する。正規表現は Java コードと同じ内容を記載する。"],
                [3, "PAYMENT_STATUSES = { unpaid, paid }", "支払状態の許可値を定義する。"],
                [4, "RESERVATION_STATUSES = { booked, checked_out, cancelled }", "予約状態の許可値を定義する。"],
                [5, "CLEANING_STATUSES = { needs_cleaning, cleaned }", "清掃状態の許可値を定義する。"],
                [6, "constructor(reservationMapper, reservationGuestMapper, roomMapper, priceRuleMapper)", "予約、同行者、客室、料金 Mapper を DI で受け取る。"],
                [7, "currentDate() -> reservationMapper.currentDate()", "DB 基準の業務日付を取得する。"],
                [8, "findRecentPage(page, pageSize) -> safePageSize() + pageOffset()", "予約中一覧を安全なページング条件で取得する。"],
                [9, "findCancelledPage(page, pageSize)", "取消済み予約一覧をページング取得する。"],
                [10, "findCheckedOutPage(page, pageSize)", "チェックアウト済み予約一覧をページング取得する。"],
                [11, "syncDueCheckouts() -> findDueCheckouts()", "宿泊終了日に達した予約を抽出する。"],
                [12, "for (dueReservation) -> markCheckedOut(id)", "対象予約を `checked_out` に更新する。"],
                [13, "roomStatus = vacant + needs_cleaning", "チェックアウト済み客室を空室・清掃待ちに戻す。"],
                [14, "create(...) -> validateReservation()", "予約本体の必須、日付、代表者、連絡先、人数を検証する。"],
                [15, "create(...) -> validateCompanions()", "宿泊人数に応じた同行者名の必須入力を検証する。"],
                [16, "create(...) -> validateCompanionContacts()", "同行者フリガナ・電話番号の形式を検証する。"],
                [17, "room = findByIdForUpdate(roomId)", "予約対象客室を行ロック付きで取得し、同時予約を抑止する。"],
                [18, "validateRoomForReservation(reservation, room)", "営業中 + 空室 + 清掃済み + 定員内を確認する。"],
                [19, "validateNoOverlappingReservation(reservation)", "同一客室・宿泊期間交差の booked 予約を拒否する。"],
                [20, "reservationNo = nextReservationNo()", "DB シーケンスから R + 6桁予約番号を採番する。"],
                [21, "reservationStatus = booked", "新規予約を予約中状態にする。"],
                [22, "if (paymentStatus is blank) -> unpaid", "支払状態未入力時は未払いに補完する。"],
                [23, "if (reservationForm is blank) -> 公式", "予約経路未入力時は公式予約に補完する。"],
                [24, "totalAmount = calculateTotalAmount(reservation, room)", "宿泊日ごとの単価を積み上げて合計金額を確定する。"],
                [25, "reservationMapper.insert(reservation)", "予約本体を DB に登録し、予約 ID を確定する。"],
                [26, "saveCompanions(reservationId, guestCount, ...)", "予約 ID に紐づく同行者明細を保存する。"],
                [27, "roomStatus = reserved + currentCleaningStatus", "予約成立後、客室を予約済みにする。"],
                [28, "updatePaymentStatus(id, paymentStatus) -> requireAllowed()", "支払状態が許可値か確認して更新する。"],
                [29, "if (updatePaymentStatus(...) == 0) -> error", "更新対象予約がない場合は業務エラーにする。"],
                [30, "updateReservationStatus(id, reservationStatus) -> requireAllowed()", "予約状態が許可値か確認する。"],
                [31, "reservation = findById(id)", "状態更新対象の予約を取得する。"],
                [32, "if (reservation == null) -> error", "対象予約が存在しない場合は業務エラーにする。"],
                [33, "if (reservationStatus == booked) -> roomStatus = reserved + cleaned", "予約中へ戻す場合、客室を予約済み・清掃済みに同期する。"],
                [34, "if (reservationStatus == checked_out) -> roomStatus = vacant + needs_cleaning", "チェックアウト済みへ変更する場合、客室を空室・清掃待ちに同期する。"],
                [35, "updateCheckoutCleaningStatus(id, cleaningStatus)", "チェックアウト済み予約だけ清掃状態を更新する。"],
                [36, "deleteCancelled(id)", "取消済み予約だけを予約一覧から完全削除する。"],
                [37, "updateRoomAfterReservationRelease(reservation, cleaningStatusWhenVacant)", "同室に他の予約中データが残る場合は予約済み状態を維持し、なければ空室へ戻す。"],
                [36, "cancel(id) -> reservationMapper.cancel(id)", "予約を取消済みに更新する。"],
                [37, "if (cancel success) -> roomStatus = vacant + cleaned", "取消後、客室を再販売可能な空室・清掃済みに戻す。"],
                [38, "safePageSize = min(100, max(1, pageSize))", "ページサイズを1〜100に補正する。"],
                [39, "pageOffset = (max(1, page) - 1) * safePageSize", "SQL OFFSET を算出する。"],
                [40, "if (roomId == null) -> error", "予約対象客室未選択を拒否する。"],
                [41, "if (checkInDate == null || checkOutDate == null) -> error", "宿泊日未入力を拒否する。"],
                [42, "if (checkInDate < currentDate()) -> error", "過去チェックイン日を拒否する。"],
                [43, "if (checkInDate >= checkOutDate) -> error", "宿泊期間が成立しない予約を拒否する。"],
                [44, "if (guestName is blank) -> error", "宿泊者名未入力を拒否する。"],
                [45, "if (!noContactInfo) -> validate(phone, email)", "連絡先保持時は電話番号とメール形式を検証する。"],
                [46, "if (noContactInfo) -> guestPhone = null, guestEmail = null", "連絡先非保持時は DB に NULL 保存する。"],
                [47, "if (guestCount < 1 || guestCount > 10) -> error", "宿泊人数の下限・上限を検証する。"],
                [48, "if (room == null || active != true) -> error", "存在しない客室または無効客室を拒否する。"],
                [49, "if (occupancyStatus != vacant) -> error", "空室以外の客室を拒否する。"],
                [50, "if (cleaningStatus != cleaned) -> error", "未清掃客室を拒否する。"],
                [51, "if (guestCount > capacity) -> error", "定員超過予約を拒否する。"],
                [52, "overlaps = countOverlapping(roomId, checkInDate, checkOutDate)", "同一客室・期間交差する予約数を取得する。"],
                [53, "if (overlaps > 0) -> error", "重複予約を拒否する。"],
                [54, "requiredCount = guestCount - 1", "代表者以外に必要な同行者数を算出する。"],
                [55, "if (companionNames size < requiredCount) -> error", "同行者情報不足を拒否する。"],
                [56, "for (i = 0; i < requiredCount; i++) -> validate companion", "同行者名、フリガナ、電話番号を人数分検証する。"],
                [57, "saveCompanions() -> new ReservationGuest() -> insert(guest)", "同行者ごとに明細オブジェクトを作成して保存する。"],
                [58, "calculateTotalAmount() -> total = 0", "料金合計を0円で初期化する。"],
                [59, "while (stayDate < checkOutDate)", "チェックアウト日前日まで日別に料金計算する。"],
                [60, "rule = findBestRule(roomId, stayDate)", "対象日の最優先料金ルールを取得する。"],
                [61, "price = rule == null ? basePrice : rule.price", "料金ルール未設定日は客室基本単価を使う。"],
                [62, "total += price * guestCount", "日別単価に人数を掛けて合計へ加算する。"],
            ],
        ),
        (
            "RoomService.java",
            "`RoomService` は客室マスタの登録、再有効化、宿泊状態・清掃状態更新、論理削除、件数集計を扱う。予約業務と清掃業務が同じ客室状態を参照するため、状態値の許可範囲をここで統一する。",
            [
                [1, "ROOM_TYPES = { washitsu, yoshitsu, suite, family }", "客室タイプの許可値を定義する。"],
                [2, "OCCUPANCY_STATUSES = { vacant, reserved, occupied }", "宿泊状態の許可値を定義する。"],
                [3, "CLEANING_STATUSES = { cleaned, needs_cleaning }", "清掃状態の許可値を定義する。"],
                [4, "findAll() / findInactive() / findActive() / findBookable()", "客室一覧、削除済み一覧、有効客室、予約可能客室を取得する。"],
                [5, "create(room) -> validate required fields", "部屋番号、部屋名、定員の必須条件を検証する。"],
                [6, "if (basePricePerPerson == null) -> 0", "基本料金未入力時は0円に補完する。"],
                [7, "if (roomType is blank) -> washitsu", "部屋タイプ未入力時は和室に補完する。"],
                [8, "requireAllowed(roomType)", "部屋タイプが許可値か検証する。"],
                [9, "if (privateBath == null) -> false", "専用風呂未入力時は false に補完する。"],
                [10, "if (occupancyStatus is blank) -> vacant", "新規客室の宿泊状態を空室に補完する。"],
                [11, "if (cleaningStatus is blank) -> cleaned", "新規客室の清掃状態を清掃済みに補完する。"],
                [12, "requireAllowed(occupancyStatus, cleaningStatus)", "宿泊状態と清掃状態の許可値を検証する。"],
                [13, "if (active == null) -> true", "有効フラグ未入力時は有効に補完する。"],
                [14, "existingRoom = findByRoomNumberIncludingInactive(roomNumber)", "同じ部屋番号の客室を削除済み含めて検索する。"],
                [15, "if (existingRoom.active == true) -> error", "有効客室の部屋番号重複を拒否する。"],
                [16, "if (existingRoom != null) -> reactivate(room) -> return", "削除済み客室がある場合は新規作成せず再有効化する。"],
                [17, "insert(room)", "既存客室がない場合は新規登録する。"],
                [18, "updateStatuses(id, occupancyStatus, cleaningStatus)", "宿泊状態と清掃状態を同時更新する。"],
                [19, "if (updateStatuses(...) == 0) -> error", "対象客室がない場合は業務エラーにする。"],
                [20, "delete(id) -> deactivate(id)", "客室を物理削除せず `active = false` にする。"],
                [21, "countAll() / countVacant()", "ダッシュボード用の有効客室数と空室数を取得する。"],
            ],
        ),
        (
            "RoomPriceRuleService.java",
            "`RoomPriceRuleService` は客室別・期間別料金ルールの登録、重複期間チェック、単一削除、一括削除を扱う。予約金額の根拠になるため、期間不正・負数・重複を登録前に排除する。",
            [
                [1, "findAllWithRoom()", "客室情報付きの料金ルール一覧を取得する。"],
                [2, "create(rule) -> validateRule(rule)", "料金ルールの必須項目、期間、金額、初期値を検証する。"],
                [3, "if (countOverlapping(roomId, startDate, endDate) > 0) -> error", "同一客室で期間が重なる有効料金ルールを拒否する。"],
                [4, "insert(rule)", "検証済み料金ルールを登録する。"],
                [5, "delete(id) -> if (id == null) -> error", "削除対象未選択を拒否する。"],
                [6, "if (delete(id) == 0) -> error", "削除対象が存在しない場合は業務エラーにする。"],
                [7, "deleteByIds(ids) -> if (ids is empty) -> error", "一括削除の未選択を拒否する。"],
                [8, "targetIds = ids.filter(nonNull).distinct()", "NULL と重複を除外して削除対象 ID を正規化する。"],
                [9, "if (targetIds is empty) -> error", "正規化後に削除対象がない場合は業務エラーにする。"],
                [10, "if (deleteByIds(targetIds) == 0) -> error", "一括削除対象が存在しない場合は業務エラーにする。"],
                [11, "if (roomId == null) -> error", "料金対象客室未選択を拒否する。"],
                [12, "if (ruleName is blank) -> error", "料金ルール名未入力を拒否する。"],
                [13, "if (startDate == null || endDate == null) -> error", "開始日・終了日未入力を拒否する。"],
                [14, "if (startDate > endDate) -> error", "開始日が終了日より後の料金ルールを拒否する。"],
                [15, "if (pricePerPerson == null || pricePerPerson < 0) -> error", "料金未入力または負数を拒否する。"],
                [16, "if (priority == null) -> priority = 10", "優先度未入力時は標準値10に補完する。"],
                [17, "if (active == null) -> active = true", "有効フラグ未入力時は有効に補完する。"],
            ],
        ),
        (
            "ReservationController.java",
            "`ReservationController` は MVC 互換の予約画面入口。GET は React shell を返し、POST は Service を実行して flash message / error を設定し、予約画面へ redirect する。",
            [
                [1, "reservations() -> 'app'", "予約管理画面 URL で React shell を返す。"],
                [2, "create(reservation, noPhoneInfo, noEmailInfo, companions)", "フォーム投稿された予約本体と同行者配列を受け取る。"],
                [3, "noContactInfo = noPhoneInfo && noEmailInfo", "電話・メールの両方が非保持指定された場合だけ非保持として扱う。"],
                [4, "reservationService.create(...)", "予約登録 Service を実行する。"],
                [5, "success -> flash.message = '予約を登録しました。'", "登録成功メッセージを画面へ渡す。"],
                [6, "catch IllegalArgumentException -> flash.error = ex.message", "業務エラーを画面表示用 error に変換する。"],
                [7, "return redirect:/reservations", "処理後に予約画面へ戻る。"],
                [8, "updatePayment(id, paymentStatus)", "支払状態更新 Service を実行する。"],
                [9, "cancel(id)", "予約取消 Service を実行する。"],
                [10, "updateStatus(id, reservationStatus)", "予約状態更新 Service を実行する。"],
                [11, "updateCleaning(id, cleaningStatus)", "清掃状態更新 Service を実行する。"],
            ],
        ),
        (
            "RoomController.java",
            "`RoomController` は MVC 互換の客室画面入口。GET は React shell を返し、POST は客室登録、状態更新、論理削除を Service へ委譲する。",
            [
                [1, "rooms() -> 'app'", "客室管理画面 URL で React shell を返す。"],
                [2, "create(room) -> roomService.create(room)", "客室登録または削除済み客室の再有効化を実行する。"],
                [3, "success -> flash.message = '部屋を登録しました。'", "登録成功メッセージを画面へ渡す。"],
                [4, "catch IllegalArgumentException -> flash.error = ex.message", "業務エラーを画面表示用 error に変換する。"],
                [5, "catch RuntimeException -> flash.error = generic message", "想定外登録失敗を汎用エラーに変換する。"],
                [6, "updateStatuses(id, occupancyStatus, cleaningStatus)", "宿泊状態と清掃状態を同時更新する。"],
                [7, "delete(id) -> roomService.delete(id)", "客室を論理削除する。"],
                [8, "return redirect:/rooms", "処理後に客室画面へ戻る。"],
            ],
        ),
        (
            "PriceRuleController.java",
            "`PriceRuleController` は MVC 互換の料金管理画面入口。料金ルール登録、旧 URL 互換を含む単一削除、複数選択削除を Service へ委譲する。",
            [
                [1, "prices() -> 'app'", "料金管理画面 URL で React shell を返す。"],
                [2, "create(rule) -> priceRuleService.create(rule)", "料金ルール登録 Service を実行する。"],
                [3, "success -> flash.message = '料金ルールを登録しました。'", "登録成功メッセージを画面へ渡す。"],
                [4, "catch IllegalArgumentException -> flash.error = ex.message", "業務エラーを画面表示用 error に変換する。"],
                [5, "@PostMapping({ '/prices/{id}', '/prices/{id}/delete' })", "旧削除 URL と現行削除 URL の両方を受け付ける。"],
                [6, "delete(id) -> priceRuleService.delete(id)", "料金ルールを1件削除する。"],
                [7, "deleteSelected(ids) -> priceRuleService.deleteByIds(ids)", "選択された料金ルールをまとめて削除する。"],
                [8, "return redirect:/prices", "処理後に料金管理画面へ戻る。"],
            ],
        ),
        (
            "ReservationServiceTest.java",
            "`ReservationServiceTest` は予約業務の単体テスト。端末ローカル日付変更による不正な過去日予約を防ぐため、業務日付は `ReservationMapper.currentDate()` 由来のシステム日付として扱うことを検証する。各テストケースは `test_XX` の `@DisplayName` と Javadoc 形式のテスト注解を持つ。",
            [
                [1, "@LoggedTest + @DisplayName('予約サービス')", "共通テストログ出力を有効化し、テストクラス名をコンソール表示用に設定する。"],
                [2, "setUp() -> lenient().when(reservationMapper.currentDate()).thenReturn(LocalDate.now())", "既存テストの既定業務日付を Mapper mock から返す。"],
                [3, "test_01 currentDateUsesSystemDateFromMapperInsteadOfLocalEnvironmentDate", "JVM default timezone を `Pacific/Pago_Pago` に変更しても、Service は `reservationMapper.currentDate()` の固定値を返すことを検証する。"],
                [4, "systemDate = LocalDate.of(2099, 1, 1)", "ローカル環境日付と衝突しない明示的なシステム日付をテスト入力にする。"],
                [5, "assert actual == systemDate", "業務日付が Mapper / DB 基準で取得されることを確認する。"],
                [6, "assert actual != LocalDate.now()", "端末ローカル環境日付を直接使用していないことを確認する。"],
                [7, "finally -> TimeZone.setDefault(originalTimeZone)", "テスト後に JVM の default timezone を元に戻し、他テストへの影響を防ぐ。"],
                [8, "test_02 createRejectsCheckInDateBeforeSystemDateEvenWhenLocalEnvironmentDateIsEarlier", "ローカル日付上は未来でも、システム日付より前のチェックイン日を拒否することを検証する。"],
                [9, "systemDate = localEnvironmentDate.plusDays(30)", "端末日付を過去方向に変更した想定を作る。"],
                [10, "checkInDate = localEnvironmentDate.plusDays(1)", "端末日付基準では未来、システム日付基準では過去となる予約入力を作る。"],
                [11, "System.out.print(systemInput.currentDate / localEnvironmentDate / input.checkInDate / input.checkOutDate)", "テスト証跡として判定に使ったシステム入力値と予約入力値を文字列出力する。"],
                [12, "assertThatThrownBy(create(...)) -> MESSAGE_PAST_CHECK_IN", "システム日付基準で過去日予約を拒否する。"],
                [13, "verify(roomMapper, never()).findByIdForUpdate(any())", "日付検証で拒否された場合、客室取得・予約登録へ進まないことを確認する。"],
                [14, "Javadoc block = テストケース名 / テスト条件 / テスト要望 / テスト結果", "各テストケース上に指定フォーマットの注解を記載し、仕様上の意図をコードから確認できるようにする。"],
                [15, "@DisplayName('test_XX ...')", "Maven / VS Code コンソールでテストケースを番号で識別可能にする。"],
            ],
        ),
        (
            "TestResultLogger.java / LoggedTest.java",
            "`TestResultLogger` は全テストクラス共通の JUnit5 Extension。`@LoggedTest` から適用され、テスト開始、テスト内容、テストコード注解、テスト文字列結果、クラス集計結果を `System.out.print` で標準出力に残す。",
            [
                [1, "LoggedTest -> @ExtendWith(TestResultLogger.class)", "テストクラスへ共通ログ出力機能を一括適用する。"],
                [2, "beforeAll(context)", "テストクラス開始時に `【テストクラス開始】` と表示名を出力する。"],
                [3, "beforeTestExecution(context)", "各テスト開始時に `テスト開始`、`テスト内容`、`テストコード注解` を出力する。"],
                [4, "testContent(context)", "メソッドの `@DisplayName` があればそれをテスト内容として使用し、なければメソッド名を分かち書きにする。"],
                [5, "testCodeNote(context)", "メソッド名の prefix / keyword から Given / When / Then の処理概要を生成する。"],
                [6, "currentDate or SystemDate", "システム日付とローカル環境日付を分けて準備し、システム日付基準で判定するテストとして注解する。"],
                [7, "create*", "登録入力と関連 mock / DB データを準備し、登録処理と保存結果または業務エラーを検証するテストとして注解する。"],
                [8, "update*", "更新対象データと更新値を準備し、更新後の状態または業務エラーを検証するテストとして注解する。"],
                [9, "delete* / cancel*", "削除・取消対象データを準備し、状態変更、遷移先、メッセージを検証するテストとして注解する。"],
                [10, "find* / query* / Returns", "検索条件と期待データを準備し、取得結果を期待値と比較するテストとして注解する。"],
                [11, "testSuccessful(context)", "成功件数を集計し、`テスト成功` と `テスト文字列結果：... / 結果=成功` を出力する。"],
                [12, "testFailed(context, cause)", "失敗件数を集計し、失敗理由を `テスト文字列結果` に含めて出力する。"],
                [13, "testDisabled(context, reason)", "スキップ件数を集計し、理由を文字列結果として出力する。"],
                [14, "afterAll(context)", "クラス単位で合計、成功、失敗、中止、スキップ件数を `【テストクラス文字列結果】` として出力する。"],
            ],
        ),
        (
            "Domain / Mapper Java",
            "`domain` は DB と画面/API の間で使うデータ保持クラス、`mapper` は MyBatis XML の SQL 呼び出し口。業務判断は Service に置き、Domain は値保持、Mapper は永続化操作に責務を限定する。",
            [
                [1, "Reservation = id + reservationNo + roomId + guest + status + totalAmount", "予約本体、代表者、状態、金額、一覧表示用項目を保持する。"],
                [2, "Reservation.getReservationStatusLabel()", "予約状態コードを画面表示用ラベルへ変換する。"],
                [3, "Reservation.getPaymentStatusLabel()", "支払状態コードを画面表示用ラベルへ変換する。"],
                [4, "ReservationGuest = reservationId + guestName + guestKana + guestGender + guestAge + guestPhone", "予約に紐づく同行者明細を保持する。"],
                [5, "Room = roomNumber + capacity + basePrice + occupancyStatus + cleaningStatus + active", "客室マスタ、販売条件、予約可否状態を保持する。"],
                [6, "RoomPriceRule = roomId + startDate + endDate + price + priority + active", "客室別・期間別料金ルールを保持する。"],
                [7, "ReservationMapper.currentDate()", "DB 基準の現在日付を取得する。"],
                [8, "ReservationMapper.nextReservationSequence()", "予約番号用の DB 連番を採番し、Service 側で R + 6桁へ整形する。"],
                [9, "ReservationMapper.countOverlapping(roomId, checkIn, checkOut)", "同一客室・期間交差する予約数を取得する。"],
                [10, "ReservationMapper.countOtherBookedByRoomId(roomId, excludedReservationId)", "取消・チェックアウト対象以外の同室予約中件数を取得する。"],
                [11, "ReservationMapper.deleteCancelled(id)", "reservation_status=cancelled の予約だけ削除する。"],
                [12, "RoomMapper.findByIdForUpdate(id)", "予約登録時に対象客室をロック付きで取得する。"],
                [13, "RoomMapper.restore(id)", "論理削除済み客室を有効状態へ戻す。"],
                [14, "RoomMapper.deletePermanently(id)", "予約履歴のない論理削除済み客室を完全削除する。"],
                [15, "RoomMapper.findBookable()", "`active = true` + `vacant` + `cleaned` の予約可能客室を取得する。"],
                [16, "RoomPriceRuleMapper.findBestRule(roomId, stayDate)", "対象日の最優先料金ルールを取得する。"],
            ],
        ),
        (
            "Reservation deleteCheckedOut / API / SQL",
            "チェックアウト済み予約の完全削除は、API 入口、Service 業務検証、Mapper SQL を合わせて追跡する。",
            [
                [1, "ApiController.deleteCheckedOutReservation(id)", "POST /api/reservations/{id}/delete-checked-out からチェックアウト済み予約削除 API を受け付ける。"],
                [2, "ReservationService.deleteCheckedOut(id)", "checked_out 予約だけを削除し、それ以外は拒否する。"],
                [3, "ReservationMapper.deleteCheckedOut(id)", "reservations から `reservation_status='checked_out'` の行だけを削除する。"],
                [4, "SQL-002C / API-014A / UT-037", "式样上の関連付けをこの処理へ集約する。"],
            ],
        ),
    ]

    for file_name, summary, rows in sections:
        add_code_section(ws, file_name, summary, rows)


def finalize(wb: Workbook):
    width_profiles = {
        "表紙": {"A": 18, "B": 46, "C": 6, "D": 24, "E": 46},
        "変更履歴": {"A": 10, "B": 16, "C": 72, "D": 18, "E": 18},
        "概要": {"A": 12, "B": 28, "C": 86, "D": 46},
        "業務要件": {"A": 16, "B": 20, "C": 28, "D": 72, "E": 12, "F": 36},
        "機能一覧": {"A": 16, "B": 18, "C": 26, "D": 72, "E": 34, "F": 34, "G": 12, "H": 12, "I": 40},
        "コード行別処理仕様": {"A": 10, "B": 72, "C": 92, "D": 18},
        "画面一覧": {"A": 14, "B": 28, "C": 22, "D": 62, "E": 52, "F": 24, "G": 36},
        "DB設計": {"A": 24, "B": 34, "C": 24, "D": 10, "E": 24, "F": 12, "G": 72},
        "API一覧": {"A": 16, "B": 18, "C": 12, "D": 42, "E": 58, "F": 34, "G": 34, "H": 12},
        "非機能要件": {"A": 18, "B": 18, "C": 42, "D": 86, "E": 12},
        "テスト仕様": {"A": 16, "B": 34, "C": 42, "D": 58, "E": 46, "F": 46, "G": 58, "H": 46, "I": 34, "J": 14, "K": 24},
        "スケジュール": {"A": 20, "B": 16, "C": 16, "D": 12, "E": 18, "F": 42, "G": 14},
        "参考資料": {"A": 12, "B": 52, "C": 74, "D": 72},
    }

    for ws in wb.worksheets:
        profile = width_profiles.get(ws.title, {})
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = profile.get(letter, 22)

        ws.freeze_panes = "A6"
        ws.sheet_view.showGridLines = False
        ws.auto_filter.ref = ws.dimensions
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        for row in range(1, ws.max_row + 1):
            max_len = max(text_len(ws.cell(row, col).value) for col in range(1, ws.max_column + 1))
            if row <= 5:
                ws.row_dimensions[row].height = 22 if row != 5 else 10
            elif ws.title == "コード行別処理仕様":
                ws.row_dimensions[row].height = min(96, max(24, 18 + (max_len // 42) * 15))
            else:
                ws.row_dimensions[row].height = min(84, max(24, 18 + (max_len // 48) * 12))

        if ws.title == "コード行別処理仕様":
            ws.freeze_panes = "A7"
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 76
            ws.column_dimensions["C"].width = 96
            ws.column_dimensions["D"].width = 14


def main():
    wb = Workbook()
    wb.properties.creator = AUTHOR
    wb.properties.lastModifiedBy = UPDATED_BY
    wb.properties.created = CREATED_AT
    wb.properties.modified = UPDATED_AT

    sheet_writers = [
        ("表紙", write_cover),
        ("変更履歴", write_history),
        ("概要", write_overview),
        ("業務要件", write_business_requirements),
        ("機能一覧", write_functions),
        ("コード行別処理仕様", write_code_trace),
        ("画面一覧", write_screens),
        ("DB設計", write_db),
        ("API一覧", write_api),
        ("非機能要件", write_non_functional),
        ("テスト仕様", write_tests),
        ("スケジュール", write_schedule),
        ("参考資料", write_references),
    ]

    while len(wb.worksheets) < len(sheet_writers):
        wb.create_sheet()
    while len(wb.worksheets) > len(sheet_writers):
        del wb[wb.worksheets[-1].title]

    for ws, (title, writer) in zip(wb.worksheets, sheet_writers):
        ws.title = title
        writer(ws)

    finalize(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    if OUT.exists():
        OUT.unlink()
    wb.save(OUT)


if __name__ == "__main__":
    main()
