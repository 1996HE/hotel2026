from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/heniantong/Documents/codex/hotel-management")
OUT = ROOT / "docs/仕様書/民宿管理システム_フロントエンド設計書.xlsx"

CREATED_AT = datetime(2026, 7, 1, 0, 0)
UPDATED_AT = datetime(2026, 7, 8, 0, 0)
CREATED_TEXT = "2026-07-01"
UPDATED_TEXT = "2026-07-08"
AUTHOR = "何念童"
UPDATED_BY = "何念童"

thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
fill_title = PatternFill("solid", fgColor="D9EAF7")
fill_header = PatternFill("solid", fgColor="EAF3F8")
fill_body = PatternFill("solid", fgColor="FFFFFF")
fill_note = PatternFill("solid", fgColor="F7F9FC")
font_title = Font(name="Meiryo", size=12, bold=True)
font_header = Font(name="Meiryo", size=10, bold=True)
font_body = Font(name="Meiryo", size=10)
font_large = Font(name="Meiryo", size=16, bold=True)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)


def set_cell(cell, value, *, font=font_body, fill=fill_body, alignment=align_left):
    cell.value = value
    cell.font = font
    cell.fill = fill
    cell.border = border
    cell.alignment = alignment


def setup_sheet(ws, title: str, max_col: int):
    ws.sheet_view.zoomScale = 90
    for col_idx in range(1, max(max_col, 5) + 1):
        letter = ws.cell(1, col_idx).column_letter
        ws.column_dimensions[letter].width = 28 if col_idx == 2 else 18

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    set_cell(ws["A1"], f"{title}　作成・更新情報", font=font_title, fill=fill_title, alignment=align_center)

    for idx, value in enumerate(["作成日時", "作成者", "更新日時", "更新者"], start=1):
        set_cell(ws.cell(2, idx), value, font=font_header, fill=fill_header, alignment=align_center)
    for idx, value in enumerate([CREATED_TEXT, AUTHOR, UPDATED_TEXT, UPDATED_BY], start=1):
        set_cell(ws.cell(3, idx), value, fill=fill_note)

    set_cell(ws["A4"], "備考", font=font_header, fill=fill_header, alignment=align_center)
    set_cell(ws["B4"], "React フロントエンド実装を基準に更新", fill=fill_note)
    set_cell(ws["C4"], "対象シート", font=font_header, fill=fill_header, alignment=align_center)
    set_cell(ws["D4"], title, fill=fill_note)
    for row in range(1, 5):
        ws.row_dimensions[row].height = 22
    ws.row_dimensions[5].height = 10


def add_table(ws, start_row: int, headers: list[str], rows: list[list[object]]):
    for col, header in enumerate(headers, start=1):
        set_cell(ws.cell(start_row, col), header, font=font_header, fill=fill_header, alignment=align_center)
    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c_idx, value in enumerate(row, start=1):
            set_cell(ws.cell(r_idx, c_idx), value)
        ws.row_dimensions[r_idx].height = 36


def add_table_after(ws, headers: list[str], rows: list[list[object]], *, gap: int = 2):
    add_table(ws, ws.max_row + gap, headers, rows)


def text_len(value: object) -> int:
    if value is None:
        return 0
    return len(str(value))


def write_cover(ws):
    setup_sheet(ws, "表紙", 5)
    ws.merge_cells("A6:E6")
    set_cell(ws["A6"], "民宿管理システム フロントエンド設計書", font=font_large, alignment=align_center)
    rows = [
        ["文書名", "民宿管理システム フロントエンド設計書", "", "仕様書構成", ""],
        ["プロジェクト名", "民宿管理システム開発", "", "シート", "内容"],
        ["対象範囲", "React 画面、API 呼び出し、状態管理、CSS 表示設計", "", "概要", "目的・対象範囲"],
        ["作成日", CREATED_TEXT, "", "画面一覧", "画面構成"],
        ["版数", "1.0", "", "機能一覧", "画面機能・コンポーネント"],
        ["詳細設計", "コード行別処理仕様", "", "コード行別処理仕様", "App.jsx / CSS / HTML の処理追跡表"],
        ["更新日", UPDATED_TEXT, "", "API一覧", "フロントから呼び出す API"],
        ["作成者", AUTHOR, "", "テスト仕様", "フロント単体確認観点"],
        ["更新者", UPDATED_BY, "", "参考資料", "関連コード"],
    ]
    for r_idx, row in enumerate(rows, start=8):
        for c_idx, value in enumerate(row, start=1):
            set_cell(ws.cell(r_idx, c_idx), value, font=font_header if c_idx in (1, 4) else font_body,
                     fill=fill_header if c_idx in (1, 4) else fill_body)
    add_table_after(ws, ["優先順", "工程区分", "対象シート", "記載目的", "確認観点"], [
        [1, "文書管理", "表紙 / 変更履歴", "文書の版、作成者、更新者、変更内容を確認する。", "提出物としての識別性があること。"],
        [2, "全体方針", "概要", "フロントエンドの目的、対象範囲、画面方式、API 方針を確認する。", "画面全体像が先に把握できること。"],
        [3, "要件定義", "業務要件", "利用者が画面に求める操作、表示、エラー処理を確認する。", "要件から画面・API・テストへ追跡できること。"],
        [4, "基本設計", "機能一覧 / 画面一覧", "画面機能、コンポーネント、画面遷移を確認する。", "画面機能と React 実装の対応が分かること。"],
        [5, "詳細設計", "コード行別処理仕様", "App.jsx / app.css / app.html の実装順に処理内容を確認する。", "コード要素と仕様行が対応していること。"],
        [6, "外部・データ設計", "API一覧 / DB設計", "フロント入出力、API、画面データ、状態値を確認する。", "画面が扱うデータ構造が分かること。"],
        [7, "品質設計", "非機能要件 / テスト仕様", "表示崩れ防止、context-path、CSRF、テスト観点を確認する。", "品質要件と検証観点が揃っていること。"],
        [8, "管理情報", "スケジュール / 参考資料", "工程、成果物、参照コードを確認する。", "提出・保守時の参照先が明確であること。"],
    ])


def write_history(ws):
    setup_sheet(ws, "変更履歴", 5)
    add_table(ws, 6, ["版数", "日付", "変更内容", "作成者", "承認者"], [
        ["1.0", CREATED_TEXT, "フロントエンド設計書初版作成", AUTHOR, ""],
        ["1.1", "2026-07-06", "React 化後の画面・API・CSS・テーブル表示仕様を反映", UPDATED_BY, ""],
        ["1.2", UPDATED_TEXT, "後端仕様書と同じ優先順・表格幅・コード行別処理仕様へ整理", UPDATED_BY, ""],
        ["1.3", UPDATED_TEXT, "予約日付 min は API の today を使用し、テストケース番号・注解・System.out.print 文字列結果の運用を反映", UPDATED_BY, ""],
        ["1.4", UPDATED_TEXT, "客室復元・完全削除、取消済み予約削除、一覧ページング・補完行、CSS レイアウト刷新を反映", UPDATED_BY, ""],
        ["1.5", UPDATED_TEXT, "チェックアウト済み予約削除、予約一覧 API、予約画面ボタン・テスト観点を反映", UPDATED_BY, ""],
        ["1.6", UPDATED_TEXT, "app.css 後半のテーブル列幅、操作ボタン、通知、同行者、レスポンシブ表示仕様を追記", UPDATED_BY, ""],
    ])


def write_overview(ws):
    setup_sheet(ws, "概要", 4)
    add_table(ws, 6, ["項目", "内容", "", ""], [
        ["目的", "民宿管理システムの React フロントエンドについて、画面構成、コンポーネント、API 連携、状態管理、CSS 方針を明確化する。", "", ""],
        ["対象ファイル", "src/main/frontend/App.jsx、src/main/resources/static/styles/app.css、src/main/resources/templates/app.html", "", ""],
        ["技術", "React、ReactDOM createRoot、Fetch API、CSS、esbuild", "", ""],
        ["画面方式", "Spring Boot が app.html を返し、React が URL パスに応じて Dashboard / Rooms / Reservations / Prices を描画する。", "", ""],
        ["API方式", "フロント共通 api() 関数で context-path、JSON header、CSRF header、エラー処理を統一する。", "", ""],
        ["設計方針", "社内業務画面として、表の視認性、列幅固定、横スクロール、フォーム操作性を重視する。", "", ""],
    ])


def write_business_requirements(ws):
    setup_sheet(ws, "業務要件", 6)
    add_table(ws, 6, ["要件ID", "分類", "要件名", "要件内容", "優先度", "備考"], [
        ["FR-001", "共通", "ナビゲーション", "ダッシュボード、部屋、予約、料金へ移動でき、現在画面を active 表示する。", "高", "Nav"],
        ["FR-002", "共通", "APIエラー表示", "API エラーは Notice に赤色メッセージとして表示する。", "高", "Notice / api"],
        ["FR-003", "共通", "CSRF送信", "POST 系 API では meta タグから取得した CSRF token を header に付与する。", "高", "api"],
        ["FR-004", "予約", "予約登録フォーム", "客室、日付、宿泊者、連絡先、同行者、支払状態を入力できる。", "高", "Reservations"],
        ["FR-005", "予約", "同行者入力", "人数が2名以上の場合、人数-1 名分の同行者入力カードを表示する。", "高", "companions useMemo"],
        ["FR-006", "予約", "予約一覧", "予約、取消、チェックアウトをそれぞれ一覧表示する。", "高", "ReservationTable / CheckoutTable"],
        ["FR-007", "予約", "同行者表示", "複数行の同行者情報を行単位のリストとして表示する。", "中", "CompanionSummary"],
        ["FR-008", "客室", "部屋登録", "部屋番号、名称、種別、定員、基本料金、専用バス、メモを登録できる。", "高", "Rooms"],
        ["FR-009", "客室", "状態更新", "宿泊状態と清掃状態を同じ行で更新できる。", "高", "RoomRow"],
        ["FR-010", "料金", "料金登録", "客室、ルール名、期間、料金、優先度、メモを登録できる。", "高", "Prices"],
        ["FR-011", "料金", "一括削除", "料金ルールをチェックボックスで複数選択し削除できる。", "中", "ids state"],
    ])


def write_functions(ws):
    setup_sheet(ws, "機能一覧", 9)
    add_table(ws, 6, ["機能ID", "大分類", "機能名", "概要", "入力", "出力", "権限", "優先度", "関連コンポーネント"], [
        ["FE-001", "共通", "context-path 解決", "script src から /jukai-internal を推定し、API とリンクに付与する。", "location/script", "URL", "社内", "高", "contextPath / withContext"],
        ["FE-002", "共通", "ルーティング", "現在 URL に応じて Dashboard / Rooms / Reservations / Prices を切り替える。", "pathname", "React view", "社内", "高", "currentRoute / App"],
        ["FE-003", "共通", "API 通信", "JSON header、CSRF header、エラー処理を共通化する。", "path/options", "JSON", "社内", "高", "api"],
        ["FE-004", "共通", "タグ表示", "状態値に応じた色付きラベルを表示する。", "status", "span.tag", "社内", "中", "Tag"],
        ["FE-005", "Dashboard", "集計表示", "部屋数、空室、有効予約、直近予約一覧を表示する。", "GET /api/dashboard", "Dashboard", "社内", "高", "Dashboard / Metric"],
        ["FE-006", "Rooms", "客室登録", "フォーム入力を POST /api/rooms へ送信し、登録後再読込する。", "Room form", "Notice", "社内", "高", "Rooms"],
        ["FE-007", "Rooms", "客室状態更新", "行内 select の値を POST /api/rooms/{id}/statuses へ送信する。", "status pair", "Notice", "社内", "高", "RoomRow"],
        ["FE-007A", "Rooms", "削除済み客室管理", "削除済み一覧から復元または完全削除を実行し、結果を Notice に表示する。", "room id", "Notice", "社内", "高", "Rooms / runAction"],
        ["FE-008", "Prices", "料金登録", "料金ルールフォームを POST /api/prices へ送信する。", "Price form", "Notice", "社内", "高", "Prices"],
        ["FE-009", "Prices", "料金選択削除", "チェック済み ID を POST /api/prices/delete-selected へ送信する。", "ids", "Notice", "社内", "中", "Prices"],
        ["FE-010", "Reservations", "予約登録", "予約本体と同行者配列を POST /api/reservations へ送信する。", "ReservationCreateRequest", "Notice", "社内", "高", "Reservations"],
        ["FE-011", "Reservations", "支払更新", "行内 select の支払状態を POST /api/reservations/{id}/payment へ送信する。", "paymentStatus", "Notice", "社内", "中", "ReservationRow"],
        ["FE-012", "Reservations", "清掃更新", "チェックアウト行の清掃状態を POST /api/reservations/{id}/cleaning へ送信する。", "cleaningStatus", "Notice", "社内", "中", "CheckoutRow"],
        ["FE-012A", "Reservations", "取消済み予約削除", "取消済み一覧から取消済み予約のみ削除 API を実行する。", "reservation id", "Notice", "社内", "中", "ReservationTable"],
        ["FE-012B", "Reservations", "チェックアウト済み予約削除", "チェックアウト済み一覧からチェックアウト済み予約のみ削除 API を実行する。", "reservation id", "Notice", "社内", "中", "CheckoutTable"],
        ["FE-015", "共通", "一覧ページング", "客室、削除済み客室、料金一覧を 5 件単位に分割し、補完行で表高さを安定させる。", "items/page", "Pager/table rows", "社内", "中", "PAGE_SIZE / slicePage / padPage"],
    ])


def write_screens(ws):
    setup_sheet(ws, "画面一覧", 7)
    add_table(ws, 6, ["画面ID", "画面名", "利用者", "概要", "主な項目", "遷移元", "遷移先"], [
        ["FS-001", "ダッシュボード", "社内担当者", "予約状況と集計を確認する初期画面。", "メトリクス、予約一覧、予約登録リンク", "-", "予約、部屋、料金"],
        ["FS-002", "部屋", "施設管理担当者", "部屋登録、部屋一覧、削除済み部屋一覧を管理する。", "登録フォーム、状態更新、削除ボタン", "Nav", "同画面"],
        ["FS-003", "予約", "予約管理担当者", "予約登録、予約一覧、取消一覧、チェックアウト一覧を管理する。", "予約フォーム、同行者カード、支払更新、取消、清掃更新", "Nav", "同画面"],
        ["FS-004", "料金", "料金管理担当者", "料金ルール登録と一覧削除を管理する。", "料金フォーム、選択削除、削除ボタン", "Nav", "同画面"],
    ])


def add_code_section(ws, file_name: str, summary: str, rows: list[list[object]]):
    start_row = ws.max_row + 2
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
    set_cell(ws.cell(start_row, 1), file_name, font=font_title, fill=fill_title, alignment=align_left)

    set_cell(ws.cell(start_row + 1, 1), "処理内容", font=font_header, fill=fill_header, alignment=align_center)
    ws.merge_cells(start_row=start_row + 1, start_column=2, end_row=start_row + 1, end_column=4)
    set_cell(ws.cell(start_row + 1, 2), summary, fill=fill_note)

    add_table(ws, start_row + 3, ["No", "コード / 要素", "入出力・状態変化"], rows)
    ws.row_dimensions[start_row].height = 26
    ws.row_dimensions[start_row + 1].height = 48


def write_code_trace(ws):
    setup_sheet(ws, "コード行別処理仕様", 4)
    add_table(ws, 6, ["記法", "意味", "例", ""], [
        ["name = value", "定数・state・初期値を示す。", "paymentStatus = unpaid", ""],
        ["method(args)", "関数・イベントハンドラ呼び出しを示す。", "api('/api/rooms')", ""],
        ["if (condition) -> result", "条件分岐と結果を示す。", "if (!response.ok) -> throw Error", ""],
        ["A -> B -> C", "処理順序または状態遷移を示す。", "submit -> api -> setNotice -> load", ""],
        ["class = value", "CSS class または React className を示す。", "class = table-scroll", ""],
    ])

    sections = [
        (
            "App.jsx / 共通初期化",
            "`App.jsx` の共通部は React 起動、CSRF、context-path、API 通信、表示ラベル変換を担当する。`/jukai-internal` 配下でも同じコードで API と画面リンクが動くようにする。",
            [
                [1, "import = React(useEffect, useMemo, useState)", "React component と state / effect / memo を使用可能にする。"],
                [2, "import = createRoot", "React を `#root` DOM に mount する API を使用可能にする。"],
                [3, "csrfToken = meta[name='_csrf']?.content", "HTML meta から CSRF token を取得する。"],
                [4, "csrfHeader = meta[name='_csrf_header']?.content || 'X-CSRF-TOKEN'", "CSRF header 名を meta から取得し、未取得時は既定値を使う。"],
                [5, "appScript = script[src$='/js/app.js']", "bundle script の URL から context-path 推定に使う src を取得する。"],
                [6, "contextPath = appScript - '/js/app.js'", "`/jukai-internal` など配備 path を算出する。"],
                [7, "withContext(path) = contextPath + path", "画面リンクと API URL に context-path を付与する。"],
                [8, "currentRoute() -> pathname - contextPath", "現在 URL から React 内部 route を取得する。"],
                [9, "if (route == '/' || route == '') -> '/dashboard'", "ルートアクセス時はダッシュボード表示に寄せる。"],
                [10, "api(path, options = {})", "Fetch API の共通 wrapper として API 通信を統一する。"],
                [11, "headers = Accept + options.headers", "全 API に JSON 受信 header を付与する。"],
                [12, "if (options.body && body is not FormData) -> Content-Type = application/json", "JSON POST 時だけ Content-Type を設定する。"],
                [13, "if (csrfToken && method != GET) -> headers[csrfHeader] = csrfToken", "POST 系 API に CSRF token を付与する。"],
                [14, "fetch(withContext(path), { ...options, headers })", "context-path 付き URL へ HTTP request を送信する。"],
                [15, "data = response.json().catch(() => ({}))", "JSON 解析失敗時も空 object として扱い、後続処理を安定させる。"],
                [16, "if (!response.ok) -> throw Error(data.error || default)", "API エラーを画面表示しやすい Error に変換する。"],
                [17, "yen(value) = `¥${value ?? 0}`", "金額表示を円表記へ整形する。"],
                [18, "text(value, fallback)", "NULL / undefined / 空文字を fallback 表示へ変換する。"],
                [19, "roomTypeLabel / occupancyLabel / cleaningLabel", "内部コードを日本語ラベルへ変換する。"],
                [20, "PAGE_SIZE = 5", "客室・料金などの一覧を 1 ページ 5 件で表示する基準値を定義する。"],
                [21, "totalPages / slicePage / padPage", "ページ数計算、表示範囲抽出、空行補完を共通処理として定義する。"],
            ],
        ),
        (
            "App.jsx / 共通コンポーネント",
            "共通コンポーネントはナビゲーション、状態タグ、通知、メトリクス表示を担当する。画面ごとの業務処理から表示部品を分離し、各画面で再利用する。",
            [
                [1, "Nav({ route })", "現在 route を受け取り、画面メニューを表示する。"],
                [2, "items = { dashboard, rooms, reservations, prices }", "ナビゲーション対象の URL とラベルを定義する。"],
                [3, "href = withContext(href)", "context-path 付きリンクを生成する。"],
                [4, "className = route == href ? active : ''", "現在画面のナビ項目を active 表示する。"],
                [5, "Tag({ children, status })", "`tag ${status}` class で状態ラベルを表示する。"],
                [6, "Notice({ message, error })", "成功またはエラー通知を表示する。"],
                [7, "if (!message && !error) -> null", "通知がない場合は DOM を出力しない。"],
                [8, "class = notice error/success", "エラー時は赤、成功時は通常通知として表示する。"],
                [9, "Metric({ label, value })", "ダッシュボードの集計カードを表示する。"],
                [10, "Pager({ page, total, onPage })", "一覧の前後ページ移動と現在ページ表示を共通化する。"],
            ],
        ),
        (
            "App.jsx / Dashboard",
            "`Dashboard` は初期画面。`GET /api/dashboard` で集計値と直近予約を取得し、メトリクスと予約一覧を表示する。",
            [
                [1, "data = useState(null)", "ダッシュボード API 結果を保持する。"],
                [2, "useEffect(() -> api('/api/dashboard').then(setData), [])", "初回表示時にダッシュボード API を取得する。"],
                [3, "if (!data) -> 読み込み中", "API 応答前は loading 表示にする。"],
                [4, "Metric(roomCount / vacantCount / bookedCount)", "部屋数、空室、有効予約をカード表示する。"],
                [5, "href = withContext('/reservations')", "予約登録画面への導線を context-path 付きで表示する。"],
                [6, "ReservationTable(data.recentReservations.items, compact = true)", "直近予約をコンパクト一覧で表示する。"],
            ],
        ),
        (
            "App.jsx / Rooms",
            "`Rooms` は客室登録・客室一覧・削除済み一覧を担当する。登録、状態更新、論理削除後は Notice を表示し、一覧を再取得する。",
            [
                [1, "data = { rooms: [], deletedRooms: [] }", "客室画面の一覧データ初期値を定義する。"],
                [2, "form = { roomType: washitsu, capacity: 2, basePricePerPerson: 8800, ... }", "客室登録フォームの初期値を定義する。"],
                [3, "loadData('/api/rooms', setData, setError)", "客室一覧 API を取得して state に反映する。"],
                [4, "useEffect(load, [])", "画面初回表示時に客室一覧を取得する。"],
                [5, "submit(event) -> preventDefault()", "フォーム標準送信を止め、React/API 処理へ切り替える。"],
                [6, "api('/api/rooms', POST, JSON.stringify(form))", "客室登録 API を実行する。"],
                [7, "success -> setNotice(message) -> reset form -> load()", "登録成功後、通知表示、フォーム初期化、一覧再取得を行う。"],
                [8, "catch(error) -> setNotice(error)", "登録失敗時は API error message を表示する。"],
                [9, "updateStatus(room, occupancyStatus, cleaningStatus)", "行内 select の状態値を API へ送信する。"],
                [10, "deleteRoom(room) -> POST /api/rooms/{id}/delete", "客室論理削除 API を実行する。"],
                [11, "restoreRoom(room) -> POST /api/rooms/{id}/restore", "削除済み客室を有効状態へ復元する API を実行する。"],
                [12, "deletePermanently(room) -> POST /api/rooms/{id}/delete-permanently", "削除済みかつ予約履歴なしの客室を完全削除する API を実行する。"],
                [13, "roomPage / deletedRoomPage", "有効客室一覧と削除済み客室一覧の現在ページを別々に保持する。"],
                [14, "visibleRooms / paddedRooms", "5 件単位の表示対象と補完行を生成して表高さを安定させる。"],
                [15, "RoomRow(room)", "1行分の客室表示、状態選択、更新、削除を担当する。"],
                [16, "occupancyStatus = useState(room.occupancyStatus)", "行内編集用に宿泊状態を local state として保持する。"],
                [17, "cleaningStatus = useState(room.cleaningStatus)", "行内編集用に清掃状態を local state として保持する。"],
                [18, "colgroup = room-col-*", "客室テーブルの列幅を CSS class で固定する。"],
            ],
        ),
        (
            "App.jsx / Prices",
            "`Prices` は料金ルール登録、一覧表示、単一削除、複数選択削除を担当する。数値項目は API 送信前に Number へ変換する。",
            [
                [1, "data = { rules: [], rooms: [] }", "料金ルール一覧と客室選択肢の初期値を定義する。"],
                [2, "ids = []", "一括削除対象 ID を保持する。"],
                [3, "form = { priority: 10, active: true }", "料金ルール登録フォームの初期値を定義する。"],
                [4, "loadData('/api/prices', setData, setError)", "料金画面初期データを取得する。"],
                [5, "submit(event) -> api('/api/prices', POST, body)", "料金ルール登録 API を実行する。"],
                [6, "roomId / pricePerPerson / priority = Number(value)", "API 送信前に数値項目を number 化する。"],
                [7, "success -> setNotice(message) -> form reset -> load()", "登録成功後、通知表示、フォーム初期化、一覧再取得を行う。"],
                [8, "remove(id) -> POST /api/prices/{id}/delete", "料金ルールを1件削除する。"],
                [9, "removeSelected() -> POST /api/prices/delete-selected { ids }", "チェック済み ID を一括削除 API へ送信する。"],
                [10, "success -> setIds([]) -> load()", "一括削除成功後、選択状態を解除し一覧を再取得する。"],
                [11, "checkbox checked = ids.includes(rule.id)", "選択済み ID を checkbox 状態に反映する。"],
                [12, "onChange -> ids add/remove", "checkbox 操作で一括削除対象 ID を更新する。"],
                [13, "rulePage / visibleRules / paddedRules", "料金一覧を 5 件単位でページングし、空行補完で高さを揃える。"],
            ],
        ),
        (
            "App.jsx / Reservations",
            "`Reservations` は予約登録、同行者入力、予約一覧、取消一覧、チェックアウト一覧、支払更新、取消、清掃更新を担当する。宿泊人数に応じて同行者配列を生成して API へ送信する。",
            [
                [1, "data = { reservations, cancelledReservations, checkedOutReservations, rooms }", "予約画面で使う一覧データの初期値を定義する。"],
                [2, "form = { guestCount: 1, reservationForm: '公式', paymentStatus: 'unpaid' }", "予約登録フォームの初期値を定義する。"],
                [3, "loadData('/api/reservations', setData, setError)", "予約画面初期データを取得する。"],
                [4, "companions = max(0, Number(guestCount || 1) - 1)", "代表者を除いた同行者入力数を算出する。"],
                [5, "submit(event) -> preventDefault()", "フォーム標準送信を止め、API 登録処理へ切り替える。"],
                [6, "companionNames/Kanas/Genders/Ages/Phones = Array.from({ length: companions })", "同行者入力を API DTO 用の配列へ変換する。"],
                [7, "reservation = { roomId, checkInDate, checkOutDate, guest*, guestCount, paymentStatus, note }", "予約本体 DTO を組み立てる。"],
                [8, "guestPhone = noPhoneInfo ? null : guestPhone", "電話なし選択時は電話番号を NULL として送る。"],
                [9, "guestEmail = noEmailInfo ? null : guestEmail", "メールなし選択時はメールを NULL として送る。"],
                [10, "api('/api/reservations', POST, { reservation, noPhoneInfo, noEmailInfo, companion* })", "予約登録 API を実行する。"],
                [11, "success -> setNotice(message) -> form reset -> load()", "登録成功後、通知表示、フォーム初期化、一覧再取得を行う。"],
                [12, "payment(item, paymentStatus) -> POST /api/reservations/{id}/payment", "予約行の支払状態を更新する。"],
                [13, "cancel(item) -> POST /api/reservations/{id}/cancel", "予約行を取消する。"],
                [14, "deleteCancelled(item) -> POST /api/reservations/{id}/delete", "取消済み予約一覧から取消済み予約を削除する。"],
                [15, "deleteCheckedOut(item) -> POST /api/reservations/{id}/delete-checked-out", "チェックアウト済み予約一覧からチェックアウト済み予約を削除する。"],
                [16, "cleaning(item, cleaningStatus) -> POST /api/reservations/{id}/cleaning", "チェックアウト行の清掃状態を更新する。"],
                [17, "input[type=date] min = data.today", "チェックイン・チェックアウト日付の最小値を業務日付にする。"],
                [18, "if (companions > 0) -> companion-panel", "宿泊人数が2名以上の場合だけ同行者入力欄を表示する。"],
            ],
        ),
        (
            "App.jsx / ReservationTable / CheckoutTable",
            "予約系テーブルは colgroup で列幅を固定し、同行者は newline text を分割してリスト表示する。状態変更操作は行内 select と button で実行する。",
            [
                [1, "ReservationTable({ reservations, onPayment, onCancel, onDelete, readonly, compact })", "予約一覧、取消一覧、ダッシュボード予約一覧を共通 table で表示する。"],
                [2, "class = reservation-table + dashboard-table/reservation-table-full", "表示場所に応じた table class を設定する。"],
                [3, "ReservationColGroup(includeActions)", "予約テーブル列幅を colgroup で固定する。"],
                [4, "if (reservations.length == 0) -> empty row", "予約データがない場合は空表示行を出す。"],
                [5, "ReservationRow -> paymentStatus = useState(item.paymentStatus)", "支払 select の行内編集 state を保持する。"],
                [6, "Tag(status = checked_out/cancelled/booked)", "予約状態コードに応じて表示色を変える。"],
                [7, "if (!readonly && !compact) -> actions", "通常予約一覧は支払更新と取消操作、取消済み一覧は削除操作を表示する。"],
                [8, "CompanionSummary(value)", "同行者要約テキストを表示用コンポーネントに渡す。"],
                [9, "lines = String(value || '').split('\\n').map(trim).filter(Boolean)", "改行区切りの同行者情報を行配列へ変換する。"],
                [10, "if (lines.length == 0) -> '同行者なし'", "同行者がない場合は専用ラベルを表示する。"],
                [11, "CheckoutTable({ reservations, onCleaning })", "チェックアウト一覧専用の列構成で table を表示する。"],
                [12, "CheckoutRow -> cleaningStatus = item.roomCleaningStatus || 'needs_cleaning'", "清掃状態 select の初期値を設定する。"],
                [13, "onCleaning(item, cleaningStatus)", "チェックアウト後の清掃状態更新を呼び出す。"],
            ],
        ),
        (
            "App.jsx / App mount",
            "`App` は現在 route に応じて画面コンポーネントを切り替え、最後に `createRoot(...).render(<App />)` で React を HTML に mount する。",
            [
                [1, "route = currentRoute()", "現在 URL から表示対象 route を取得する。"],
                [2, "<Nav route={route} />", "全画面共通のナビゲーションを表示する。"],
                [3, "route == '/rooms' -> <Rooms />", "客室画面を表示する。"],
                [4, "route == '/reservations' -> <Reservations />", "予約画面を表示する。"],
                [5, "route == '/prices' -> <Prices />", "料金画面を表示する。"],
                [6, "else -> <Dashboard />", "上記以外はダッシュボードを表示する。"],
                [7, "createRoot(document.getElementById('root')).render(<App />)", "React アプリを `#root` に mount する。"],
            ],
        ),
        (
            "app.html",
            "`app.html` は Spring Boot / Thymeleaf が返す React shell。CSRF meta、CSS、React mount 先、bundle script を定義する。",
            [
                [1, "meta charset = UTF-8", "日本語画面を UTF-8 で表示する。"],
                [2, "meta viewport = width=device-width, initial-scale=1.0", "レスポンシブ表示の基準を設定する。"],
                [3, "meta[name='_csrf'] = ${_csrf.token}", "React が読み取る CSRF token を出力する。"],
                [4, "meta[name='_csrf_header'] = ${_csrf.headerName}", "React が使用する CSRF header 名を出力する。"],
                [5, "link stylesheet = @{/styles/app.css}", "画面 CSS を context-path 対応で読み込む。"],
                [6, "div#root", "React の mount 先 DOM を定義する。"],
                [7, "script = @{/js/app.js}", "bundle 済み React アプリを context-path 対応で読み込む。"],
            ],
        ),
        (
            "app.css / 表示設計",
            "`app.css` は社内業務画面向けの密度、表の読みやすさ、横スクロール、状態タグ、フォーム配置を制御する。テーブル列幅を class で固定し、表示崩れを抑止する。",
            [
                [1, ":root = color variables", "色、罫線、背景、影、状態色を CSS 変数として定義する。"],
                [2, "body = font + background + line-height", "業務画面全体の文字と背景を設定する。"],
                [3, ".topbar = sticky + nav background", "上部ナビゲーションを固定表示する。"],
                [4, ".page = min(1400px, calc(100% - 40px))", "画面本体の最大幅と左右余白を制御する。"],
                [5, ".grid-page = 324px + minmax(0, 1fr)", "フォーム列と一覧列の2カラムレイアウトにする。"],
                [6, ".panel = background + border + shadow", "入力フォームや一覧領域を業務画面パネルとして表示する。"],
                [7, ".form / .form-grid / .form-grid-2", "フォーム項目を縦並び・2列配置で整理する。"],
                [8, "input/select/textarea = min-height + focus ring", "入力欄の高さ、余白、フォーカス時の枠線を統一する。"],
                [9, ".field-hint / .field-hint.is-visible", "入力補助メッセージは通常透明にし、エラー時だけ表示する。"],
                [10, "button/.button/.danger", "通常操作と危険操作のボタン色、サイズ、折返し抑止を定義する。"],
                [11, ".table-scroll = overflow-x auto", "列数が多い表を横スクロールで表示し、セル潰れを防ぐ。"],
                [12, ".table = table-layout fixed", "表全体を固定レイアウトにし、列幅指定が安定して効くようにする。"],
                [13, ".table th = sticky header", "横長一覧でも見出し行を読みやすく固定表示する。"],
                [14, ".table-placeholder-row", "ページング時の空行を非表示に近い表示で補完し、表高さの変動を抑える。"],
                [15, ".reservation-table/.dashboard-table/.checkout-table", "予約系テーブルの文字サイズとセル余白を調整し、情報密度を上げる。"],
                [16, ".table-scroll .reservation-table = min-width 1320px", "予約一覧は全列を保持し、画面幅不足時は横スクロールに逃がす。"],
                [17, ".table-scroll .dashboard-table = min-width 1040px", "ダッシュボード予約一覧は同じ情報をより狭い列幅で表示する。"],
                [18, ".table-scroll .checkout-table = min-width 1180px", "チェックアウト一覧は連絡先と同行者列を潰さない幅を確保する。"],
                [19, ".table-scroll .room-table/.deleted-room-table", "通常客室一覧と削除済み客室一覧で必要な最小幅を分ける。"],
                [20, ".price-table = min-width 1120px", "料金一覧の列数に合わせて最小幅とフォーム部品サイズを調整する。"],
                [21, ".col-* reservation widths", "予約番号、部屋、宿泊者、連絡先、同行者、日程、金額、状態、操作の列幅を固定する。"],
                [22, ".room-col-* widths", "部屋番号、部屋名、種別、定員、料金、状態、更新、削除、復元操作の列幅を固定する。"],
                [23, ".price-col-* widths", "料金ルール一覧の選択、部屋、ルール、日付、金額、優先度、状態、メモ、操作列を固定する。"],
                [24, ".dashboard-table .col-* overrides", "ダッシュボードだけ列幅を狭め、初期画面の一覧密度を上げる。"],
                [25, ".tag.is-*", "予約状態、支払状態、客室状態、清掃状態を色付きラベルで表示する。"],
                [26, ".companion-summary", "同行者要約は折返しを許可し、複数人分の情報をセル内で読めるようにする。"],
                [27, ".companion-list / li", "バックエンドの改行区切り同行者情報をカード状の行リストで表示する。"],
                [28, ".companion-meta-item", "同行者のフリガナ、性別、年齢、電話を小さな補助ラベルで並べる。"],
                [29, ".person-summary / .person-meta-item", "代表宿泊者の氏名と補助情報を一覧内で縦配置する。"],
                [30, ".inline-form", "支払状態や清掃状態の select と更新ボタンを行内で横並びにする。"],
                [31, ".reservation-actions", "予約一覧の支払更新と取消ボタンを縦に整理し、操作列内に収める。"],
                [32, ".checkout-actions", "チェックアウト済み予約の清掃更新と削除操作を横並び・折返し可能にする。"],
                [33, ".room-status-form", "客室状態と清掃状態の select、更新ボタンを同じ操作ブロックにまとめる。"],
                [34, ".room-actions", "削除済み客室一覧の復元・完全削除ボタンを横並びにする。"],
                [35, ".pager / .pager-link / .pager-state", "5件ページングの前後ボタンと現在ページ表示を定義する。"],
                [36, ".notice / .notice-toast", "成功・エラー通知を固定位置の toast として表示する。"],
                [37, ".error-panel / .error-copy", "エラー画面の余白、最大幅、説明文色を設定する。"],
                [38, ".empty", "データなし行は中央揃えと十分な余白で表示する。"],
                [39, ".companion-panel", "予約フォーム内の同行者入力エリアを独立した入力ブロックとして表示する。"],
                [40, ".companion-card", "同行者1名分の入力欄をカード状に区切り、氏名・属性・電話をまとめる。"],
                [41, ".companion-grid", "同行者入力欄を2列グリッドにし、狭い画面では1列へ落とす。"],
                [42, ".contact-toggle", "電話なし・メールなしチェックボックスを入力欄の下端に揃える。"],
                [43, "@media (max-width: 980px)", "画面幅が狭い場合、ナビ、メトリクス、フォーム、同行者グリッドを1列にする。"],
            ],
        ),
        (
            "package.json / build",
            "`package.json` は React / ReactDOM / esbuild の依存と build script を定義する。`npm run build` で `App.jsx` を Spring Boot 静的配信用 `app.js` に bundle する。",
            [
                [1, "scripts.build = esbuild src/main/frontend/App.jsx --bundle --minify --format=iife", "React ソースを IIFE 形式で bundle/minify する。"],
                [2, "outfile = src/main/resources/static/js/app.js", "Spring Boot が配信する静的 JS として出力する。"],
                [3, "dependencies = react + react-dom + esbuild", "React 実行と bundle に必要な npm dependency を定義する。"],
            ],
        ),
    ]

    for file_name, summary, rows in sections:
        add_code_section(ws, file_name, summary, rows)


def write_db(ws):
    setup_sheet(ws, "DB設計", 7)
    add_table(ws, 6, ["画面データ", "フロント項目", "型/形式", "表示", "更新", "NULL", "説明"], [
        ["Room", "roomNumber, roomName", "string", "部屋一覧", "登録", "NO", "部屋登録フォームと一覧表示に使用。"],
        ["Room", "roomType", "enum", "和室等ラベル", "登録", "NO", "roomTypeLabel で日本語変換。"],
        ["Room", "occupancyStatus", "enum", "Tag", "状態更新", "NO", "occupancyLabel と Tag 色に使用。"],
        ["Room", "cleaningStatus", "enum", "Tag", "状態更新", "NO", "cleaningLabel と Tag 色に使用。"],
        ["Reservation", "reservationNo", "string", "予約一覧", "なし", "NO", "一覧の主識別子。"],
        ["Reservation", "guestName, guestPhone, guestEmail", "string", "予約一覧", "登録", "一部YES", "代表宿泊者情報。"],
        ["Reservation", "companionSummary", "newline text", "同行者リスト", "なし", "YES", "CompanionSummary で行ごとに表示。"],
        ["Reservation", "paymentStatus", "enum", "Tag/select", "支払更新", "NO", "unpaid / paid。"],
        ["Reservation", "roomCleaningStatus", "enum", "Tag/select", "清掃更新", "YES", "checked_out 行で使用。"],
        ["RoomPriceRule", "ruleName, startDate, endDate, pricePerPerson", "string/date/number", "料金一覧", "登録", "NO", "料金ルールフォームと一覧に使用。"],
    ])


def write_api(ws):
    setup_sheet(ws, "API一覧", 8)
    add_table(ws, 6, ["API ID", "分類", "メソッド", "URL", "概要", "リクエスト", "レスポンス", "呼び出し元"], [
        ["FAPI-001", "Dashboard", "GET", "/api/dashboard", "集計と直近予約取得", "-", "DashboardResponse", "Dashboard.useEffect"],
        ["FAPI-002", "Room", "GET", "/api/rooms", "客室画面初期表示", "-", "RoomsResponse", "Rooms.load"],
        ["FAPI-003", "Room", "POST", "/api/rooms", "客室登録", "form", "MessageResponse", "Rooms.submit"],
        ["FAPI-004", "Room", "POST", "/api/rooms/{id}/statuses", "客室状態更新", "occupancyStatus, cleaningStatus", "MessageResponse", "RoomRow"],
        ["FAPI-005", "Room", "POST", "/api/rooms/{id}/delete", "客室削除", "id", "MessageResponse", "Rooms.deleteRoom"],
        ["FAPI-005A", "Room", "POST", "/api/rooms/{id}/restore", "削除済み客室復元", "id", "MessageResponse", "Rooms.restoreRoom"],
        ["FAPI-005B", "Room", "POST", "/api/rooms/{id}/delete-permanently", "削除済み客室完全削除", "id", "MessageResponse", "Rooms.deletePermanently"],
        ["FAPI-006", "Price", "GET", "/api/prices", "料金画面初期表示", "-", "PricesResponse", "Prices.load"],
        ["FAPI-007", "Price", "POST", "/api/prices", "料金登録", "form", "MessageResponse", "Prices.submit"],
        ["FAPI-008", "Price", "POST", "/api/prices/{id}/delete", "料金単一削除", "id", "MessageResponse", "Prices.remove"],
        ["FAPI-009", "Price", "POST", "/api/prices/delete-selected", "料金一括削除", "ids", "MessageResponse", "Prices.removeSelected"],
        ["FAPI-010", "Reservation", "GET", "/api/reservations", "予約画面初期表示", "-", "ReservationsResponse", "Reservations.load"],
        ["FAPI-011", "Reservation", "POST", "/api/reservations", "予約登録", "reservation + companions", "MessageResponse", "Reservations.submit"],
        ["FAPI-012", "Reservation", "POST", "/api/reservations/{id}/payment", "支払更新", "paymentStatus", "MessageResponse", "ReservationRow"],
        ["FAPI-013", "Reservation", "POST", "/api/reservations/{id}/cancel", "予約取消", "id", "MessageResponse", "ReservationRow"],
        ["FAPI-013A", "Reservation", "POST", "/api/reservations/{id}/delete", "取消済み予約削除", "id", "MessageResponse", "ReservationTable.onDelete"],
        ["FAPI-014A", "Reservation", "POST", "/api/reservations/{id}/delete-checked-out", "チェックアウト済み予約削除", "id", "MessageResponse", "CheckoutTable.onDelete"],
        ["FAPI-014", "Reservation", "POST", "/api/reservations/{id}/cleaning", "清掃更新", "cleaningStatus", "MessageResponse", "CheckoutRow"],
    ])


def write_non_functional(ws):
    setup_sheet(ws, "非機能要件", 5)
    add_table(ws, 6, ["要件ID", "分類", "要件", "基準", "優先度"], [
        ["FNF-001", "表示", "表の列崩れ防止", "予約、客室、料金テーブルは colgroup と min-width で列幅を固定する。", "高"],
        ["FNF-002", "表示", "横スクロール", "幅が足りない場合は .table-scroll で横スクロールし、セルを潰さない。", "高"],
        ["FNF-003", "表示", "同行者可読性", "同行者情報は複数行をリスト表示し、詰まった文章にしない。", "中"],
        ["FNF-004", "操作", "フォーム初期化", "登録成功後はフォームを初期値に戻し、一覧を再読込する。", "高"],
        ["FNF-005", "通信", "CSRF", "POST/DELETE 相当の操作は CSRF header を付与する。", "高"],
        ["FNF-006", "互換", "context-path", "アプリ配備パスが /jukai-internal でも API とリンクが正しく動作する。", "高"],
        ["FNF-007", "ビルド", "esbuild", "App.jsx を static/js/app.js に bundle/minify する。", "中"],
        ["FNF-008", "予約日付", "システム日付基準", "予約フォームの日付 min は端末ローカル日付ではなく /api/reservations の today を使用する。", "高"],
        ["FNF-009", "テスト証跡", "コンソール識別", "Maven / VS Code コンソールでは test_XX 番号、テスト内容、テストコード注解、文字列結果を出力する。", "中"],
        ["FNF-010", "表示", "一覧高さ安定", "5 件ページング対象の表は padPage による補完行でページ移動時の高さ変動を抑える。", "中"],
        ["FNF-011", "操作", "削除制御", "復元、完全削除、取消済み予約削除は API の業務制約結果を Notice へ表示する。", "高"],
    ])


def write_tests(ws):
    setup_sheet(ws, "テスト仕様", 8)
    add_table(ws, 6, ["テストID", "対象機能", "テスト観点", "前提条件", "操作手順", "期待結果", "結果", "備考"], [
        ["FUT-001", "api", "GET 正常", "API 200 応答", "api('/api/dashboard') を実行", "JSON が返る", "未実施", "単体観点"],
        ["FUT-002", "api", "POST CSRF", "meta csrf あり", "POST API を実行", "CSRF header が付与される", "未実施", "単体観点"],
        ["FUT-003", "currentRoute", "context-path", "/jukai-internal/reservations", "currentRoute を実行", "/reservations を返す", "未実施", "単体観点"],
        ["FUT-004", "Reservations", "同行者数", "guestCount=3", "フォーム人数を変更", "同行者カードが2件表示される", "未実施", "単体観点"],
        ["FUT-005", "CompanionSummary", "空表示", "value empty", "コンポーネント表示", "同行者なしを表示", "未実施", "単体観点"],
        ["FUT-006", "ReservationTable", "列幅", "予約一覧データあり", "画面表示", "予約列が colgroup に従う", "未実施", "単体観点"],
        ["FUT-007", "Rooms", "登録成功", "API 成功", "部屋登録フォーム送信", "成功 Notice と一覧再読込", "未実施", "単体観点"],
        ["FUT-008", "Prices", "一括削除", "ids 選択済み", "選択削除を実行", "ids を POST し成功 Notice", "未実施", "単体観点"],
        ["FUT-009", "Reservations", "日付 min", "/api/reservations が today を返す", "予約フォームを表示", "チェックイン・チェックアウト input の min が data.today になる", "仕様反映済", "端末ローカル日付ではなくシステム日付を利用"],
        ["FUT-010", "テスト証跡", "test_XX 表示", "@LoggedTest が有効", "mvn test / VS Code task を実行", "テスト開始・テスト内容・テストコード注解・テスト文字列結果が System.out.print で出力される", "実施済", "後端テスト共通 logger と連動"],
        ["FUT-011", "Rooms", "削除済み客室復元", "削除済み客室データあり", "復元ボタンを実行", "POST /api/rooms/{id}/restore を呼び出し成功 Notice と一覧再読込", "未実施", "API 制約は RoomServiceTest で確認"],
        ["FUT-012", "Rooms", "完全削除", "削除済みかつ予約履歴なしの客室データあり", "完全削除ボタンを実行", "POST /api/rooms/{id}/delete-permanently を呼び出し成功 Notice と一覧再読込", "未実施", "予約履歴ありは API エラー"],
        ["FUT-013", "Reservations", "取消済み予約削除", "取消済み予約データあり", "取消済み一覧の削除ボタンを実行", "POST /api/reservations/{id}/delete を呼び出し成功 Notice と一覧再読込", "未実施", "通常予約は API エラー"],
        ["FUT-014", "Reservations", "チェックアウト済み予約削除", "チェックアウト済み予約データあり", "チェックアウト一覧の削除ボタンを実行", "POST /api/reservations/{id}/delete-checked-out を呼び出し成功 Notice と一覧再読込", "未実施", "通常予約は API エラー"],
    ])


def write_schedule(ws):
    setup_sheet(ws, "スケジュール", 7)
    add_table(ws, 6, ["工程", "開始日", "終了日", "期間(日)", "担当", "成果物", "状態"], [
        ["画面要件整理", "2026-07-01", "2026-07-01", 1, AUTHOR, "画面一覧", "完了"],
        ["React 実装", "2026-07-02", "2026-07-06", 5, AUTHOR, "App.jsx", "完了"],
        ["CSS 調整", "2026-07-02", "2026-07-06", 5, AUTHOR, "app.css", "完了"],
        ["前端設計書", "2026-07-06", "2026-07-06", 1, AUTHOR, "フロントエンド設計書", "完了"],
    ])


def write_references(ws):
    setup_sheet(ws, "参考資料", 4)
    add_table(ws, 6, ["No", "資料名", "内容", "URL"], [
        [1, "日本民宿予約管理システム_仕様書_ヘッダー追加.xlsx", "仕様書フォーマット参考。", "docs/日本民宿予約管理システム_仕様書_ヘッダー追加.xlsx"],
        [2, "App.jsx", "React 画面、コンポーネント、API 呼び出し。", "src/main/frontend/App.jsx"],
        [3, "app.css", "レイアウト、テーブル、フォーム、状態タグの CSS。", "src/main/resources/static/styles/app.css"],
        [4, "app.html", "React マウント先、CSRF meta、bundle 読み込み。", "src/main/resources/templates/app.html"],
        [5, "package.json", "React / ReactDOM / esbuild 依存と build script。", "package.json"],
        [6, "ApiController.reservations", "予約フォームの today を返し、フロントの日付 min に利用する API。", "src/main/java/com/example/minshuku/controller/ApiController.java"],
        [7, "TestResultLogger.java", "Maven / VS Code コンソールへ test_XX、テスト内容、テストコード注解、文字列結果を出力する共通テスト logger。", "src/test/java/com/example/minshuku/support/TestResultLogger.java"],
        [8, "RoomService.java", "客室復元、完全削除、予約中削除制限の業務制約。", "src/main/java/com/example/minshuku/service/RoomService.java"],
        [9, "ReservationService.java", "取消済み予約削除、清掃更新制限、予約番号形式、システム日付取得の業務制約。", "src/main/java/com/example/minshuku/service/ReservationService.java"],
    ])


def finalize(wb: Workbook):
    width_profiles = {
        "表紙": {"A": 18, "B": 48, "C": 6, "D": 26, "E": 52},
        "変更履歴": {"A": 10, "B": 16, "C": 76, "D": 18, "E": 18},
        "概要": {"A": 14, "B": 34, "C": 86, "D": 36},
        "業務要件": {"A": 16, "B": 18, "C": 28, "D": 72, "E": 12, "F": 36},
        "機能一覧": {"A": 16, "B": 18, "C": 28, "D": 72, "E": 36, "F": 30, "G": 12, "H": 12, "I": 42},
        "コード行別処理仕様": {"A": 10, "B": 76, "C": 96, "D": 14},
        "画面一覧": {"A": 14, "B": 28, "C": 24, "D": 70, "E": 62, "F": 24, "G": 28},
        "DB設計": {"A": 24, "B": 36, "C": 24, "D": 22, "E": 18, "F": 12, "G": 72},
        "API一覧": {"A": 16, "B": 18, "C": 12, "D": 44, "E": 58, "F": 38, "G": 34, "H": 34},
        "非機能要件": {"A": 18, "B": 18, "C": 42, "D": 86, "E": 12},
        "テスト仕様": {"A": 16, "B": 34, "C": 40, "D": 46, "E": 46, "F": 58, "G": 14, "H": 28},
        "スケジュール": {"A": 20, "B": 16, "C": 16, "D": 12, "E": 18, "F": 42, "G": 14},
        "参考資料": {"A": 12, "B": 52, "C": 76, "D": 72},
    }
    for ws in wb.worksheets:
        profile = width_profiles.get(ws.title, {})
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = profile.get(letter, 22)

        ws.freeze_panes = "A6"
        ws.sheet_view.showGridLines = False
        ws.auto_filter.ref = ws.dimensions
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        for row in range(1, ws.max_row + 1):
            max_len = max(text_len(ws.cell(row, col).value) for col in range(1, ws.max_column + 1))
            if row <= 5:
                ws.row_dimensions[row].height = 22 if row != 5 else 10
            elif ws.title == "コード行別処理仕様":
                ws.row_dimensions[row].height = min(96, max(24, 18 + (max_len // 42) * 15))
            else:
                ws.row_dimensions[row].height = min(84, max(24, 18 + (max_len // 48) * 12))

        if ws.title == "コード行別処理仕様":
            ws.freeze_panes = "A7"
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 76
            ws.column_dimensions["C"].width = 96
            ws.column_dimensions["D"].width = 14


def main():
    wb = Workbook()
    wb.properties.creator = AUTHOR
    wb.properties.lastModifiedBy = UPDATED_BY
    wb.properties.created = CREATED_AT
    wb.properties.modified = UPDATED_AT
    sheet_writers = [
        ("表紙", write_cover),
        ("変更履歴", write_history),
        ("概要", write_overview),
        ("業務要件", write_business_requirements),
        ("機能一覧", write_functions),
        ("コード行別処理仕様", write_code_trace),
        ("画面一覧", write_screens),
        ("DB設計", write_db),
        ("API一覧", write_api),
        ("非機能要件", write_non_functional),
        ("テスト仕様", write_tests),
        ("スケジュール", write_schedule),
        ("参考資料", write_references),
    ]
    while len(wb.worksheets) < len(sheet_writers):
        wb.create_sheet()
    while len(wb.worksheets) > len(sheet_writers):
        del wb[wb.worksheets[-1].title]
    for ws, (title, writer) in zip(wb.worksheets, sheet_writers):
        ws.title = title
        writer(ws)
    finalize(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    if OUT.exists():
        OUT.unlink()
    wb.save(OUT)


if __name__ == "__main__":
    main()
